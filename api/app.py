from datetime import datetime, timezone, date
from typing import Optional, List, Tuple, Any, Dict

from fastapi import FastAPI, HTTPException, Depends, Header, UploadFile, File, Form
from fastapi.responses import Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field, validator
import psycopg
import os
from dotenv import load_dotenv
from pathlib import Path
import jwt
from passlib.context import CryptContext
import base64
from io import BytesIO
import urllib.request
import urllib.error
import re
import imghdr
import ssl
import threading
import socket
import time
import uuid
import json
import redis
import urllib.request
import urllib.error
from openpyxl import Workbook, load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
load_dotenv(ROOT_DIR / ".env.dev", override=True)

app = FastAPI(title="GameSales API", version="0.1.0")

@app.on_event("startup")
def ensure_analytics_schema():
    try:
        with psycopg.connect(DB_DSN) as conn:
            exec1(conn, "ALTER TABLE app.regions ADD COLUMN IF NOT EXISTS purchase_cost_rate numeric(12,6) NOT NULL DEFAULT 1.0")
            exec1(conn, "ALTER TABLE app.deals ADD COLUMN IF NOT EXISTS completed_at timestamptz")
            exec1(
                conn,
                """
                CREATE TABLE IF NOT EXISTS tg.dialog_snapshot (
                  chat_id       bigint PRIMARY KEY,
                  title         text NOT NULL DEFAULT '',
                  unread_count  integer NOT NULL DEFAULT 0,
                  is_group      boolean NOT NULL DEFAULT false,
                  is_channel    boolean NOT NULL DEFAULT false,
                  updated_at    timestamptz NOT NULL DEFAULT now()
                )
                """,
            )
            exec1(
                conn,
                "CREATE INDEX IF NOT EXISTS idx_tg_dialog_snapshot_updated_at ON tg.dialog_snapshot(updated_at DESC)",
            )
            conn.commit()
    except Exception:
        # Ignore startup migration errors to avoid breaking app boot
        pass
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Compose service name: "postgres"
DB_DSN = os.getenv(
    "DATABASE_URL",
    "postgresql://gamesales_app:najTylth1@postgres:5432/gamesales",
)

JWT_SECRET = os.getenv("JWT_SECRET", "")
JWT_ALG = os.getenv("JWT_ALG", "HS256")
JWT_TTL_MIN = int(os.getenv("JWT_TTL_MIN", "720"))

pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")

def now_utc():
    return datetime.now(timezone.utc)

MIN_DATE = date(2020, 1, 1)

def validate_date_in_range(value, field_name: str):
    if value is None:
        return
    if isinstance(value, datetime):
        check_date = value.date()
    else:
        check_date = value
    max_date = now_utc().date()
    if check_date < MIN_DATE or check_date > max_date:
        raise HTTPException(400, f"{field_name} must be between {MIN_DATE.isoformat()} and {max_date.isoformat()}")

def validate_date_not_future(value, field_name: str):
    if value is None:
        return
    if isinstance(value, datetime):
        check_date = value.date()
    else:
        check_date = value
    max_date = now_utc().date()
    if check_date > max_date:
        raise HTTPException(400, f"{field_name} must be <= {max_date.isoformat()}")

def validate_date_range(start_value, end_value, field_name: str):
    if not start_value or not end_value:
        return
    if isinstance(start_value, datetime):
        start_date = start_value.date()
    else:
        start_date = start_value
    if isinstance(end_value, datetime):
        end_date = end_value.date()
    else:
        end_date = end_value
    if end_date < start_date:
        raise HTTPException(400, f"{field_name} must be >= start_at")

def normalize_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=timezone.utc)
    return value

def normalize_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return value

# ----------------------------
# Models
# ----------------------------
class AccountCreate(BaseModel):
    region_code: Optional[str] = Field(None, description="RU/TR/US/EU")
    login_name: Optional[str] = None
    domain_code: Optional[str] = Field(None, description="email domain, e.g. example.com")
    account_date: Optional[date] = None
    notes: Optional[str] = None

    @validator("account_date", pre=False)
    def normalize_account_date(cls, v):
        return normalize_date(v)

class AccountUpdate(BaseModel):
    region_code: Optional[str] = None
    login_name: Optional[str] = None
    domain_code: Optional[str] = None
    status_code: Optional[str] = None
    account_date: Optional[date] = None
    notes: Optional[str] = None

    @validator("account_date", pre=False)
    def normalize_account_date(cls, v):
        return normalize_date(v)

class AccountPlatformSlots(BaseModel):
    platform_code: str
    slot_capacity: int
    occupied_slots: int
    free_slots: int

class AccountSlotStatusOut(BaseModel):
    slot_type_code: str
    platform_code: str
    mode: str
    capacity: int
    occupied: int
    free: int

class AccountOut(BaseModel):
    account_id: int
    region_code: Optional[str]
    status: str
    login_name: Optional[str]
    domain_code: Optional[str]
    login_full: Optional[str]
    platform_slots: List[AccountPlatformSlots]
    slot_status: List[AccountSlotStatusOut] = []
    game_titles: Optional[List[str]] = None
    platform_codes: Optional[List[str]] = None
    account_date: Optional[date] = None
    notes: Optional[str] = None

class AccountListOut(BaseModel):
    total: int
    items: List[AccountOut]

class GameCreate(BaseModel):
    title: str
    short_title: Optional[str] = None
    link: Optional[str] = None
    logo_url: Optional[str] = None
    text_lang: Optional[str] = None
    audio_lang: Optional[str] = None
    vr_support: Optional[str] = None
    platform_codes: Optional[List[str]] = None
    region_code: Optional[str] = None

class GameUpdate(BaseModel):
    title: Optional[str] = None
    short_title: Optional[str] = None
    link: Optional[str] = None
    logo_url: Optional[str] = None
    text_lang: Optional[str] = None
    audio_lang: Optional[str] = None
    vr_support: Optional[str] = None
    platform_codes: Optional[List[str]] = None
    region_code: Optional[str] = None

class GameOut(BaseModel):
    game_id: int
    title: str
    short_title: Optional[str] = None
    link: Optional[str] = None
    logo_url: Optional[str] = None
    text_lang: Optional[str] = None
    audio_lang: Optional[str] = None
    vr_support: Optional[str] = None
    platform_codes: List[str]
    region_code: Optional[str]

class GameListOut(BaseModel):
    total: int
    items: List[GameOut]

class PlatformOut(BaseModel):
    code: str
    name: str
    slot_capacity: int

class SourceOut(BaseModel):
    source_id: int
    code: str
    name: str

class RegionOut(BaseModel):
    code: str
    name: str
    purchase_cost_rate: float
    purchase_cost_rate: float

class PlatformIn(BaseModel):
    code: str
    name: str
    slot_capacity: int = 0

class PlatformUpdate(BaseModel):
    name: Optional[str] = None
    slot_capacity: Optional[int] = None

class RegionIn(BaseModel):
    code: str
    name: str
    purchase_cost_rate: float = 1.0

class RegionUpdate(BaseModel):
    name: Optional[str] = None
    purchase_cost_rate: Optional[float] = None

class RentalCreate(BaseModel):
    account_id: int
    customer_nickname: str
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    slots_used: int = 1
    price: float = 0
    game_id: Optional[int] = None
    platform_code: Optional[str] = None
    slot_type_code: Optional[str] = None
    source_id: Optional[int] = None
    purchase_at: Optional[datetime] = None

    @validator("purchase_at", "start_at", "end_at", pre=False)
    def normalize_dt(cls, v):
        return normalize_datetime(v)

class LoginIn(BaseModel):
    username: str
    password: str

class UserOut(BaseModel):
    username: str
    role: str


class TelegramAuthStartIn(BaseModel):
    phone: str


class TelegramAuthConfirmIn(BaseModel):
    code: str


class TelegramAuthPasswordIn(BaseModel):
    password: str


class TelegramDialogsOut(BaseModel):
    items: list
    counts: Dict[str, int] = {}
    sync_running: bool = False
    last_sync_at: Optional[datetime] = None
    sync_loaded: int = 0
    sync_batches: int = 0


class TelegramMessagesOut(BaseModel):
    items: list


class TelegramSendMessageIn(BaseModel):
    chat_id: int
    text: str

class TelegramDialogStatusIn(BaseModel):
    status: str

class TelegramContactIn(BaseModel):
    sender_id: int
    title: str = ""
    info: str = ""

class ImportIssue(BaseModel):
    row: int
    field: str
    value: Optional[str] = None
    message: str

class ImportReportIn(BaseModel):
    errors: List[ImportIssue] = []
    warnings: List[ImportIssue] = []

class LoginOut(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserOut

class UserCreate(BaseModel):
    username: str
    password: str
    role_code: str = "manager"

class UserListOut(BaseModel):
    username: str
    role: str
    created_at: datetime

class ChangePasswordIn(BaseModel):
    current_password: str
    new_password: str

class ResetPasswordIn(BaseModel):
    new_password: str

class RoleOut(BaseModel):
    code: str
    name: str

class DealCreate(BaseModel):
    deal_type_code: str = Field(..., description="sale/rental")
    account_id: Optional[int] = None
    game_id: Optional[int] = None
    customer_nickname: str
    source_id: Optional[int] = None
    region_code: Optional[str] = None
    platform_code: Optional[str] = None
    slot_type_code: Optional[str] = None
    price: float = 0
    purchase_cost: float = 0
    game_link: Optional[str] = None
    purchase_at: Optional[datetime] = None
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    slots_used: int = 1
    notes: Optional[str] = None

    @validator("purchase_at", "start_at", "end_at", pre=False)
    def normalize_dt(cls, v):
        return normalize_datetime(v)

class DealUpdate(BaseModel):
    deal_type_code: Optional[str] = None
    account_id: Optional[int] = None
    game_id: Optional[int] = None
    customer_nickname: Optional[str] = None
    source_id: Optional[int] = None
    region_code: Optional[str] = None
    platform_code: Optional[str] = None
    slot_type_code: Optional[str] = None
    price: Optional[float] = None
    purchase_cost: Optional[float] = None
    game_link: Optional[str] = None
    purchase_at: Optional[datetime] = None
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    slots_used: Optional[int] = None
    notes: Optional[str] = None
    flow_status_code: Optional[str] = None

    @validator("purchase_at", "start_at", "end_at", pre=False)
    def normalize_dt(cls, v):
        return normalize_datetime(v)

class DealListItem(BaseModel):
    deal_id: int
    deal_type: str
    deal_type_code: Optional[str] = None
    status: str
    flow_status: Optional[str] = None
    flow_status_code: Optional[str] = None
    account_id: Optional[int] = None
    account_login: Optional[str]
    region_code: Optional[str]
    game_id: Optional[int]
    game_title: Optional[str]
    game_short_title: Optional[str] = None
    platform_code: Optional[str]
    slot_type_code: Optional[str] = None
    customer_nickname: Optional[str]
    source_id: Optional[int]
    price: float
    purchase_cost: Optional[float] = None
    game_link: Optional[str] = None
    purchase_at: Optional[datetime]
    created_at: datetime
    slots_used: Optional[int] = None
    notes: Optional[str] = None

class SalesAnalyticsTotals(BaseModel):
    revenue: float
    purchase_cost: float
    margin: float
    count: int
    avg_check: float

class SalesAnalyticsPoint(BaseModel):
    date: date
    revenue: float
    purchase_cost: float
    margin: float
    count: int

class SalesAnalyticsByType(BaseModel):
    deal_type_code: str
    revenue: float
    purchase_cost: float
    margin: float
    count: int

class SalesAnalyticsOut(BaseModel):
    totals: SalesAnalyticsTotals
    by_day: List[SalesAnalyticsPoint]
    by_type: List[SalesAnalyticsByType]

class SourceAnalyticsItem(BaseModel):
    source_id: Optional[int]
    source_code: Optional[str]
    source_name: Optional[str]
    deals_count: int
    revenue: float

class RepeatCustomersOut(BaseModel):
    repeat_count: int
    total_customers: int
    repeat_share: float

class SourcesAnalyticsOut(BaseModel):
    top_by_count: List[SourceAnalyticsItem]
    top_by_revenue: List[SourceAnalyticsItem]
    repeat_customers: RepeatCustomersOut

class AuditItemOut(BaseModel):
    deal_id: Optional[int]
    table_name: str
    action: str
    changed_at: datetime
    changed_by: Optional[str]

class AuditAnalyticsOut(BaseModel):
    total: int
    items: List[AuditItemOut]

class SlotTypeOut(BaseModel):
    code: str
    name: str
    platform_code: str
    mode: str
    capacity: int

class AccountSlotAssignmentOut(BaseModel):
    assignment_id: int
    account_id: int
    account_login: Optional[str] = None
    slot_type_code: str
    customer_id: Optional[int]
    customer_nickname: Optional[str]
    game_id: Optional[int]
    game_title: Optional[str]
    deal_id: Optional[int]
    deal_item_id: Optional[int]
    assigned_at: datetime
    released_at: Optional[datetime]
    assigned_by: Optional[str]
    released_by: Optional[str]

class DealListOut(BaseModel):
    total: int
    items: List[DealListItem]

class DomainIn(BaseModel):
    name: str

class SourceIn(BaseModel):
    code: str
    name: str

class SourceUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None

class NameUpdate(BaseModel):
    name: str

class AccountSecretIn(BaseModel):
    secret_key: str
    secret_value: str

class AccountSecretOut(BaseModel):
    secret_key: str
    secret_value_b64: str
    created_at: datetime

class AccountSecretsBatchIn(BaseModel):
    account_ids: List[int]

class AccountSecretsBatchItem(BaseModel):
    account_id: int
    secrets: List[AccountSecretOut]

class SlotAvailabilityOut(BaseModel):
    slot_type_code: str
    has_free: bool

class AccountGamesIn(BaseModel):
    game_ids: List[int]

class GameAccountOut(BaseModel):
    account_id: int
    login_full: Optional[str]
    platform_code: str
    free_slots: int
    occupied_slots: int

# ----------------------------
# DB helpers
# ----------------------------
def q1(conn, sql, params=None):
    with conn.cursor() as cur:
        cur.execute(sql, params or ())
        row = cur.fetchone()
        return row

def qall(conn, sql, params=None):
    with conn.cursor() as cur:
        cur.execute(sql, params or ())
        return cur.fetchall()

def exec1(conn, sql, params=None):
    with conn.cursor() as cur:
        cur.execute(sql, params or ())
        return cur.rowcount

def get_platform_id(conn, code: str) -> int:
    row = q1(conn, "SELECT platform_id FROM app.platforms WHERE code=%s AND is_archived IS NOT TRUE", (code,))
    if not row:
        raise HTTPException(400, f"Unknown platform_code: {code}")
    return int(row[0])

def get_platform_info(conn, code: str) -> tuple[int, int]:
    row = q1(conn, "SELECT platform_id, slot_capacity FROM app.platforms WHERE code=%s AND is_archived IS NOT TRUE", (code,))
    if not row:
        raise HTTPException(400, f"Unknown platform_code: {code}")
    return int(row[0]), int(row[1] or 0)

def get_region_id(conn, code: Optional[str]) -> Optional[int]:
    if not code:
        return None
    row = q1(conn, "SELECT region_id FROM app.regions WHERE code=%s AND is_archived IS NOT TRUE", (code,))
    if not row:
        raise HTTPException(400, f"Unknown region_code: {code}")
    return int(row[0])

def get_domain_id(conn, code: Optional[str]) -> Optional[int]:
    if not code:
        return None
    row = q1(conn, "SELECT domain_id FROM app.domains WHERE name=%s AND is_archived IS NOT TRUE", (code,))
    if not row:
        raise HTTPException(400, f"Unknown domain: {code}")
    return int(row[0])

def get_platform_id_optional(conn, code: Optional[str]) -> Optional[int]:
    if not code:
        return None
    row = q1(conn, "SELECT platform_id FROM app.platforms WHERE code=%s AND is_archived IS NOT TRUE", (code,))
    if not row:
        raise HTTPException(400, f"Unknown platform_code: {code}")
    return int(row[0])

def ensure_account_exists(conn, account_id: int):
    row = q1(conn, "SELECT 1 FROM app.accounts WHERE account_id=%s", (account_id,))
    if not row:
        raise HTTPException(400, f"Unknown account_id: {account_id}")

def ensure_game_exists(conn, game_id: Optional[int]):
    if not game_id:
        return
    row = q1(conn, "SELECT 1 FROM app.game_titles WHERE game_id=%s", (game_id,))
    if not row:
        raise HTTPException(400, f"Unknown game_id: {game_id}")

def ensure_game_active(conn, game_id: Optional[int]):
    if not game_id:
        return
    row = q1(conn, "SELECT 1 FROM app.game_titles WHERE game_id=%s AND is_archived IS NOT TRUE", (game_id,))
    if not row:
        raise HTTPException(400, f"Unknown game_id: {game_id}")

def ensure_customer(conn, nickname: Optional[str], source_id: Optional[int]) -> Optional[int]:
    if not nickname:
        return None
    row = q1(conn, "SELECT customer_id, source_id FROM app.customers WHERE nickname=%s", (nickname,))
    if row:
        customer_id = int(row[0])
        if source_id and row[1] != source_id:
            exec1(conn, "UPDATE app.customers SET source_id=%s WHERE customer_id=%s", (source_id, customer_id))
        return customer_id
    row = q1(
        conn,
        "INSERT INTO app.customers(nickname, source_id) VALUES (%s, %s) RETURNING customer_id",
        (nickname, source_id),
    )
    return int(row[0])

def ensure_source_exists(conn, source_id: Optional[int]):
    if not source_id:
        return
    row = q1(conn, "SELECT 1 FROM app.sources WHERE source_id=%s AND is_archived IS NOT TRUE", (source_id,))
    if not row:
        raise HTTPException(400, f"Unknown source_id: {source_id}")

def build_deals_filters(
    account_id: Optional[int],
    game_id: Optional[int],
    platform_code: Optional[str],
    q: Optional[str],
    deal_type_code: Optional[str],
    status_code: Optional[str],
    flow_status_code: Optional[str],
    customer_q: Optional[str],
    source_id: Optional[int],
    purchase_from: Optional[date],
    purchase_to: Optional[date],
    price_min: Optional[float],
    price_max: Optional[float],
    notes_q: Optional[str],
    account_q: Optional[str],
    region_q: Optional[str],
    game_q: Optional[str],
    platform_q: Optional[str],
    type_q: Optional[str],
    status_q: Optional[str],
    flow_status_q: Optional[str],
    source_q: Optional[str],
    date_q: Optional[str],
    price_q: Optional[str],
) -> Tuple[str, List]:
    where = []
    params: List = []
    if account_id:
        where.append("di.account_id = %s")
        params.append(account_id)
    if game_id:
        where.append("di.game_id = %s")
        params.append(game_id)
    if platform_code:
        where.append("p.code = %s")
        params.append(platform_code)
    if q:
        where.append("""
          (
            c.nickname ILIKE %s
            OR COALESCE(rd.code, ra.code) ILIKE %s
            OR dt.name ILIKE %s
            OR dt.code ILIKE %s
            OR fs.name ILIKE %s
            OR fs.code ILIKE %s
            OR COALESCE(di.purchase_at, d.created_at)::text ILIKE %s
          )
        """)
        like = f"%{q}%"
        params.extend([like, like, like, like, like, like, like])
    if deal_type_code:
        where.append("d.deal_type_code = %s")
        params.append(deal_type_code)
    if status_code:
        where.append("d.status_code = %s")
        params.append(status_code)
    if flow_status_code:
        where.append("d.flow_status_code = %s")
        params.append(flow_status_code)
    if customer_q:
        where.append("c.nickname ILIKE %s")
        params.append(f"%{customer_q}%")
    if source_id:
        where.append("c.source_id = %s")
        params.append(source_id)
    if notes_q:
        where.append("di.notes ILIKE %s")
        params.append(f"%{notes_q}%")
    if account_q:
        like = f"%{account_q}%"
        where.append("(a.login_name ILIKE %s OR dm.name ILIKE %s)")
        params.extend([like, like])
    if region_q:
        like = f"%{region_q}%"
        where.append("COALESCE(rd.code, ra.code) ILIKE %s")
        params.append(like)
    if game_q:
        where.append("g.title ILIKE %s")
        params.append(f"%{game_q}%")
    if platform_q:
        like = f"%{platform_q}%"
        where.append("(p.code ILIKE %s OR p.name ILIKE %s)")
        params.extend([like, like])
    if type_q:
        like = f"%{type_q}%"
        where.append("(dt.code ILIKE %s OR dt.name ILIKE %s)")
        params.extend([like, like])
    if status_q:
        like = f"%{status_q}%"
        where.append("(ds.code ILIKE %s OR ds.name ILIKE %s)")
        params.extend([like, like])
    if flow_status_q:
        like = f"%{flow_status_q}%"
        where.append("(fs.code ILIKE %s OR fs.name ILIKE %s)")
        params.extend([like, like])
    if source_q:
        like = f"%{source_q}%"
        where.append("(src.code ILIKE %s OR src.name ILIKE %s)")
        params.extend([like, like])
    if date_q:
        where.append("COALESCE(di.purchase_at, d.created_at)::text ILIKE %s")
        params.append(f"%{date_q}%")
    if price_q:
        where.append("di.price::text ILIKE %s")
        params.append(f"%{price_q}%")
    if purchase_from:
        where.append("COALESCE(di.purchase_at, d.created_at)::date >= %s")
        params.append(purchase_from)
    if purchase_to:
        where.append("COALESCE(di.purchase_at, d.created_at)::date <= %s")
        params.append(purchase_to)
    if price_min is not None:
        where.append("di.price >= %s")
        params.append(price_min)
    if price_max is not None:
        where.append("di.price <= %s")
        params.append(price_max)
    where_sql = "WHERE " + " AND ".join(where) if where else ""
    return where_sql, params

def slots_summary(conn, account_id: int, platform_id: int):
    row = q1(
        conn,
        """
        SELECT occupied_slots, free_slots
        FROM app.v_account_platform_slots
        WHERE account_id=%s AND platform_id=%s
        """,
        (account_id, platform_id),
    )
    if not row:
        return (0, 0)
    return int(row[0]), int(row[1])

def get_account_platform_slots(conn, account_id: int) -> List[AccountPlatformSlots]:
    rows = qall(
        conn,
        """
        SELECT p.code, s.slot_capacity, s.occupied_slots, s.free_slots
        FROM app.v_account_platform_slots s
        JOIN app.platforms p ON p.platform_id = s.platform_id
        WHERE s.account_id=%s
        ORDER BY p.code
        """,
        (account_id,),
    )
    return [
        AccountPlatformSlots(
            platform_code=r0,
            slot_capacity=int(r1 or 0),
            occupied_slots=int(r2 or 0),
            free_slots=int(r3 or 0),
        )
        for (r0, r1, r2, r3) in rows
    ]

def get_account_slot_status(conn, account_id: int) -> List[AccountSlotStatusOut]:
    rows = qall(
        conn,
        """
        SELECT slot_type_code, platform_code, mode, capacity, occupied, free
        FROM app.v_account_slot_status
        WHERE account_id=%s
        ORDER BY slot_type_code
        """,
        (account_id,),
    )
    return [
        AccountSlotStatusOut(
            slot_type_code=r0,
            platform_code=r1,
            mode=r2,
            capacity=int(r3 or 0),
            occupied=int(r4 or 0),
            free=int(r5 or 0),
        )
        for (r0, r1, r2, r3, r4, r5) in rows
    ]

def normalize_platform_codes(codes: Optional[List[str]]) -> List[str]:
    if not codes:
        return []
    uniq = []
    seen = set()
    for code in codes:
        if not code:
            continue
        val = str(code).strip().lower()
        if not val or val in seen:
            continue
        seen.add(val)
        uniq.append(val)
    return uniq

def get_game_platform_codes(conn, game_id: int) -> List[str]:
    rows = qall(
        conn,
        """
        SELECT p.code
        FROM app.game_platforms gp
        JOIN app.platforms p ON p.platform_id = gp.platform_id
        WHERE gp.game_id=%s
        ORDER BY p.code
        """,
        (game_id,),
    )
    return [r[0] for r in rows]

def get_account_platform_codes(conn, account_id: int) -> List[str]:
    rows = qall(
        conn,
        """
        SELECT DISTINCT p.code
        FROM app.account_assets aa
        JOIN app.game_platforms gp ON gp.game_id = aa.game_id
        JOIN app.platforms p ON p.platform_id = gp.platform_id
        WHERE aa.account_id=%s AND aa.asset_type_code='game'
        ORDER BY p.code
        """,
        (account_id,),
    )
    return [r[0] for r in rows]

def account_has_ps4(conn, account_id: int) -> bool:
    row = q1(
        conn,
        """
        SELECT 1
        FROM app.account_assets aa
        JOIN app.game_platforms gp ON gp.game_id = aa.game_id
        JOIN app.platforms p ON p.platform_id = gp.platform_id
        WHERE aa.account_id=%s AND aa.asset_type_code='game' AND p.code='ps4'
        LIMIT 1
        """,
        (account_id,),
    )
    return bool(row)

def ensure_account_allows_slot_type(conn, account_id: int, slot_type_code: str):
    slot_type = get_slot_type(conn, slot_type_code)
    platform_code = slot_type[1]
    if platform_code == "ps4" and not account_has_ps4(conn, account_id):
        raise HTTPException(400, "Account does not support PS4 slot type")
    return slot_type

def encode_b64(value: bytes) -> str:
    return base64.b64encode(value).decode("ascii")

def get_slot_type(conn, slot_type_code: str):
    row = q1(
        conn,
        "SELECT code, platform_code, mode, capacity FROM app.slot_types WHERE code=%s",
        (slot_type_code,),
    )
    if not row:
        raise HTTPException(400, "Unknown slot_type_code")
    return row

def get_account_slot_free(conn, account_id: int, slot_type_code: str) -> int:
    row = q1(
        conn,
        """
        SELECT free
        FROM app.v_account_slot_status
        WHERE account_id=%s AND slot_type_code=%s
        """,
        (account_id, slot_type_code),
    )
    return int(row[0] or 0) if row else 0

def release_slot_assignment(conn, deal_item_id: int, released_by: Optional[str]):
    exec1(
        conn,
        """
        UPDATE app.account_slot_assignments
        SET released_at=now(), released_by=%s
        WHERE deal_item_id=%s AND released_at IS NULL
        """,
        (released_by, deal_item_id),
    )

MAX_LOGO_BYTES = 5 * 1024 * 1024
ALLOWED_LOGO_MIME = {"image/jpeg", "image/png", "image/webp"}

ACCOUNT_IMPORT_HEADER_MAP = {
    "Аккаунт": "account",
    "аккаунт": "account",
    "Login": "account",
    "Логин": "account",
    "логин": "account",
    "Пароль": "password",
    "пароль": "password",
    "Password": "password",
    "Игра": "game",
    "игра": "game",
    "Game": "game",
}
SLOT_IMPORT_HEADER_MAP = {
    "Аккаунт": "account",
    "аккаунт": "account",
    "Игра": "game",
    "игра": "game",
    "Статус": "status",
    "статус": "status",
    "Пользователь": "customer",
    "пользователь": "customer",
    "Откуда": "source",
    "откуда": "source",
    "Дата": "date",
    "дата": "date",
    "Слот": "slot",
    "слот": "slot",
}
LOGO_DOWNLOAD_DELAY_SEC = 0.25
IMPORT_STATUS_TTL_SEC = 24 * 60 * 60
_REDIS_CLIENT = None
_QUEUE_API_URL = os.getenv("QUEUE_API_URL", "")
_QUEUE_API_KEY = os.getenv("QUEUE_API_KEY", "")
_TELEGRAM_API_URL = os.getenv("TELEGRAM_API_URL", "")
_TELEGRAM_API_KEY = os.getenv("TELEGRAM_API_KEY", "")

_TELEGRAM_DIALOGS_SYNC_LOCK = threading.Lock()
_TELEGRAM_DIALOGS_SYNC_RUNNING = False
_TELEGRAM_DIALOGS_SYNC_LAST_STARTED_MONO = 0.0
_TELEGRAM_DIALOGS_SYNC_LOADED = 0
_TELEGRAM_DIALOGS_SYNC_BATCHES = 0
TELEGRAM_DIALOGS_SYNC_LIMIT = int(os.getenv("TELEGRAM_DIALOGS_SYNC_LIMIT", "0") or "0")
TELEGRAM_DIALOGS_SYNC_BATCH = int(os.getenv("TELEGRAM_DIALOGS_SYNC_BATCH", "100") or "100")
TELEGRAM_DIALOGS_SYNC_COOLDOWN_SEC = int(os.getenv("TELEGRAM_DIALOGS_SYNC_COOLDOWN_SEC", "45") or "45")
TELEGRAM_DIALOGS_SYNC_BATCH_DELAY_MS = int(os.getenv("TELEGRAM_DIALOGS_SYNC_BATCH_DELAY_MS", "250") or "250")

def get_redis():
    global _REDIS_CLIENT
    if _REDIS_CLIENT is None:
        url = os.getenv("REDIS_URL", "redis://localhost:6379/0")
        _REDIS_CLIENT = redis.Redis.from_url(url, decode_responses=True)
    return _REDIS_CLIENT

def _job_key(job_id: str) -> str:
    return f"games_import:{job_id}"

def _queue_api_request(method: str, path: str, data: Optional[dict] = None) -> Optional[dict]:
    if not _QUEUE_API_URL:
        return None
    url = _QUEUE_API_URL.rstrip("/") + path
    body = None
    headers = {"Content-Type": "application/json"}
    if _QUEUE_API_KEY:
        headers["X-API-Key"] = _QUEUE_API_KEY
    if data is not None:
        body = json.dumps(data).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            content = resp.read()
            if not content:
                return None
            return json.loads(content.decode("utf-8"))
    except urllib.error.HTTPError as exc:
        raw = exc.read()
        details = raw.decode("utf-8", errors="ignore") if raw else ""
        message = f"Queue API {method} {path} failed: {exc.code} {exc.reason}"
        if details:
            message = f"{message}. {details}"
        raise RuntimeError(message) from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"Queue API {method} {path} failed: {exc.reason}") from exc


def _telegram_api_request(method: str, path: str, data: Optional[dict] = None, timeout_sec: int = 15) -> dict:
    if not _TELEGRAM_API_URL:
        raise HTTPException(500, "Telegram service is not configured")
    url = _TELEGRAM_API_URL.rstrip("/") + path
    body = None
    headers = {"Content-Type": "application/json"}
    if _TELEGRAM_API_KEY:
        headers["X-API-Key"] = _TELEGRAM_API_KEY
    if data is not None:
        body = json.dumps(data).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=timeout_sec) as resp:
            content = resp.read()
            return json.loads(content.decode("utf-8")) if content else {}
    except urllib.error.HTTPError as exc:
        raw = exc.read()
        details = raw.decode("utf-8", errors="ignore") if raw else ""
        message = f"Telegram service {method} {path} failed: {exc.code} {exc.reason}"
        if details:
            message = f"{message}. {details}"
        raise HTTPException(exc.code, message)
    except urllib.error.URLError as exc:
        raise HTTPException(502, f"Telegram service {method} {path} failed: {exc.reason}")
    except (TimeoutError, socket.timeout):
        raise HTTPException(504, f"Telegram service {method} {path} timed out")

def _telegram_api_request_raw(method: str, path: str, data: Optional[dict] = None) -> Tuple[bytes, Optional[str]]:
    if not _TELEGRAM_API_URL:
        raise HTTPException(500, "Telegram service is not configured")
    url = _TELEGRAM_API_URL.rstrip("/") + path
    body = None
    headers = {"Content-Type": "application/json"}
    if _TELEGRAM_API_KEY:
        headers["X-API-Key"] = _TELEGRAM_API_KEY
    if data is not None:
        body = json.dumps(data).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            content = resp.read()
            content_type = resp.headers.get("Content-Type")
            return content or b"", content_type
    except urllib.error.HTTPError as exc:
        raw = exc.read()
        details = raw.decode("utf-8", errors="ignore") if raw else ""
        message = f"Telegram service {method} {path} failed: {exc.code} {exc.reason}"
        if details:
            message = f"{message}. {details}"
        raise HTTPException(exc.code, message)
    except urllib.error.URLError as exc:
        raise HTTPException(502, f"Telegram service {method} {path} failed: {exc.reason}")
    except (TimeoutError, socket.timeout):
        raise HTTPException(504, f"Telegram service {method} {path} timed out")


def _upsert_telegram_dialog_snapshot(conn, user_id: int, items: List[dict]):
    if not items:
        return
    payload = []
    chat_ids = []
    archived_with_unread = []
    for item in items:
        raw_id = item.get("id")
        if raw_id is None:
            continue
        chat_id = int(raw_id)
        title = str(item.get("title") or "")
        unread_count = int(item.get("unread_count") or 0)
        is_group = bool(item.get("is_group"))
        is_channel = bool(item.get("is_channel"))
        payload.append((chat_id, title, unread_count, is_group, is_channel))
        chat_ids.append(chat_id)
        if unread_count > 0:
            archived_with_unread.append(chat_id)

    if not payload:
        return

    with conn.cursor() as cur:
        cur.executemany(
            """
            INSERT INTO tg.dialog_snapshot(chat_id, title, unread_count, is_group, is_channel, updated_at)
            VALUES (%s, %s, %s, %s, %s, now())
            ON CONFLICT (chat_id)
            DO UPDATE SET
              title=excluded.title,
              unread_count=excluded.unread_count,
              is_group=excluded.is_group,
              is_channel=excluded.is_channel,
              updated_at=now()
            """,
            payload,
        )
        # First time seen contact -> new
        cur.executemany(
            """
            INSERT INTO tg.dialog_states(chat_id, status, updated_at, updated_by_user_id)
            VALUES (%s, 'new', now(), %s)
            ON CONFLICT (chat_id) DO NOTHING
            """,
            [(cid, user_id) for cid in chat_ids],
        )

    # If archived contact writes again, move it to accepted.
    if archived_with_unread:
        exec1(
            conn,
            """
            UPDATE tg.dialog_states
            SET status='accepted', updated_at=now(), updated_by_user_id=%s
            WHERE chat_id = ANY(%s) AND status='archived'
            """,
            (user_id, archived_with_unread),
        )

def _delete_unseen_dialogs_from_snapshot(conn, seen_chat_ids: List[int]):
    if not seen_chat_ids:
        return
    exec1(
        conn,
        "DELETE FROM tg.dialog_snapshot WHERE NOT (chat_id = ANY(%s))",
        (seen_chat_ids,),
    )

def _delete_dead_dialog(conn, chat_id: int):
    exec1(conn, "DELETE FROM tg.dialog_snapshot WHERE chat_id=%s", (chat_id,))
    # Keep status row for audit/history semantics; it will be reused if chat appears again.

def _telegram_fetch_dialog_batch_with_retry(
    session_string: str,
    limit: int,
    offset_date: Optional[str],
    offset_id: int,
) -> dict:
    delays = (0, 2, 5)
    last_exc: Optional[HTTPException] = None
    for idx, delay_sec in enumerate(delays):
        if delay_sec > 0:
            time.sleep(delay_sec)
        try:
            return _telegram_api_request(
                "POST",
                "/dialogs",
                {
                    "session_string": session_string,
                    "limit": limit,
                    "offset_date": offset_date,
                    "offset_id": offset_id,
                    "only_private": True,
                },
                timeout_sec=45,
            )
        except HTTPException as exc:
            last_exc = exc
            # Retry only for transient transport/rate limit conditions.
            if exc.status_code not in (429, 502, 504) or idx == len(delays) - 1:
                raise
    if last_exc:
        raise last_exc
    return {"items": [], "has_more": False, "next_offset_date": None, "next_offset_id": 0}

def _telegram_dialogs_sync_worker(session_string: str, user_id: int):
    global _TELEGRAM_DIALOGS_SYNC_RUNNING, _TELEGRAM_DIALOGS_SYNC_LOADED, _TELEGRAM_DIALOGS_SYNC_BATCHES
    try:
        # 0 or less means "sync all dialogs" in batches.
        sync_limit: Optional[int] = TELEGRAM_DIALOGS_SYNC_LIMIT if TELEGRAM_DIALOGS_SYNC_LIMIT > 0 else None
        batch_size = TELEGRAM_DIALOGS_SYNC_BATCH if TELEGRAM_DIALOGS_SYNC_BATCH > 0 else 100
        batch_delay_sec = max(TELEGRAM_DIALOGS_SYNC_BATCH_DELAY_MS, 0) / 1000.0
        offset_date: Optional[str] = None
        offset_id = 0
        total_loaded = 0
        rounds = 0
        seen_chat_ids: List[int] = []
        completed_full_scan = False
        with psycopg.connect(DB_DSN) as conn:
            while True:
                if sync_limit is not None:
                    remaining = max(sync_limit - total_loaded, 0)
                    if remaining <= 0:
                        break
                    current_batch = min(batch_size, remaining)
                else:
                    current_batch = batch_size
                resp = _telegram_fetch_dialog_batch_with_retry(session_string, current_batch, offset_date, offset_id)
                items = resp.get("items", []) or []
                if not items:
                    completed_full_scan = True
                    break
                _upsert_telegram_dialog_snapshot(conn, user_id, items)
                for item in items:
                    if item.get("id") is not None:
                        seen_chat_ids.append(int(item["id"]))
                conn.commit()
                total_loaded += len(items)
                rounds += 1
                with _TELEGRAM_DIALOGS_SYNC_LOCK:
                    _TELEGRAM_DIALOGS_SYNC_LOADED = total_loaded
                    _TELEGRAM_DIALOGS_SYNC_BATCHES = rounds
                has_more = bool(resp.get("has_more"))
                if not has_more:
                    completed_full_scan = True
                    break
                offset_date = resp.get("next_offset_date")
                offset_id = int(resp.get("next_offset_id") or 0)
                if not offset_date or rounds > 500:
                    break
                if batch_delay_sec > 0:
                    time.sleep(batch_delay_sec)
            if completed_full_scan and seen_chat_ids:
                _delete_unseen_dialogs_from_snapshot(conn, seen_chat_ids)
            exec1(conn, "UPDATE tg.shared_session SET last_used_at=now() WHERE id=1")
            conn.commit()
    except Exception as exc:
        # Background sync errors must not break API response path.
        print(f"Telegram dialogs sync failed: {exc}")
    finally:
        with _TELEGRAM_DIALOGS_SYNC_LOCK:
            _TELEGRAM_DIALOGS_SYNC_RUNNING = False


def _trigger_telegram_dialogs_sync(session_string: str, user_id: int) -> bool:
    global _TELEGRAM_DIALOGS_SYNC_RUNNING, _TELEGRAM_DIALOGS_SYNC_LAST_STARTED_MONO, _TELEGRAM_DIALOGS_SYNC_LOADED, _TELEGRAM_DIALOGS_SYNC_BATCHES
    now_mono = time.monotonic()
    with _TELEGRAM_DIALOGS_SYNC_LOCK:
        if _TELEGRAM_DIALOGS_SYNC_RUNNING:
            return False
        if (now_mono - _TELEGRAM_DIALOGS_SYNC_LAST_STARTED_MONO) < TELEGRAM_DIALOGS_SYNC_COOLDOWN_SEC:
            return False
        _TELEGRAM_DIALOGS_SYNC_RUNNING = True
        _TELEGRAM_DIALOGS_SYNC_LAST_STARTED_MONO = now_mono
        _TELEGRAM_DIALOGS_SYNC_LOADED = 0
        _TELEGRAM_DIALOGS_SYNC_BATCHES = 0
    thread = threading.Thread(
        target=_telegram_dialogs_sync_worker,
        args=(session_string, user_id),
        daemon=True,
    )
    thread.start()
    return True


def is_import_cancelled(job_id: str) -> bool:
    status = get_import_progress(job_id)
    return bool(status.get("cancelled"))

def set_import_progress(job_id: str, owner: str, payload: dict):
    data = {"owner": owner}
    data.update(payload)
    if _QUEUE_API_URL:
        try:
            _queue_api_request("POST", f"/status/{job_id}", {"data": data})
        except Exception as exc:
            print(f"Queue API status update failed: {exc}")
        return
    r = get_redis()
    r.setex(_job_key(job_id), IMPORT_STATUS_TTL_SEC, json.dumps(data))

def get_import_progress(job_id: str) -> dict:
    if _QUEUE_API_URL:
        try:
            return _queue_api_request("GET", f"/status/{job_id}") or {}
        except Exception as exc:
            print(f"Queue API status read failed: {exc}")
            return {}
    r = get_redis()
    raw = r.get(_job_key(job_id))
    if not raw:
        return {}
    try:
        return json.loads(raw)
    except Exception:
        return {}

def clear_import_progress(job_id: str):
    if _QUEUE_API_URL:
        try:
            _queue_api_request("DELETE", f"/status/{job_id}")
        except Exception as exc:
            print(f"Queue API status delete failed: {exc}")
        return
    r = get_redis()
    r.delete(_job_key(job_id))

GAME_IMPORT_HEADERS = [
    "Игра",
    "Платформа",
]

GAME_IMPORT_HEADER_MAP = {
    "Игра": "title",
    "Платформа": "platform_codes",
    # legacy/tech headers
    "title": "title",
    "platform_codes": "platform_codes",
}

def find_game_title_platform_conflicts(conn, title: str, platform_codes: List[str], exclude_game_id: Optional[int] = None) -> List[str]:
    if not title or not platform_codes:
        return []
    normalized_codes = normalize_platform_codes(platform_codes)
    params = [title, normalized_codes]
    extra = ""
    if exclude_game_id is not None:
        extra = "AND g.game_id <> %s"
        params.append(exclude_game_id)
    rows = qall(
        conn,
        f"""
        SELECT DISTINCT p.code
        FROM app.game_titles g
        JOIN app.game_platforms gp ON gp.game_id = g.game_id
        JOIN app.platforms p ON p.platform_id = gp.platform_id
        WHERE lower(g.title) = lower(%s)
          AND g.is_archived IS NOT TRUE
          AND lower(p.code) = ANY(%s)
          {extra}
        ORDER BY p.code
        """,
        params,
    )
    return [r[0] for r in rows]

def find_game_id_by_title_platforms(conn, title: str, platform_codes: List[str]) -> Optional[int]:
    if not title or not platform_codes:
        return None
    normalized_codes = normalize_platform_codes(platform_codes)
    target_count = len(normalized_codes)
    row = q1(
        conn,
        """
        SELECT g.game_id
        FROM app.game_titles g
        JOIN app.game_platforms gp ON gp.game_id = g.game_id
        JOIN app.platforms p ON p.platform_id = gp.platform_id
        WHERE lower(g.title) = lower(%s)
          AND g.is_archived IS NOT TRUE
        GROUP BY g.game_id
        HAVING count(DISTINCT lower(p.code)) = %s
           AND bool_and(lower(p.code) = ANY(%s))
        ORDER BY g.game_id
        LIMIT 1
        """,
        (title, target_count, normalized_codes),
    )
    return int(row[0]) if row else None

def parse_import_platforms(value: str) -> List[str]:
    if not value:
        return []
    raw = str(value).strip()
    if raw in {"-", "—"}:
        return []
    parts = []
    for chunk in raw.replace(";", ",").split(","):
        val = chunk.strip().lower()
        if val:
            parts.append(val)
    return normalize_platform_codes(parts)

def fetch_logo_from_url(url: str) -> Tuple[bytes, str]:
    if not url:
        raise ValueError("logo url is empty")
    context = ssl.create_default_context()
    context.check_hostname = False
    context.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "gamesales-import"},
    )
    try:
        with urllib.request.urlopen(req, timeout=10, context=context) as resp:
            content_type = (resp.headers.get("Content-Type") or "").split(";")[0].strip().lower()
            data = resp.read(MAX_LOGO_BYTES + 1)
    except urllib.error.URLError as exc:
        raise ValueError(f"logo download failed: {exc}") from exc
    if len(data) > MAX_LOGO_BYTES:
        raise ValueError("logo too large")
    mime = content_type if content_type in ALLOWED_LOGO_MIME else ""
    if not mime:
        kind = imghdr.what(None, data)
        if kind == "jpeg":
            mime = "image/jpeg"
        elif kind == "png":
            mime = "image/png"
        elif kind == "webp":
            mime = "image/webp"
    if mime not in ALLOWED_LOGO_MIME:
        raise ValueError("logo type not allowed")
    return data, mime

def validate_game_import_rows(conn, rows: List[dict], progress_cb=None, check_logo=False) -> Tuple[List[dict], List[dict]]:
    errors = []
    warnings = []
    platform_rows = qall(conn, "SELECT code FROM app.platforms")
    platforms = {str(r[0]).strip().lower() for r in platform_rows}
    for idx, row in enumerate(rows, start=2):
        report_row = idx - 1
        title = (row.get("title") or "").strip()
        platform_raw = row.get("platform_codes") or ""
        platform_codes = parse_import_platforms(platform_raw)
        if not title:
            errors.append({"row": report_row, "field": "Игра", "value": title, "message": "Название обязательно"})
        if not platform_codes:
            warnings.append({"row": report_row, "field": "Платформа", "value": str(platform_raw).strip(), "message": "Платформы не указаны — строка будет пропущена"})
        for code in platform_codes:
            if code not in platforms:
                errors.append({"row": report_row, "field": "Платформа", "value": code, "message": f"Неизвестная платформа: {code}"})
        if progress_cb:
            progress_cb(idx - 1)
    return errors, warnings

def read_games_from_excel(content: bytes) -> List[dict]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        if cell.value is None:
            headers.append("")
        else:
            raw = str(cell.value).strip()
            headers.append(GAME_IMPORT_HEADER_MAP.get(raw, raw))
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        item = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            item[key] = row[idx] if idx < len(row) else None
        data.append(item)
    return data

def read_accounts_from_excel(content: bytes) -> List[dict]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        if cell.value is None:
            headers.append("")
        else:
            raw = str(cell.value).strip()
            headers.append(ACCOUNT_IMPORT_HEADER_MAP.get(raw, raw))
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        item = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            item[key] = row[idx] if idx < len(row) else None
        data.append(item)
    return data

def read_slots_from_excel(content: bytes) -> List[dict]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        if cell.value is None:
            headers.append("")
        else:
            raw = str(cell.value).strip()
            headers.append(SLOT_IMPORT_HEADER_MAP.get(raw, raw))
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        item = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            item[key] = row[idx] if idx < len(row) else None
        data.append(item)
    return data

def clean_slots_excel(content: bytes) -> bytes:
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    status_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        header = str(cell.value).strip()
        if header.lower() == "статус":
            status_col = idx
            break
    if not status_col:
        raise HTTPException(400, 'Колонка "Статус" не найдена')
    for row_idx in range(ws.max_row, 1, -1):
        raw = ws.cell(row=row_idx, column=status_col).value
        value = "" if raw is None else str(raw).strip()
        if not value or value.lower() == "свободен":
            ws.delete_rows(row_idx, 1)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()

def split_account(value: str) -> Tuple[Optional[str], Optional[str]]:
    if not value:
        return None, None
    raw = str(value).strip()
    if not raw or "@" not in raw:
        return None, None
    login, domain = raw.split("@", 1)
    login = login.strip()
    domain = domain.strip()
    if not login or not domain:
        return None, None
    return login, domain

def validate_account_import_rows(conn, rows: List[dict], progress_cb=None) -> Tuple[List[dict], List[dict]]:
    errors = []
    warnings = []
    for idx, row in enumerate(rows, start=2):
        report_row = idx - 1
        account_val = (row.get("account") or "").strip()
        password = (row.get("password") or "").strip()
        game_title = (row.get("game") or "").strip()
        login, domain = split_account(account_val)
        if not account_val or not login or not domain:
            warnings.append({"row": report_row, "field": "Аккаунт", "value": account_val, "message": "Нужно значение в формате login@domain — строка будет пропущена"})
        # Пароль не обязателен при импорте аккаунтов
        if not game_title:
            warnings.append({"row": report_row, "field": "Игра", "value": game_title, "message": "Игра не указана — строка будет пропущена"})
        else:
            row_game = q1(conn, "SELECT game_id FROM app.game_titles WHERE lower(title)=lower(%s)", (game_title,))
            if not row_game:
                warnings.append({"row": report_row, "field": "Игра", "value": game_title, "message": "Не найдена в списке игр — строка будет пропущена"})
        if progress_cb:
            progress_cb(idx - 1)
    return errors, warnings

SLOT_FILE_TO_TYPE = {
    "п4": ("activate_ps4", 1),
    "п5": ("activate_ps5", 1),
    "ps4(1)": ("play_ps4", 1),
    "ps4(2)": ("play_ps4", 2),
    "ps5(1)": ("play_ps5", 1),
    "ps5(2)": ("play_ps5", 2),
}

def normalize_slot_file_value(value: str) -> str:
    return str(value or "").strip().lower()

def normalize_cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()

RU_MONTHS = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}

def normalize_slot_date(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = str(value).strip()
    if not raw:
        return None
    m = re.search(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", raw)
    if m:
        d = int(m.group(1))
        mo = int(m.group(2))
        y = int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            return None
    m = re.match(r"^\s*([A-Za-zА-Яа-яёЁ]+)\s+(\d{2,4})\s*$", raw)
    if m:
        month_name = m.group(1).strip().lower()
        year = int(m.group(2))
        if year < 100:
            year += 2000
        month = RU_MONTHS.get(month_name)
        if month:
            try:
                return date(year, month, 1).isoformat()
            except ValueError:
                return None
    return None

def normalize_slot_date_to_dt(value) -> Optional[datetime]:
    normalized = normalize_slot_date(value)
    if not normalized:
        return None
    try:
        return datetime.fromisoformat(normalized).replace(tzinfo=timezone.utc)
    except ValueError:
        return None

def resolve_source_id(conn, source_val: str) -> Optional[int]:
    if not source_val:
        return None
    parts = [p for p in str(source_val).strip().split() if p]
    row = None
    if len(parts) >= 2:
        name_part = parts[0]
        code_part = parts[-1]
        row = q1(
            conn,
            """
            SELECT source_id
            FROM app.sources
            WHERE lower(name) = lower(%s) AND lower(code) = lower(%s)
            """,
            (name_part, code_part),
        )
        if not row:
            row = q1(
                conn,
                """
                SELECT source_id
                FROM app.sources
                WHERE lower(name) = lower(%s) AND lower(code) = lower(%s)
                """,
                (code_part, name_part),
            )
    if not row:
        row = q1(
            conn,
            """
            SELECT source_id
            FROM app.sources
            WHERE lower(code) = lower(%s) OR lower(name) = lower(%s)
            """,
            (source_val, source_val),
        )
    return int(row[0]) if row else None

def validate_slot_import_rows(conn, rows: List[dict], progress_cb=None) -> Tuple[List[dict], List[dict], int]:
    errors = []
    warnings = []
    total = 0

    slot_rows = qall(conn, "SELECT code FROM app.slot_types")
    slot_types = {str(r[0]).strip().lower() for r in slot_rows}

    for idx, row in enumerate(rows, start=2):
        report_row = idx - 1
        if progress_cb:
            progress_cb(idx - 1)
        status_raw = row.get("status")
        status = "" if status_raw is None else str(status_raw).strip().lower()
        if not status or status == "свободен":
            continue

        total += 1
        account_val = normalize_cell_text(row.get("account"))
        slot_val = normalize_cell_text(row.get("slot"))
        game_title = normalize_cell_text(row.get("game"))
        customer = normalize_cell_text(row.get("customer"))
        source_val = normalize_cell_text(row.get("source"))
        date_val = row.get("date")

        if not account_val:
            errors.append({"row": report_row, "field": "Аккаунт", "value": account_val, "message": "Аккаунт обязателен"})
        else:
            login, domain = split_account(account_val)
            if not login or not domain:
                warnings.append({"row": report_row, "field": "Аккаунт", "value": account_val, "message": "Нужно значение в формате login@domain"})
            else:
                row_acc = q1(
                    conn,
                    """
                    SELECT 1
                    FROM app.accounts a
                    JOIN app.domains d ON d.domain_id = a.domain_id
                    WHERE lower(a.login_name) = lower(%s) AND lower(d.name) = lower(%s)
                    """,
                    (login, domain),
                )
                if not row_acc:
                    warnings.append({"row": report_row, "field": "Аккаунт", "value": account_val, "message": "Аккаунт не найден"})

        if not slot_val:
            errors.append({"row": report_row, "field": "Слот", "value": slot_val, "message": "Слот обязателен"})
        else:
            normalized_slot = normalize_slot_file_value(slot_val)
            slot_mapping = SLOT_FILE_TO_TYPE.get(normalized_slot)
            if not slot_mapping:
                warnings.append({"row": report_row, "field": "Слот", "value": slot_val, "message": "Неизвестный слот"})
            elif slot_mapping[0].lower() not in slot_types:
                warnings.append({"row": report_row, "field": "Слот", "value": slot_val, "message": "Слот не найден в БД"})

        if game_title:
            row_game = q1(conn, "SELECT 1 FROM app.game_titles WHERE lower(title) = lower(%s)", (game_title,))
            if not row_game:
                warnings.append({"row": report_row, "field": "Игра", "value": game_title, "message": "Игра не найдена"})
        else:
            warnings.append({"row": report_row, "field": "Игра", "value": game_title, "message": "Игра не указана"})

        if not customer:
            warnings.append({"row": report_row, "field": "Пользователь", "value": customer, "message": "Пользователь не указан"})

        if source_val:
            parts = [p for p in str(source_val).strip().split() if p]
            row_source = None
            if len(parts) >= 2:
                name_part = parts[0]
                code_part = parts[-1]
                row_source = q1(
                    conn,
                    """
                    SELECT 1
                    FROM app.sources
                    WHERE lower(name) = lower(%s) AND lower(code) = lower(%s)
                    """,
                    (name_part, code_part),
                )
                if not row_source:
                    row_source = q1(
                        conn,
                        """
                        SELECT 1
                        FROM app.sources
                        WHERE lower(name) = lower(%s) AND lower(code) = lower(%s)
                        """,
                        (code_part, name_part),
                    )
            if not row_source:
                row_source = q1(
                    conn,
                    """
                    SELECT 1
                    FROM app.sources
                    WHERE lower(code) = lower(%s) OR lower(name) = lower(%s)
                    """,
                    (source_val, source_val),
                )
            if not row_source:
                warnings.append({"row": report_row, "field": "Откуда", "value": source_val, "message": "Источник не найден"})

        if date_val not in (None, ""):
            normalized_date = normalize_slot_date(date_val)
            original_date = normalize_cell_text(date_val)
            if normalized_date:
                if normalized_date != original_date:
                    warnings.append({"row": report_row, "field": "Дата", "value": original_date, "message": f"Будет сохранено как {normalized_date}"})
            else:
                warnings.append({"row": report_row, "field": "Дата", "value": original_date, "message": "Не удалось распознать дату"})

        if progress_cb:
            progress_cb(idx - 1)

    return errors, warnings, total

def init_auth_schema(conn):
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS app.user_roles (
              code text PRIMARY KEY,
              name text NOT NULL
            );
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS app.users (
              user_id bigserial PRIMARY KEY,
              username text NOT NULL UNIQUE,
              password_hash text NOT NULL,
              role_code text NOT NULL DEFAULT 'manager' REFERENCES app.user_roles(code),
              created_at timestamptz NOT NULL DEFAULT now()
            );
            """
        )
        cur.execute(
            """
            INSERT INTO app.user_roles(code, name)
            VALUES ('admin','Admin'),('manager','Manager')
            ON CONFLICT (code) DO NOTHING;
            """
        )
        conn.commit()

def get_user_by_username(conn, username: str):
    return q1(
        conn,
        "SELECT user_id, username, password_hash, role_code FROM app.users WHERE username=%s",
        (username,),
    )


def get_user_id(conn, username: str) -> int:
    row = get_user_by_username(conn, username)
    if not row:
        raise HTTPException(404, "User not found")
    return int(row[0])

def role_exists(conn, role_code: str) -> bool:
    row = q1(conn, "SELECT 1 FROM app.user_roles WHERE code=%s", (role_code,))
    return bool(row)

def ensure_admin_user(conn):
    admin_username = os.getenv("ADMIN_USERNAME")
    admin_password = os.getenv("ADMIN_PASSWORD")
    admin_role = os.getenv("ADMIN_ROLE", "admin")
    if not admin_username or not admin_password:
        return
    if not role_exists(conn, admin_role):
        exec1(
            conn,
            "INSERT INTO app.user_roles(code, name) VALUES (%s, %s) ON CONFLICT (code) DO NOTHING",
            (admin_role, admin_role.title()),
        )
    row = get_user_by_username(conn, admin_username)
    if row:
        return
    password_hash = pwd_context.hash(admin_password)
    exec1(
        conn,
        "INSERT INTO app.users(username, password_hash, role_code) VALUES (%s, %s, %s)",
        (admin_username, password_hash, admin_role),
    )
    conn.commit()

def create_access_token(user_id: int, username: str, role: str):
    if not JWT_SECRET:
        raise HTTPException(500, "JWT_SECRET is not set")
    exp = int((now_utc().timestamp()) + JWT_TTL_MIN * 60)
    payload = {"sub": str(user_id), "username": username, "role": role, "exp": exp}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)

def get_current_user(authorization: str = Header(None)) -> UserOut:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Missing bearer token")
    token = authorization.split(" ", 1)[1].strip()
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except Exception:
        raise HTTPException(401, "Invalid token")
    return UserOut(username=payload.get("username"), role=payload.get("role"))

def b64_encode(value: str) -> str:
    return base64.b64encode(value.encode("utf-8")).decode("utf-8")

def require_role(*roles):
    def _dep(user: UserOut = Depends(get_current_user)) -> UserOut:
        if user.role not in roles:
            raise HTTPException(403, "Forbidden")
        return user
    return _dep

# ----------------------------
# Endpoints
# ----------------------------
@app.get("/health")
def health():
    with psycopg.connect(DB_DSN) as conn:
        init_auth_schema(conn)
        ensure_admin_user(conn)
        v = q1(conn, "SELECT 1")
    return {"ok": True}

@app.post("/auth/login", response_model=LoginOut)
def login(payload: LoginIn):
    with psycopg.connect(DB_DSN) as conn:
        init_auth_schema(conn)
        ensure_admin_user(conn)
        row = get_user_by_username(conn, payload.username)
        if not row:
            raise HTTPException(401, "Invalid credentials")
        user_id, username, password_hash, role = row
        if not pwd_context.verify(payload.password, password_hash):
            raise HTTPException(401, "Invalid credentials")
        token = create_access_token(int(user_id), str(username), str(role))
        return LoginOut(access_token=token, user=UserOut(username=str(username), role=str(role)))

@app.get("/auth/me", response_model=UserOut)
def me(user: UserOut = Depends(get_current_user)):
    return user


@app.get("/tg/status")
def telegram_status(user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT phone, status FROM tg.shared_session WHERE id=1")
    if not row:
        return {"status": "not_connected"}
    return {"status": row[1], "phone": row[0]}


@app.post("/tg/auth/start")
def telegram_auth_start(payload: TelegramAuthStartIn, user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT session_string FROM tg.shared_session WHERE id=1")
        session_string = row[0] if row else ""
        resp = _telegram_api_request("POST", "/auth/start", {"phone": payload.phone, "session_string": session_string})
        phone_code_hash = resp.get("phone_code_hash")
        session_string = resp.get("session_string") or session_string
        exec1(
            conn,
            """
            INSERT INTO tg.shared_session(id, phone, session_string, phone_code_hash, status, updated_at, updated_by_user_id)
            VALUES (1, %s, %s, %s, 'pending', now(), %s)
            ON CONFLICT (id)
            DO UPDATE SET phone=excluded.phone,
                          session_string=excluded.session_string,
                          phone_code_hash=excluded.phone_code_hash,
                          status='pending',
                          updated_at=now(),
                          updated_by_user_id=excluded.updated_by_user_id
            """,
            (payload.phone, session_string, phone_code_hash, get_user_id(conn, user.username)),
        )
        conn.commit()
    return {"status": "code_sent"}


@app.post("/tg/auth/confirm")
def telegram_auth_confirm(payload: TelegramAuthConfirmIn, user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(
            conn,
            "SELECT phone, session_string, phone_code_hash, status FROM tg.shared_session WHERE id=1",
        )
        if not row:
            raise HTTPException(404, "Telegram session not found")
        phone, session_string, phone_code_hash, status = row
        if status not in ("pending", "password_required"):
            raise HTTPException(400, "Telegram auth not in progress")
        resp = _telegram_api_request(
            "POST",
            "/auth/confirm",
            {
                "phone": phone,
                "code": payload.code,
                "phone_code_hash": phone_code_hash,
                "session_string": session_string,
            },
        )
        new_session = resp.get("session_string") or session_string
        if resp.get("status") == "password_required":
            exec1(
                conn,
                """
                UPDATE tg.shared_session
                SET session_string=%s, status='password_required', updated_at=now(), updated_by_user_id=%s
                WHERE id=1
                """,
                (new_session, get_user_id(conn, user.username)),
            )
            conn.commit()
            return {"status": "password_required"}
        exec1(
            conn,
            """
            UPDATE tg.shared_session
            SET session_string=%s, status='ready', phone_code_hash=NULL, updated_at=now(), updated_by_user_id=%s
            WHERE id=1
            """,
            (new_session, get_user_id(conn, user.username)),
        )
        conn.commit()
    return {"status": "ready"}


@app.post("/tg/auth/password")
def telegram_auth_password(payload: TelegramAuthPasswordIn, user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(
            conn,
            "SELECT session_string, status FROM tg.shared_session WHERE id=1",
        )
        if not row:
            raise HTTPException(404, "Telegram session not found")
        session_string, status = row
        if status != "password_required":
            raise HTTPException(400, "Password not required")
        resp = _telegram_api_request(
            "POST",
            "/auth/password",
            {"password": payload.password, "session_string": session_string},
        )
        new_session = resp.get("session_string") or session_string
        exec1(
            conn,
            """
            UPDATE tg.shared_session
            SET session_string=%s, status='ready', phone_code_hash=NULL, updated_at=now(), updated_by_user_id=%s
            WHERE id=1
            """,
            (new_session, get_user_id(conn, user.username)),
        )
        conn.commit()
    return {"status": "ready"}

@app.post("/tg/auth/disconnect")
def telegram_auth_disconnect(user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        exec1(
            conn,
            """
            UPDATE tg.shared_session
            SET phone='',
                session_string='',
                phone_code_hash=NULL,
                status='pending',
                updated_at=now(),
                updated_by_user_id=%s
            WHERE id=1
            """,
            (get_user_id(conn, user.username),),
        )
        conn.commit()
    return {"status": "disconnected"}


@app.get("/tg/dialogs", response_model=TelegramDialogsOut)
def telegram_dialogs(status: Optional[str] = None, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT session_string, status FROM tg.shared_session WHERE id=1")
        if not row or row[1] != "ready":
            raise HTTPException(400, "Telegram is not connected")
        session_string = row[0]
        user_id = get_user_id(conn, user.username)
        rows = qall(
            conn,
            """
            SELECT
              s.chat_id,
              s.title,
              s.unread_count,
              s.is_group,
              s.is_channel,
              COALESCE(ds.status, 'new') AS status
            FROM tg.dialog_snapshot s
            LEFT JOIN tg.dialog_states ds ON ds.chat_id = s.chat_id
            ORDER BY s.updated_at DESC, s.chat_id DESC
            """,
        )
        items = [
            {
                "id": int(r[0]),
                "title": str(r[1] or ""),
                "unread_count": int(r[2] or 0),
                "is_group": bool(r[3]),
                "is_channel": bool(r[4]),
                "status": str(r[5] or "new"),
            }
            for r in rows
        ]
        last_sync_row = q1(conn, "SELECT MAX(updated_at) FROM tg.dialog_snapshot")
        last_sync_at = last_sync_row[0] if last_sync_row else None

        # Fast bootstrap so first screen is not empty on fresh DB.
        if not items:
            resp = _telegram_api_request(
                "POST",
                "/dialogs",
                {"session_string": session_string, "limit": 50, "only_private": True},
                timeout_sec=30,
            )
            bootstrap_items = resp.get("items", []) or []
            if bootstrap_items:
                _upsert_telegram_dialog_snapshot(conn, user_id, bootstrap_items)
                exec1(conn, "UPDATE tg.shared_session SET last_used_at=now() WHERE id=1")
                conn.commit()
                rows = qall(
                    conn,
                    """
                    SELECT
                      s.chat_id,
                      s.title,
                      s.unread_count,
                      s.is_group,
                      s.is_channel,
                      COALESCE(ds.status, 'new') AS status
                    FROM tg.dialog_snapshot s
                    LEFT JOIN tg.dialog_states ds ON ds.chat_id = s.chat_id
                    ORDER BY s.updated_at DESC, s.chat_id DESC
                    """,
                )
                items = [
                    {
                        "id": int(r[0]),
                        "title": str(r[1] or ""),
                        "unread_count": int(r[2] or 0),
                        "is_group": bool(r[3]),
                        "is_channel": bool(r[4]),
                        "status": str(r[5] or "new"),
                    }
                    for r in rows
                ]
                last_sync_row = q1(conn, "SELECT MAX(updated_at) FROM tg.dialog_snapshot")
                last_sync_at = last_sync_row[0] if last_sync_row else None

        _trigger_telegram_dialogs_sync(session_string, user_id)
        with _TELEGRAM_DIALOGS_SYNC_LOCK:
            sync_running = bool(_TELEGRAM_DIALOGS_SYNC_RUNNING)
            sync_loaded = int(_TELEGRAM_DIALOGS_SYNC_LOADED)
            sync_batches = int(_TELEGRAM_DIALOGS_SYNC_BATCHES)
    # Tab counters should show unread messages count, not contacts count.
    # Keep "all" as contacts count for the "Контактов" badge in UI.
    counts = {"new": 0, "accepted": 0, "archived": 0, "all": len(items)}
    for item in items:
        key = str(item.get("status") or "new")
        unread = int(item.get("unread_count") or 0)
        if key in ("new", "accepted", "archived"):
            counts[key] += unread
    if status in ("new", "accepted", "archived"):
        items = [i for i in items if str(i.get("status") or "new") == status]
    return {
        "items": items,
        "counts": counts,
        "sync_running": sync_running,
        "last_sync_at": last_sync_at,
        "sync_loaded": sync_loaded,
        "sync_batches": sync_batches,
    }

@app.put("/tg/dialogs/{chat_id}/status")
def telegram_dialog_status_set(chat_id: int, payload: TelegramDialogStatusIn, user: UserOut = Depends(get_current_user)):
    status = (payload.status or "").strip().lower()
    if status not in ("new", "accepted", "archived"):
        raise HTTPException(400, "Invalid status")
    with psycopg.connect(DB_DSN) as conn:
        user_id = get_user_id(conn, user.username)
        exec1(
            conn,
            """
            INSERT INTO tg.dialog_states(chat_id, status, updated_at, updated_by_user_id)
            VALUES (%s, %s, now(), %s)
            ON CONFLICT (chat_id)
            DO UPDATE SET status=excluded.status, updated_at=now(), updated_by_user_id=excluded.updated_by_user_id
            """,
            (chat_id, status, user_id),
        )
        conn.commit()
    return {"ok": True, "chat_id": chat_id, "status": status}

@app.get("/tg/contact")
def telegram_contact(sender_id: int, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(
            conn,
            "SELECT title, info FROM tg.contact_notes_shared WHERE sender_id=%s",
            (sender_id,),
        )
    if not row:
        return {"title": "", "info": ""}
    return {"title": row[0], "info": row[1]}

@app.put("/tg/contact")
def telegram_contact_upsert(payload: TelegramContactIn, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        user_id = get_user_id(conn, user.username)
        exec1(
            conn,
            """
            INSERT INTO tg.contact_notes_shared(sender_id, title, info, updated_at, updated_by_user_id)
            VALUES (%s, %s, %s, now(), %s)
            ON CONFLICT (sender_id)
            DO UPDATE SET title=excluded.title,
                          info=excluded.info,
                          updated_at=now(),
                          updated_by_user_id=excluded.updated_by_user_id
            """,
            (payload.sender_id, payload.title or "", payload.info or "", user_id),
        )
        conn.commit()
    return {"ok": True}


@app.get("/tg/messages", response_model=TelegramMessagesOut)
def telegram_messages(chat_id: int, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT session_string, status FROM tg.shared_session WHERE id=1")
        if not row or row[1] != "ready":
            raise HTTPException(400, "Telegram is not connected")
        session_string = row[0]
        try:
            resp = _telegram_api_request("POST", "/messages", {"session_string": session_string, "chat_id": chat_id, "limit": 100})
        except HTTPException as exc:
            if exc.status_code == 404:
                _delete_dead_dialog(conn, chat_id)
                conn.commit()
                raise HTTPException(404, "Chat is unavailable and was removed from the list")
            raise
        items = resp.get("items", [])
        message_ids = [int(i["id"]) for i in items if i.get("id")]
        sent_map = {}
        if message_ids:
            rows = qall(
                conn,
                """
                SELECT message_id, sent_by_user_id
                FROM tg.sent_messages
                WHERE chat_id=%s AND message_id = ANY(%s)
                """,
                (chat_id, message_ids),
            )
            if rows:
                user_ids = [r[1] for r in rows]
                user_rows = qall(
                    conn,
                    "SELECT user_id, username FROM app.users WHERE user_id = ANY(%s)",
                    (user_ids,),
                )
                user_map = {r[0]: r[1] for r in user_rows}
                for msg_id, sent_by_user_id in rows:
                    sent_map[int(msg_id)] = user_map.get(sent_by_user_id, "")
        for item in items:
            sent_by = sent_map.get(int(item.get("id", 0)))
            if sent_by:
                item["sent_by"] = sent_by
        exec1(conn, "UPDATE tg.shared_session SET last_used_at=now() WHERE id=1")
        conn.commit()
    return {"items": items}


@app.post("/tg/messages")
def telegram_send_message(payload: TelegramSendMessageIn, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        user_id = get_user_id(conn, user.username)
        row = q1(conn, "SELECT session_string, status FROM tg.shared_session WHERE id=1")
        if not row or row[1] != "ready":
            raise HTTPException(400, "Telegram is not connected")
        session_string = row[0]
        resp = _telegram_api_request("POST", "/messages/send", {"session_string": session_string, "chat_id": payload.chat_id, "text": payload.text})
        message_id = resp.get("message_id")
        if message_id:
            exec1(
                conn,
                """
                INSERT INTO tg.sent_messages(message_id, chat_id, sent_by_user_id, sent_at)
                VALUES (%s, %s, %s, now())
                ON CONFLICT (message_id, chat_id) DO NOTHING
                """,
                (message_id, payload.chat_id, user_id),
            )
        exec1(
            conn,
            """
            INSERT INTO tg.dialog_states(chat_id, status, updated_at, updated_by_user_id)
            VALUES (%s, 'accepted', now(), %s)
            ON CONFLICT (chat_id)
            DO UPDATE SET status='accepted', updated_at=now(), updated_by_user_id=excluded.updated_by_user_id
            """,
            (payload.chat_id, user_id),
        )
        exec1(conn, "UPDATE tg.shared_session SET last_used_at=now() WHERE id=1")
        conn.commit()
    return {"ok": True}

@app.get("/tg/media")
def telegram_media(chat_id: int, message_id: int, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT session_string, status FROM tg.shared_session WHERE id=1")
        if not row or row[1] != "ready":
            raise HTTPException(400, "Telegram is not connected")
        session_string = row[0]
        data, content_type = _telegram_api_request_raw(
            "POST",
            "/media",
            {"session_string": session_string, "chat_id": chat_id, "message_id": message_id},
        )
        exec1(conn, "UPDATE tg.shared_session SET last_used_at=now() WHERE id=1")
        conn.commit()
    return Response(content=data, media_type=content_type or "application/octet-stream")

@app.get("/platforms", response_model=List[PlatformOut])
def list_platforms(user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        rows = qall(conn, "SELECT code, name, slot_capacity FROM app.platforms WHERE is_archived IS NOT TRUE ORDER BY code")
    return [PlatformOut(code=r0, name=r1, slot_capacity=int(r2 or 0)) for (r0, r1, r2) in rows]

@app.get("/regions", response_model=List[RegionOut])
def list_regions(user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        rows = qall(conn, "SELECT code, name, purchase_cost_rate FROM app.regions WHERE is_archived IS NOT TRUE ORDER BY code")
    return [RegionOut(code=r0, name=r1, purchase_cost_rate=float(r2 or 1.0)) for (r0, r1, r2) in rows]

@app.post("/platforms", response_model=PlatformOut)
def create_platform(payload: PlatformIn, user: UserOut = Depends(require_role("admin"))):
    code = (payload.code or "").strip().lower()
    name = (payload.name or "").strip()
    if not code or not name:
        raise HTTPException(400, "Platform code and name are required")
    with psycopg.connect(DB_DSN) as conn:
        exec1(
            conn,
            """
            INSERT INTO app.platforms(code, name, slot_capacity, is_archived)
            VALUES (%s, %s, %s, false)
            ON CONFLICT (code)
            DO UPDATE SET name=excluded.name, slot_capacity=excluded.slot_capacity, is_archived=false
            """,
            (code, name, payload.slot_capacity),
        )
        conn.commit()
    return PlatformOut(code=code, name=name, slot_capacity=payload.slot_capacity)

@app.put("/platforms/{code}", response_model=PlatformOut)
def update_platform(code: str, payload: PlatformUpdate, user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT name, slot_capacity FROM app.platforms WHERE code=%s AND is_archived IS NOT TRUE", (code,))
        if not row:
            raise HTTPException(404, "Platform not found")
        new_name = (payload.name or row[0]).strip()
        new_slots = payload.slot_capacity if payload.slot_capacity is not None else row[1]
        exec1(conn, "UPDATE app.platforms SET name=%s, slot_capacity=%s WHERE code=%s", (new_name, new_slots, code))
        conn.commit()
    return PlatformOut(code=code, name=new_name, slot_capacity=new_slots)

@app.delete("/platforms/{code}")
def delete_platform(code: str, user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT 1 FROM app.platforms WHERE code=%s", (code,))
        if not row:
            raise HTTPException(404, "Platform not found")
        exec1(conn, "UPDATE app.platforms SET is_archived=true WHERE code=%s", (code,))
        conn.commit()
    return {"ok": True}

@app.post("/regions", response_model=RegionOut)
def create_region(payload: RegionIn, user: UserOut = Depends(require_role("admin"))):
    code = (payload.code or "").strip().upper()
    name = (payload.name or "").strip()
    rate = float(payload.purchase_cost_rate or 1.0)
    if not code or not name:
        raise HTTPException(400, "Region code and name are required")
    with psycopg.connect(DB_DSN) as conn:
        exec1(
            conn,
            """
            INSERT INTO app.regions(code, name, purchase_cost_rate, is_archived)
            VALUES (%s, %s, %s, false)
            ON CONFLICT (code)
            DO UPDATE SET name=excluded.name, purchase_cost_rate=excluded.purchase_cost_rate, is_archived=false
            """,
            (code, name, rate),
        )
        conn.commit()
    return RegionOut(code=code, name=name, purchase_cost_rate=rate)

@app.put("/regions/{code}", response_model=RegionOut)
def update_region(code: str, payload: RegionUpdate, user: UserOut = Depends(require_role("admin"))):
    name = (payload.name or "").strip() if payload.name is not None else None
    rate = payload.purchase_cost_rate
    if name is not None and not name:
        raise HTTPException(400, "Name is required")
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT name, purchase_cost_rate FROM app.regions WHERE code=%s AND is_archived IS NOT TRUE", (code,))
        if not row:
            raise HTTPException(404, "Region not found")
        new_name = name if name is not None else row[0]
        new_rate = float(rate) if rate is not None else float(row[1] or 1.0)
        exec1(conn, "UPDATE app.regions SET name=%s, purchase_cost_rate=%s WHERE code=%s", (new_name, new_rate, code))
        conn.commit()
    return RegionOut(code=code, name=new_name, purchase_cost_rate=new_rate)

@app.delete("/regions/{code}")
def delete_region(code: str, user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT 1 FROM app.regions WHERE code=%s", (code,))
        if not row:
            raise HTTPException(404, "Region not found")
        exec1(conn, "UPDATE app.regions SET is_archived=true WHERE code=%s", (code,))
        conn.commit()
    return {"ok": True}

@app.get("/domains", response_model=List[PlatformOut])
def list_domains(user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        rows = qall(conn, "SELECT name, name FROM app.domains WHERE is_archived IS NOT TRUE ORDER BY name")
    return [PlatformOut(code=r0, name=r1, slot_capacity=0) for (r0, r1) in rows]

@app.get("/sources", response_model=List[SourceOut])
def list_sources(user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        rows = qall(conn, "SELECT source_id, code, name FROM app.sources WHERE is_archived IS NOT TRUE ORDER BY source_id")
    return [SourceOut(source_id=int(r0), code=r1, name=r2) for (r0, r1, r2) in rows]

@app.post("/domains", response_model=PlatformOut)
def create_domain(payload: DomainIn, user: UserOut = Depends(require_role("admin"))):
    name = (payload.name or "").strip().lower()
    if not name:
        raise HTTPException(400, "Domain name is required")
    with psycopg.connect(DB_DSN) as conn:
        exec1(
            conn,
            """
            INSERT INTO app.domains(name, is_archived)
            VALUES (%s, false)
            ON CONFLICT (name)
            DO UPDATE SET is_archived=false
            """,
            (name,),
        )
        conn.commit()
    return PlatformOut(code=name, name=name, slot_capacity=0)

@app.put("/domains/{name}", response_model=PlatformOut)
def update_domain(name: str, payload: NameUpdate, user: UserOut = Depends(require_role("admin"))):
    new_name = (payload.name or "").strip().lower()
    if not new_name:
        raise HTTPException(400, "Name is required")
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT 1 FROM app.domains WHERE name=%s AND is_archived IS NOT TRUE", (name,))
        if not row:
            raise HTTPException(404, "Domain not found")
        exec1(conn, "UPDATE app.domains SET name=%s WHERE name=%s", (new_name, name))
        conn.commit()
    return PlatformOut(code=new_name, name=new_name, slot_capacity=0)

@app.delete("/domains/{name}")
def delete_domain(name: str, user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT 1 FROM app.domains WHERE name=%s", (name,))
        if not row:
            raise HTTPException(404, "Domain not found")
        exec1(conn, "UPDATE app.domains SET is_archived=true WHERE name=%s", (name,))
        conn.commit()
    return {"ok": True}

@app.post("/sources", response_model=SourceOut)
def create_source(payload: SourceIn, user: UserOut = Depends(require_role("admin"))):
    code = (payload.code or "").strip().lower()
    name = (payload.name or "").strip()
    if not code or not name:
        raise HTTPException(400, "Source code and name are required")
    with psycopg.connect(DB_DSN) as conn:
        row = q1(
            conn,
            """
            INSERT INTO app.sources(code, name, is_archived)
            VALUES (%s, %s, false)
            RETURNING source_id, code, name
            """,
            (code, name),
        )
        conn.commit()
    if not row:
        raise HTTPException(500, "Failed to create source")
    return SourceOut(source_id=int(row[0]), code=row[1], name=row[2])

@app.put("/sources/{source_id}", response_model=SourceOut)
def update_source(source_id: int, payload: SourceUpdate, user: UserOut = Depends(require_role("admin"))):
    code = (payload.code or "").strip().lower() if payload.code is not None else None
    name = (payload.name or "").strip() if payload.name is not None else None
    if code is not None and not code:
        raise HTTPException(400, "Source code is required")
    if name is not None and not name:
        raise HTTPException(400, "Name is required")
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT code, name FROM app.sources WHERE source_id=%s AND is_archived IS NOT TRUE", (source_id,))
        if not row:
            raise HTTPException(404, "Source not found")
        new_code = code if code is not None else row[0]
        new_name = name if name is not None else row[1]
        exec1(conn, "UPDATE app.sources SET code=%s, name=%s WHERE source_id=%s", (new_code, new_name, source_id))
        conn.commit()
    return SourceOut(source_id=source_id, code=new_code, name=new_name)

@app.delete("/sources/{source_id}")
def delete_source(source_id: int, user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT 1 FROM app.sources WHERE source_id=%s", (source_id,))
        if not row:
            raise HTTPException(404, "Source not found")
        exec1(conn, "UPDATE app.sources SET is_archived=true WHERE source_id=%s", (source_id,))
        conn.commit()
    return {"ok": True}

@app.post("/auth/change-password")
def change_password(payload: ChangePasswordIn, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        row = get_user_by_username(conn, user.username)
        if not row:
            raise HTTPException(404, "User not found")
        user_id, username, password_hash, role = row
        if not pwd_context.verify(payload.current_password, password_hash):
            raise HTTPException(401, "Invalid credentials")
        new_hash = pwd_context.hash(payload.new_password)
        exec1(
            conn,
            "UPDATE app.users SET password_hash=%s WHERE user_id=%s",
            (new_hash, user_id),
        )
        conn.commit()
    return {"ok": True}

@app.get("/user-roles", response_model=List[RoleOut])
def list_roles(user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        rows = qall(conn, "SELECT code, name FROM app.user_roles ORDER BY code")
    return [RoleOut(code=r0, name=r1) for (r0, r1) in rows]

@app.get("/users", response_model=List[UserListOut])
def list_users(user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        rows = qall(conn, "SELECT username, role_code, created_at FROM app.users ORDER BY user_id")
    return [UserListOut(username=r0, role=r1, created_at=r2) for (r0, r1, r2) in rows]

@app.post("/users", response_model=UserOut)
def create_user(payload: UserCreate, user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        if not role_exists(conn, payload.role_code):
            raise HTTPException(400, "Unknown role")
        row = get_user_by_username(conn, payload.username)
        if row:
            raise HTTPException(409, "User already exists")
        password_hash = pwd_context.hash(payload.password)
        exec1(
            conn,
            "INSERT INTO app.users(username, password_hash, role_code) VALUES (%s, %s, %s)",
            (payload.username, password_hash, payload.role_code),
        )
        conn.commit()
    return UserOut(username=payload.username, role=payload.role_code)

@app.post("/users/{username}/password")
def reset_password(username: str, payload: ResetPasswordIn, user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        row = get_user_by_username(conn, username)
        if not row:
            raise HTTPException(404, "User not found")
        user_id = int(row[0])
        new_hash = pwd_context.hash(payload.new_password)
        exec1(
            conn,
            "UPDATE app.users SET password_hash=%s WHERE user_id=%s",
            (new_hash, user_id),
        )
        conn.commit()
    return {"ok": True}

@app.get("/accounts", response_model=AccountListOut)
def list_accounts(
    q: Optional[str] = None,
    login_q: Optional[str] = None,
    game_q: Optional[str] = None,
    region_q: Optional[str] = None,
    status_q: Optional[str] = None,
    slots_q: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    sort_key: str = "login",
    sort_dir: str = "asc",
    page: int = 1,
    page_size: int = 50,
    all: bool = False,
    user: UserOut = Depends(get_current_user),
):
    page = max(1, page)
    if all:
        page_size = 0
        offset = 0
    else:
        page_size = max(1, min(int(page_size or 50), 200))
        offset = (page - 1) * page_size

    sort_map = {
        "login": "login_name",
        "region": "region_code",
        "status": "status_code",
        "slots": "free_total",
        "games": "game_titles_text",
        "date": "account_date",
    }
    sort_col = sort_map.get(sort_key, "login_name")
    sort_dir = "desc" if str(sort_dir).lower() == "desc" else "asc"

    filters = []
    params = []
    if login_q:
        filters.append("(a.login_name ILIKE %s OR d.name ILIKE %s OR (a.login_name || '@' || d.name) ILIKE %s)")
        params.extend([f"%{login_q}%", f"%{login_q}%", f"%{login_q}%"])
    if q:
        filters.append("(a.login_name ILIKE %s OR d.name ILIKE %s OR (a.login_name || '@' || d.name) ILIKE %s OR r.code ILIKE %s OR g.title ILIKE %s)")
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"])
    if game_q:
        filters.append("g.title ILIKE %s")
        params.append(f"%{game_q}%")
    if region_q:
        filters.append("r.code ILIKE %s")
        params.append(f"%{region_q}%")
    if status_q:
        filters.append("a.status_code ILIKE %s")
        params.append(f"%{status_q}%")
    else:
        filters.append("a.status_code <> 'archived'")
    if date_from:
        filters.append("a.account_date >= %s")
        params.append(date_from)
    if date_to:
        filters.append("a.account_date <= %s")
        params.append(date_to)

    where_sql = f"WHERE {' AND '.join(filters)}" if filters else ""

    with psycopg.connect(DB_DSN) as conn:
        rows = qall(
            conn,
            f"""
            WITH account_platforms AS (
              SELECT
                aa.account_id,
                BOOL_OR(p.code = 'ps4') AS has_ps4,
                COALESCE(array_agg(DISTINCT p.code ORDER BY p.code) FILTER (WHERE p.code IS NOT NULL), '{{}}'::text[]) AS platform_codes
              FROM app.account_assets aa
              JOIN app.game_platforms gp ON gp.game_id = aa.game_id
              JOIN app.platforms p ON p.platform_id = gp.platform_id
              WHERE aa.asset_type_code = 'game'
              GROUP BY aa.account_id
            ),
            base AS (
              SELECT
                a.account_id,
                a.region_id,
                a.status_code,
                a.login_name,
                a.domain_id,
                a.account_date,
                a.notes,
                r.code as region_code,
                d.name as domain_name,
                COALESCE(array_agg(DISTINCT g.title ORDER BY g.title) FILTER (WHERE g.title IS NOT NULL), '{{}}'::text[]) as game_titles,
                COALESCE(string_agg(DISTINCT g.title, ' · ' ORDER BY g.title), '') as game_titles_text,
                ap.platform_codes,
                COALESCE(SUM(s.free), 0) as free_total,
                COALESCE(string_agg(st.code || ' ' || s.occupied || '/' || s.capacity, ' · ' ORDER BY st.code), '') as slots_text
              FROM app.accounts a
              LEFT JOIN app.regions r ON r.region_id = a.region_id
              LEFT JOIN app.domains d ON d.domain_id = a.domain_id
              LEFT JOIN account_platforms ap ON ap.account_id = a.account_id
              LEFT JOIN app.account_assets aa ON aa.account_id = a.account_id AND aa.asset_type_code = 'game'
              LEFT JOIN app.game_titles g ON g.game_id = aa.game_id
              LEFT JOIN app.v_account_slot_status s
                ON s.account_id = a.account_id
               AND (COALESCE(ap.has_ps4, false) OR s.platform_code = 'ps5')
              LEFT JOIN app.slot_types st ON st.code = s.slot_type_code
              {where_sql}
              GROUP BY a.account_id, a.region_id, a.status_code, a.login_name, a.domain_id, a.account_date, a.notes, r.code, d.name, ap.platform_codes
            ),
            filtered AS (
              SELECT * FROM base
              {"WHERE slots_text ILIKE %s" if slots_q else ""}
            ),
            page AS (
              SELECT
                filtered.*,
                COUNT(*) OVER() AS total_count
              FROM filtered
              ORDER BY {sort_col} {sort_dir}, account_id DESC
              {"" if all else "LIMIT %s OFFSET %s"}
            )
            SELECT
              page.account_id,
              page.region_code,
              page.status_code,
              page.login_name,
              page.domain_name,
              page.account_date,
              page.notes,
              page.game_titles,
              page.platform_codes,
              s.slot_type_code,
              s.platform_code,
              s.mode,
              s.capacity,
              s.occupied,
              s.free,
              page.total_count
            FROM page
            LEFT JOIN app.v_account_slot_status s ON s.account_id = page.account_id
            ORDER BY {sort_col} {sort_dir}, page.account_id DESC, s.slot_type_code
            """,
            params + ([f"%{slots_q}%"] if slots_q else []) + ([] if all else [page_size, offset]),
        )

    acc_map = {}
    acc_list = []
    total = 0
    for row in rows:
        if not total:
            total = int(row[15] or 0)
        account_id = row[0]
        if account_id not in acc_map:
            acc = AccountOut(
                account_id=account_id,
                region_code=row[1],
                status=row[2],
                login_name=row[3],
                domain_code=row[4],
                login_full=f"{row[3]}@{row[4]}" if row[3] and row[4] else None,
                platform_slots=[],
                slot_status=[],
                game_titles=list(row[7] or []),
                platform_codes=list(row[8] or []),
                account_date=row[5],
                notes=row[6],
            )
            acc_map[account_id] = acc
            acc_list.append(acc)
        slot_type_code = row[9]
        if slot_type_code:
            acc_map[account_id].slot_status.append(
                AccountSlotStatusOut(
                    slot_type_code=slot_type_code,
                    platform_code=row[10],
                    mode=row[11],
                    capacity=int(row[12] or 0),
                    occupied=int(row[13] or 0),
                    free=int(row[14] or 0),
                )
            )
    return {"total": total, "items": acc_list}

@app.post("/accounts", response_model=AccountOut)
def create_account(payload: AccountCreate, user: UserOut = Depends(get_current_user)):
    if not payload.login_name or not payload.domain_code:
        raise HTTPException(400, "login_name and domain_code are required")
    validate_date_not_future(payload.account_date, "account_date")
    with psycopg.connect(DB_DSN) as conn:
        region_id = get_region_id(conn, payload.region_code)
        domain_id = get_domain_id(conn, payload.domain_code)

        row = q1(conn, """
            INSERT INTO app.accounts(login_name, domain_id, region_id, notes, account_date)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING account_id
        """, (payload.login_name, domain_id, region_id, payload.notes, payload.account_date))
        account_id = int(row[0])

        ps4_id, ps4_slots = get_platform_info(conn, "ps4")
        ps5_id, ps5_slots = get_platform_info(conn, "ps5")
        exec1(
            conn,
            """
            INSERT INTO app.account_platforms(account_id, platform_id, slot_capacity)
            VALUES (%s, %s, %s), (%s, %s, %s)
            ON CONFLICT (account_id, platform_id) DO UPDATE SET slot_capacity=excluded.slot_capacity
            """,
            (account_id, ps4_id, ps4_slots, account_id, ps5_id, ps5_slots),
        )
        conn.commit()

        platform_slots = get_account_platform_slots(conn, account_id)
        slot_status = get_account_slot_status(conn, account_id)
        return AccountOut(
            account_id=account_id,
            region_code=payload.region_code,
            status="active",
            login_name=payload.login_name,
            domain_code=payload.domain_code,
            login_full=f"{payload.login_name}@{payload.domain_code}" if payload.login_name and payload.domain_code else None,
            platform_slots=platform_slots,
            slot_status=slot_status,
            game_titles=None,
            account_date=payload.account_date,
            notes=payload.notes
        )

@app.put("/accounts/{account_id}", response_model=AccountOut)
def update_account(
    account_id: int,
    payload: AccountUpdate,
    user: UserOut = Depends(require_role("admin")),
):
    validate_date_not_future(payload.account_date, "account_date")
    with psycopg.connect(DB_DSN) as conn:
        current = q1(
            conn,
            """
            SELECT login_name, domain_id, region_id, status_code, account_date, notes
            FROM app.accounts
            WHERE account_id=%s
            """,
            (account_id,),
        )
        if not current:
            raise HTTPException(404, "Account not found")

        region_id = get_region_id(conn, payload.region_code) if payload.region_code else current[2]
        domain_id = get_domain_id(conn, payload.domain_code) if payload.domain_code else current[1]

        new_login = payload.login_name if payload.login_name is not None else current[0]
        new_status = payload.status_code if payload.status_code is not None else current[3]
        new_date = payload.account_date if payload.account_date is not None else current[4]
        new_notes = payload.notes if payload.notes is not None else current[5]

        exec1(
            conn,
            """
            UPDATE app.accounts
            SET login_name=%s,
                domain_id=%s,
                region_id=%s,
                status_code=%s,
                notes=%s,
                account_date=%s
            WHERE account_id=%s
            """,
            (
                new_login,
                domain_id,
                region_id,
                new_status,
                new_notes,
                new_date,
                account_id,
            ),
        )
        conn.commit()

        row = q1(
            conn,
            """
            SELECT
              a.account_id,
              r.code as region_code,
              a.status_code,
              a.login_name,
              d.name as domain_name,
              a.account_date,
              a.notes
            FROM app.accounts a
            LEFT JOIN app.regions r ON r.region_id = a.region_id
            LEFT JOIN app.domains d ON d.domain_id = a.domain_id
            WHERE a.account_id=%s
            """,
            (account_id,),
        )
        if not row:
            raise HTTPException(404, "Account not found")

        platform_slots = get_account_platform_slots(conn, row[0])
        slot_status = get_account_slot_status(conn, row[0])
        return AccountOut(
            account_id=row[0], region_code=row[1],
            status=row[2], login_name=row[3], domain_code=row[4],
            login_full=f"{row[3]}@{row[4]}" if row[3] and row[4] else None,
            platform_slots=platform_slots,
            slot_status=slot_status,
            game_titles=None,
            account_date=row[5],
            notes=row[6]
        )

@app.delete("/accounts/{account_id}")
def archive_account(account_id: int, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT 1 FROM app.accounts WHERE account_id=%s", (account_id,))
        if not row:
            raise HTTPException(404, "Account not found")
        exec1(conn, "UPDATE app.accounts SET status_code='archived' WHERE account_id=%s", (account_id,))
        conn.commit()
    return {"ok": True}

@app.get("/accounts/{account_id}/secrets", response_model=List[AccountSecretOut])
def list_account_secrets(account_id: int, user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        rows = qall(
            conn,
            """
            SELECT secret_key, secret_value, created_at
            FROM app.account_secrets
            WHERE account_id=%s
            ORDER BY secret_key
            """,
            (account_id,),
        )
    return [AccountSecretOut(secret_key=r0, secret_value_b64=r1, created_at=r2) for (r0, r1, r2) in rows]

@app.post("/accounts/secrets/batch", response_model=List[AccountSecretsBatchItem])
def list_account_secrets_batch(payload: AccountSecretsBatchIn, user: UserOut = Depends(require_role("admin"))):
    account_ids = list({int(a) for a in (payload.account_ids or [])})
    if not account_ids:
        return []
    with psycopg.connect(DB_DSN) as conn:
        rows = qall(
            conn,
            """
            SELECT account_id, secret_key, secret_value, created_at
            FROM app.account_secrets
            WHERE account_id = ANY(%s)
            ORDER BY account_id, secret_key
            """,
            (account_ids,),
        )
    out_map = {aid: [] for aid in account_ids}
    for account_id, secret_key, secret_value, created_at in rows:
        out_map.setdefault(account_id, []).append(
            AccountSecretOut(
                secret_key=secret_key,
                secret_value_b64=secret_value,
                created_at=created_at,
            )
        )
    return [AccountSecretsBatchItem(account_id=aid, secrets=out_map.get(aid, [])) for aid in account_ids]

@app.post("/accounts/{account_id}/secrets", response_model=AccountSecretOut)
def upsert_account_secret(
    account_id: int,
    payload: AccountSecretIn,
    user: UserOut = Depends(require_role("admin")),
):
    value_b64 = b64_encode(payload.secret_value)
    with psycopg.connect(DB_DSN) as conn:
        q1(
            conn,
            """
            INSERT INTO app.account_secrets(account_id, secret_key, secret_value)
            VALUES (%s, %s, %s)
            ON CONFLICT (account_id, secret_key)
            DO UPDATE SET secret_value=excluded.secret_value
            RETURNING secret_key, secret_value, created_at
            """,
            (account_id, payload.secret_key, value_b64),
        )
        conn.commit()
        row = q1(
            conn,
            "SELECT secret_key, secret_value, created_at FROM app.account_secrets WHERE account_id=%s AND secret_key=%s",
            (account_id, payload.secret_key),
        )
    return AccountSecretOut(secret_key=row[0], secret_value_b64=row[1], created_at=row[2])

@app.delete("/accounts/{account_id}/secrets/{secret_key}")
def delete_account_secret(
    account_id: int,
    secret_key: str,
    user: UserOut = Depends(require_role("admin")),
):
    with psycopg.connect(DB_DSN) as conn:
        exec1(
            conn,
            "DELETE FROM app.account_secrets WHERE account_id=%s AND secret_key=%s",
            (account_id, secret_key),
        )
        conn.commit()
    return {"ok": True}

@app.get("/accounts/{account_id}/games", response_model=List[GameOut])
def list_account_games(account_id: int, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        ensure_account_exists(conn, account_id)
        rows = qall(
            conn,
            """
            SELECT g.game_id, g.title, g.short_title, g.link, g.logo_url, g.text_lang, g.audio_lang, g.vr_support, r.code,
                   COALESCE(array_agg(p.code ORDER BY p.code) FILTER (WHERE p.code IS NOT NULL), '{}'::text[]) AS platform_codes
            FROM app.account_assets aa
            JOIN app.game_titles g ON g.game_id = aa.game_id
            LEFT JOIN app.regions r ON r.region_id = g.region_id
            LEFT JOIN app.game_platforms gp ON gp.game_id = g.game_id
            LEFT JOIN app.platforms p ON p.platform_id = gp.platform_id
            WHERE aa.account_id=%s AND aa.asset_type_code='game'
            GROUP BY g.game_id, g.title, g.short_title, g.link, g.logo_url, g.text_lang, g.audio_lang, g.vr_support, r.code
            ORDER BY g.title
            """,
            (account_id,),
        )
    return [
        GameOut(
            game_id=r0,
            title=r1,
            short_title=r2,
            link=r3,
            logo_url=r4,
            text_lang=r5,
            audio_lang=r6,
            vr_support=r7,
            platform_codes=list(r9 or []),
            region_code=r8,
        )
        for (r0, r1, r2, r3, r4, r5, r6, r7, r8, r9) in rows
    ]

@app.get("/accounts/{account_id}/slot-status", response_model=List[AccountSlotStatusOut])
def list_account_slot_status(account_id: int, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        ensure_account_exists(conn, account_id)
        rows = qall(
            conn,
            """
            SELECT s.slot_type_code, s.platform_code, s.mode, s.capacity, s.occupied, s.free
            FROM app.v_account_slot_status s
            LEFT JOIN (
              SELECT
                aa.account_id,
                BOOL_OR(p.code = 'ps4') AS has_ps4
              FROM app.account_assets aa
              JOIN app.game_platforms gp ON gp.game_id = aa.game_id
              JOIN app.platforms p ON p.platform_id = gp.platform_id
              WHERE aa.asset_type_code = 'game' AND aa.account_id = %s
              GROUP BY aa.account_id
            ) ap ON ap.account_id = s.account_id
            WHERE s.account_id=%s
              AND (COALESCE(ap.has_ps4, false) OR s.platform_code = 'ps5')
            ORDER BY slot_type_code
            """,
            (account_id, account_id),
        )
    return [
        AccountSlotStatusOut(
            slot_type_code=r[0],
            platform_code=r[1],
            mode=r[2],
            capacity=int(r[3] or 0),
            occupied=int(r[4] or 0),
            free=int(r[5] or 0),
        )
        for r in rows
    ]

@app.get("/accounts/{account_id}/slot-assignments", response_model=List[AccountSlotAssignmentOut])
def list_account_slot_assignments(account_id: int, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        ensure_account_exists(conn, account_id)
        rows = qall(
            conn,
            """
            SELECT
              asa.assignment_id,
              asa.account_id,
              a.login_name,
              d.name as domain_name,
              asa.slot_type_code,
              asa.customer_id,
              c.nickname,
              asa.game_id,
              g.title,
              asa.deal_id,
              asa.deal_item_id,
              asa.assigned_at,
              asa.released_at,
              asa.assigned_by,
              asa.released_by
            FROM app.account_slot_assignments asa
            LEFT JOIN app.accounts a ON a.account_id = asa.account_id
            LEFT JOIN app.domains d ON d.domain_id = a.domain_id
            LEFT JOIN app.customers c ON c.customer_id = asa.customer_id
            LEFT JOIN app.game_titles g ON g.game_id = asa.game_id
            WHERE asa.account_id=%s
            ORDER BY asa.released_at IS NULL DESC, asa.assigned_at DESC
            """,
            (account_id,),
        )
    return [
        AccountSlotAssignmentOut(
            assignment_id=r[0],
            account_id=r[1],
            account_login=f"{r[2]}@{r[3]}" if r[2] and r[3] else None,
            slot_type_code=r[4],
            customer_id=r[5],
            customer_nickname=r[6],
            game_id=r[7],
            game_title=r[8],
            deal_id=r[9],
            deal_item_id=r[10],
            assigned_at=r[11],
            released_at=r[12],
            assigned_by=r[13],
            released_by=r[14],
        )
        for r in rows
    ]

@app.get("/accounts/for-deal", response_model=List[AccountOut])
def list_accounts_for_deal(
    game_id: int,
    slot_type_code: Optional[str] = None,
    user: UserOut = Depends(get_current_user),
):
    slot_type_code = (slot_type_code or "").strip() or None
    with psycopg.connect(DB_DSN) as conn:
        rows = qall(
            conn,
            """
            WITH account_platforms AS (
              SELECT
                aa.account_id,
                BOOL_OR(p.code = 'ps4') AS has_ps4,
                COALESCE(array_agg(DISTINCT p.code ORDER BY p.code) FILTER (WHERE p.code IS NOT NULL), '{}'::text[]) AS platform_codes
              FROM app.account_assets aa
              JOIN app.game_platforms gp ON gp.game_id = aa.game_id
              JOIN app.platforms p ON p.platform_id = gp.platform_id
              WHERE aa.asset_type_code = 'game'
              GROUP BY aa.account_id
            ),
            base AS (
              SELECT
                a.account_id,
                a.region_id,
                a.status_code,
                a.login_name,
                a.domain_id,
                a.account_date,
                a.notes,
                r.code as region_code,
                d.name as domain_name,
                ap.platform_codes,
                COALESCE(ap.has_ps4, false) as has_ps4
              FROM app.accounts a
              LEFT JOIN app.regions r ON r.region_id = a.region_id
              LEFT JOIN app.domains d ON d.domain_id = a.domain_id
              LEFT JOIN account_platforms ap ON ap.account_id = a.account_id
              JOIN app.account_assets aa
                ON aa.account_id = a.account_id
               AND aa.asset_type_code = 'game'
               AND aa.game_id = %s
              WHERE a.status_code <> 'archived'
                AND EXISTS (
                SELECT 1
                FROM app.v_account_slot_status ss
                WHERE ss.account_id = a.account_id
                  AND ss.free > 0
                  AND (COALESCE(ap.has_ps4, false) OR ss.platform_code = 'ps5')
                  AND (%s::text IS NULL OR ss.slot_type_code = %s)
              )
            )
            SELECT
              base.account_id,
              base.region_code,
              base.status_code,
              base.login_name,
              base.domain_name,
              base.account_date,
              base.notes,
              base.platform_codes,
              p.code as platform_code,
              s.slot_capacity,
              s.occupied_slots,
              s.free_slots
            FROM base
            LEFT JOIN app.v_account_platform_slots s ON s.account_id = base.account_id
            LEFT JOIN app.platforms p ON p.platform_id = s.platform_id
            ORDER BY base.account_id DESC, p.code
            """,
            (game_id, slot_type_code, slot_type_code),
        )

    acc_map = {}
    acc_list = []
    for row in rows:
        account_id = row[0]
        if account_id not in acc_map:
            acc = AccountOut(
                account_id=account_id,
                region_code=row[1],
                status=row[2],
                login_name=row[3],
                domain_code=row[4],
                login_full=f"{row[3]}@{row[4]}" if row[3] and row[4] else None,
                platform_slots=[],
                platform_codes=list(row[7] or []),
                account_date=row[5],
                notes=row[6],
            )
            acc_map[account_id] = acc
            acc_list.append(acc)
        platform_code = row[8]
        if platform_code:
            acc_map[account_id].platform_slots.append(
                AccountPlatformSlots(
                    platform_code=platform_code,
                    slot_capacity=int(row[9] or 0),
                    occupied_slots=int(row[10] or 0),
                    free_slots=int(row[11] or 0),
                )
            )
    return acc_list

@app.get("/accounts/for-deal/availability", response_model=List[SlotAvailabilityOut])
def list_slot_availability_for_deal(
    game_id: int,
    user: UserOut = Depends(get_current_user),
):
    with psycopg.connect(DB_DSN) as conn:
        rows = qall(
            conn,
            """
            WITH account_platforms AS (
              SELECT
                aa.account_id,
                BOOL_OR(p.code = 'ps4') AS has_ps4
              FROM app.account_assets aa
              JOIN app.game_platforms gp ON gp.game_id = aa.game_id
              JOIN app.platforms p ON p.platform_id = gp.platform_id
              WHERE aa.asset_type_code = 'game'
              GROUP BY aa.account_id
            ),
            base_accounts AS (
              SELECT a.account_id
              FROM app.accounts a
              JOIN app.account_assets aa
                ON aa.account_id = a.account_id
               AND aa.asset_type_code = 'game'
               AND aa.game_id = %s
              WHERE a.status_code <> 'archived'
            )
            SELECT
              st.code AS slot_type_code,
              BOOL_OR(COALESCE(ss.free, 0) > 0) AS has_free
            FROM base_accounts ba
            LEFT JOIN account_platforms ap ON ap.account_id = ba.account_id
            JOIN app.slot_types st
              ON (COALESCE(ap.has_ps4, false) OR st.platform_code = 'ps5')
            LEFT JOIN app.v_account_slot_status ss
              ON ss.account_id = ba.account_id AND ss.slot_type_code = st.code
            GROUP BY st.code
            ORDER BY st.code
            """,
            (game_id,),
        )
    return [SlotAvailabilityOut(slot_type_code=r0, has_free=bool(r1)) for (r0, r1) in rows]

@app.put("/accounts/{account_id}/games")
def set_account_games(
    account_id: int,
    payload: AccountGamesIn,
    user: UserOut = Depends(require_role("admin")),
):
    with psycopg.connect(DB_DSN) as conn:
        ensure_account_exists(conn, account_id)
        game_ids = list({int(g) for g in (payload.game_ids or [])})
        if game_ids:
            rows = qall(conn, "SELECT game_id FROM app.game_titles WHERE game_id = ANY(%s)", (game_ids,))
            valid_ids = [int(r[0]) for r in rows]
        else:
            valid_ids = []

        exec1(
            conn,
            "DELETE FROM app.account_assets WHERE account_id=%s AND asset_type_code='game'",
            (account_id,),
        )
        if valid_ids:
            with conn.cursor() as cur:
                cur.executemany(
                    "INSERT INTO app.account_assets(account_id, game_id, asset_type_code) VALUES (%s, %s, 'game')",
                    [(account_id, gid) for gid in valid_ids],
                )
        conn.commit()
    return {"ok": True}

@app.get("/games/{game_id}/accounts", response_model=List[GameAccountOut])
def list_game_accounts(game_id: int, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        ensure_game_exists(conn, game_id)
        rows = qall(
            conn,
            """
            SELECT
              a.account_id,
              a.login_name,
              d.name as domain_name,
              p.code as platform_code,
              s.free_slots,
              s.occupied_slots
            FROM (
              SELECT DISTINCT ON (di.account_id, di.platform_id)
                di.account_id,
                di.platform_id
              FROM app.deal_items di
              WHERE di.game_id=%s
                AND di.account_id IS NOT NULL
                AND di.platform_id IS NOT NULL
              ORDER BY di.account_id DESC, di.platform_id
            ) dg
            JOIN app.accounts a ON a.account_id = dg.account_id
            LEFT JOIN app.domains d ON d.domain_id = a.domain_id
            JOIN app.platforms p ON p.platform_id = dg.platform_id
            LEFT JOIN app.v_account_platform_slots s ON s.account_id = dg.account_id AND s.platform_id = dg.platform_id
            ORDER BY a.account_id DESC, p.code
            """,
            (game_id,),
        )
    return [
        GameAccountOut(
            account_id=r0,
            login_full=f"{r1}@{r2}" if r1 and r2 else None,
            platform_code=r3,
            free_slots=int(r4 or 0),
            occupied_slots=int(r5 or 0),
        )
        for (r0, r1, r2, r3, r4, r5) in rows
    ]

@app.get("/games/{game_id}/slot-assignments", response_model=List[AccountSlotAssignmentOut])
def list_game_slot_assignments(game_id: int, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        ensure_game_exists(conn, game_id)
        rows = qall(
            conn,
            """
            SELECT
              asa.assignment_id,
              asa.account_id,
              a.login_name,
              d.name as domain_name,
              asa.slot_type_code,
              asa.customer_id,
              c.nickname,
              asa.game_id,
              g.title,
              asa.deal_id,
              asa.deal_item_id,
              asa.assigned_at,
              asa.released_at,
              asa.assigned_by,
              asa.released_by
            FROM app.account_slot_assignments asa
            LEFT JOIN app.accounts a ON a.account_id = asa.account_id
            LEFT JOIN app.domains d ON d.domain_id = a.domain_id
            LEFT JOIN app.customers c ON c.customer_id = asa.customer_id
            LEFT JOIN app.game_titles g ON g.game_id = asa.game_id
            WHERE asa.game_id=%s
            ORDER BY asa.released_at IS NULL DESC, asa.assigned_at DESC
            """,
            (game_id,),
        )
    return [
        AccountSlotAssignmentOut(
            assignment_id=r[0],
            account_id=r[1],
            account_login=f"{r[2]}@{r[3]}" if r[2] and r[3] else None,
            slot_type_code=r[4],
            customer_id=r[5],
            customer_nickname=r[6],
            game_id=r[7],
            game_title=r[8],
            deal_id=r[9],
            deal_item_id=r[10],
            assigned_at=r[11],
            released_at=r[12],
            assigned_by=r[13],
            released_by=r[14],
        )
        for r in rows
    ]

@app.post("/slot-assignments/{assignment_id}/release")
def release_slot_assignment_api(assignment_id: int, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(
            conn,
            "SELECT assignment_id FROM app.account_slot_assignments WHERE assignment_id=%s AND released_at IS NULL",
            (assignment_id,),
        )
        if not row:
            return {"ok": True}
        exec1(
            conn,
            "UPDATE app.account_slot_assignments SET released_at=now(), released_by=%s WHERE assignment_id=%s",
            (user.username, assignment_id),
        )
        conn.commit()
    return {"ok": True}

@app.get("/games", response_model=GameListOut)
def list_games(
    q: Optional[str] = None,
    platform_code: Optional[str] = None,
    region_code: Optional[str] = None,
    sort_key: str = "id",
    sort_dir: str = "desc",
    page: int = 1,
    page_size: int = 50,
    all: bool = False,
    user: UserOut = Depends(get_current_user),
):
    page = max(1, page)
    if all:
        page_size = 0
        offset = 0
    else:
        page_size = max(1, min(int(page_size or 50), 200))
        offset = (page - 1) * page_size

    sort_map = {
        "id": "game_id",
        "title": "title",
        "platform": "platform_sort",
        "region": "region_code",
    }
    sort_col = sort_map.get(sort_key, "g.game_id")
    sort_dir = "desc" if str(sort_dir).lower() == "desc" else "asc"

    filters = []
    params = []
    if q:
        filters.append("(g.title ILIKE %s OR g.short_title ILIKE %s)")
        params.extend([f"%{q}%", f"%{q}%"])
    if region_code:
        filters.append("r.code ILIKE %s")
        params.append(f"%{region_code}%")
    if platform_code:
        filters.append("""
            EXISTS (
              SELECT 1
              FROM app.game_platforms gp2
              JOIN app.platforms p2 ON p2.platform_id = gp2.platform_id
              WHERE gp2.game_id = g.game_id AND lower(p2.code) = lower(%s)
            )
        """)
        params.append(platform_code)
    filters.append("g.is_archived IS NOT TRUE")

    where_sql = f"WHERE {' AND '.join(filters)}" if filters else ""

    with psycopg.connect(DB_DSN) as conn:
        rows = qall(
            conn,
            f"""
            WITH base AS (
              SELECT
                g.game_id,
                g.title,
                g.short_title,
                g.link,
                g.logo_url,
                g.text_lang,
                g.audio_lang,
                g.vr_support,
                r.code as region_code,
                MIN(p.code) as platform_sort
              FROM app.game_titles g
              LEFT JOIN app.regions r ON r.region_id = g.region_id
              LEFT JOIN app.game_platforms gp ON gp.game_id = g.game_id
              LEFT JOIN app.platforms p ON p.platform_id = gp.platform_id
              {where_sql}
              GROUP BY g.game_id, g.title, g.short_title, g.link, g.logo_url, g.text_lang, g.audio_lang, g.vr_support, r.code
            ),
            total AS (
              SELECT COUNT(*) AS total_count FROM base
            ),
            page AS (
              SELECT * FROM base
              ORDER BY {sort_col} {sort_dir}, game_id DESC
              {"" if all else "LIMIT %s OFFSET %s"}
            )
            SELECT
              page.game_id,
              page.title,
              page.short_title,
              page.link,
              page.logo_url,
              page.text_lang,
              page.audio_lang,
              page.vr_support,
              page.region_code,
              page.platform_sort,
              COALESCE(array_agg(p2.code ORDER BY p2.code) FILTER (WHERE p2.code IS NOT NULL), '{{}}'::text[]) AS platform_codes,
              (SELECT total_count FROM total) as total_count
            FROM page
            LEFT JOIN app.game_platforms gp2 ON gp2.game_id = page.game_id
            LEFT JOIN app.platforms p2 ON p2.platform_id = gp2.platform_id
            GROUP BY page.game_id, page.title, page.short_title, page.link, page.logo_url, page.text_lang, page.audio_lang, page.vr_support, page.region_code, page.platform_sort
            ORDER BY {sort_col} {sort_dir}, page.game_id DESC
            """,
            params + ([] if all else [page_size, offset]),
        )

    items = []
    total = 0
    for (r0, r1, r2, r3, r4, r5, r6, r7, r8, _platform_sort, r9, r10) in rows:
        total = int(r10 or 0)
        items.append(
            GameOut(
                game_id=r0,
                title=r1,
                short_title=r2,
                link=r3,
                logo_url=r4,
                text_lang=r5,
                audio_lang=r6,
                vr_support=r7,
                platform_codes=list(r9 or []),
                region_code=r8,
            )
        )
    if not rows:
        with psycopg.connect(DB_DSN) as conn:
            total_row = q1(
                conn,
                f"""
                SELECT COUNT(*) FROM (
                  SELECT g.game_id
                  FROM app.game_titles g
                  LEFT JOIN app.regions r ON r.region_id = g.region_id
                  LEFT JOIN app.game_platforms gp ON gp.game_id = g.game_id
                  LEFT JOIN app.platforms p ON p.platform_id = gp.platform_id
                  {where_sql}
                  GROUP BY g.game_id
                ) t
                """,
                params,
            )
            total = int(total_row[0] or 0) if total_row else 0
    return {"total": total, "items": items}

@app.get("/slot-types", response_model=List[SlotTypeOut])
def list_slot_types(user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        rows = qall(
            conn,
            """
            SELECT code, name, platform_code, mode, capacity
            FROM app.slot_types
            ORDER BY
              CASE WHEN mode = 'activate' THEN 1 ELSE 2 END,
              platform_code,
              code
            """,
        )
    return [SlotTypeOut(code=r[0], name=r[1], platform_code=r[2], mode=r[3], capacity=int(r[4] or 0)) for r in rows]

@app.post("/games", response_model=GameOut)
def create_game(payload: GameCreate, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        platform_codes = normalize_platform_codes(payload.platform_codes)
        conflicts = find_game_title_platform_conflicts(conn, payload.title, platform_codes)
        if conflicts:
            raise HTTPException(409, f"Game already exists for platforms: {', '.join(conflicts)}")
        platform_ids = [get_platform_id(conn, code) for code in platform_codes]
        region_id = get_region_id(conn, payload.region_code)

        row = q1(conn, """
            INSERT INTO app.game_titles(title, short_title, link, logo_url, text_lang, audio_lang, vr_support, region_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING game_id
        """, (
            payload.title,
            payload.short_title,
            payload.link,
            payload.logo_url,
            payload.text_lang,
            payload.audio_lang,
            payload.vr_support,
            region_id,
        ))
        gid = int(row[0])
        if platform_ids:
            with conn.cursor() as cur:
                cur.executemany(
                    "INSERT INTO app.game_platforms(game_id, platform_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                    [(gid, pid) for pid in platform_ids],
                )
        conn.commit()

    return GameOut(
        game_id=gid,
        title=payload.title,
        short_title=payload.short_title,
        link=payload.link,
        logo_url=payload.logo_url,
        text_lang=payload.text_lang,
        audio_lang=payload.audio_lang,
        vr_support=payload.vr_support,
        platform_codes=platform_codes,
        region_code=payload.region_code,
    )

@app.put("/games/{game_id}", response_model=GameOut)
def update_game(game_id: int, payload: GameUpdate, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(
            conn,
            """
            SELECT g.title, g.short_title, g.link, g.logo_url, g.text_lang, g.audio_lang, g.vr_support, r.code
            FROM app.game_titles g
            LEFT JOIN app.regions r ON r.region_id = g.region_id
            WHERE g.game_id=%s
            """,
            (game_id,),
        )
        if not row:
            raise HTTPException(404, "Game not found")

        new_title = payload.title if payload.title is not None else row[0]
        if not new_title:
            raise HTTPException(400, "title is required")
        new_short = payload.short_title if payload.short_title is not None else row[1]
        new_link = payload.link if payload.link is not None else row[2]
        new_logo = payload.logo_url if payload.logo_url is not None else row[3]
        new_text_lang = payload.text_lang if payload.text_lang is not None else row[4]
        new_audio_lang = payload.audio_lang if payload.audio_lang is not None else row[5]
        new_vr_support = payload.vr_support if payload.vr_support is not None else row[6]
        platform_codes = normalize_platform_codes(payload.platform_codes)
        platforms_for_check = platform_codes if payload.platform_codes is not None else get_game_platform_codes(conn, game_id)
        conflicts = find_game_title_platform_conflicts(conn, new_title, platforms_for_check, exclude_game_id=game_id)
        if conflicts:
            raise HTTPException(409, f"Game already exists for platforms: {', '.join(conflicts)}")

        if payload.region_code is None:
            region_code = row[7]
        else:
            region_code = payload.region_code or None

        region_id = get_region_id(conn, region_code)

        exec1(
            conn,
            """
            UPDATE app.game_titles
            SET title=%s,
                short_title=%s,
                link=%s,
                logo_url=%s,
                text_lang=%s,
                audio_lang=%s,
                vr_support=%s,
                region_id=%s
            WHERE game_id=%s
            """,
            (new_title, new_short, new_link, new_logo, new_text_lang, new_audio_lang, new_vr_support, region_id, game_id),
        )
        if payload.platform_codes is not None:
            exec1(conn, "DELETE FROM app.game_platforms WHERE game_id=%s", (game_id,))
            if platform_codes:
                platform_ids = [get_platform_id(conn, code) for code in platform_codes]
                with conn.cursor() as cur:
                    cur.executemany(
                        "INSERT INTO app.game_platforms(game_id, platform_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                        [(game_id, pid) for pid in platform_ids],
                    )
        else:
            rows = qall(conn, """
                SELECT p.code
                FROM app.game_platforms gp
                JOIN app.platforms p ON p.platform_id = gp.platform_id
                WHERE gp.game_id=%s
                ORDER BY p.code
            """, (game_id,))
            platform_codes = [r[0] for r in rows]
        conn.commit()
    return GameOut(
        game_id=game_id,
        title=new_title,
        short_title=new_short,
        link=new_link,
        logo_url=new_logo,
        text_lang=new_text_lang,
        audio_lang=new_audio_lang,
        vr_support=new_vr_support,
        platform_codes=platform_codes,
        region_code=region_code,
    )

@app.delete("/games/{game_id}")
def archive_game(game_id: int, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(conn, "SELECT 1 FROM app.game_titles WHERE game_id=%s", (game_id,))
        if not row:
            raise HTTPException(404, "Game not found")
        exec1(conn, "UPDATE app.game_titles SET is_archived=true WHERE game_id=%s", (game_id,))
        conn.commit()
    return {"ok": True}

@app.get("/games/{game_id}/logo")
def get_game_logo(game_id: int, user: UserOut = Depends(get_current_user)):
    with psycopg.connect(DB_DSN) as conn:
        row = q1(
            conn,
            "SELECT logo_blob, logo_mime FROM app.game_titles WHERE game_id=%s",
            (game_id,),
        )
        if not row:
            raise HTTPException(404, "Game not found")
        blob, mime = row
        if not blob or not mime:
            raise HTTPException(404, "Logo not found")
    return {"mime": mime, "data_b64": encode_b64(blob)}

@app.post("/games/{game_id}/logo")
def upload_game_logo(game_id: int, file: UploadFile = File(...), user: UserOut = Depends(require_role("admin"))):
    if not file or not file.content_type:
        raise HTTPException(400, "file is required")
    if file.content_type not in ("image/jpeg", "image/png", "image/webp"):
        raise HTTPException(400, "Only jpg, png, webp are allowed")
    content = file.file.read()
    if content is None:
        raise HTTPException(400, "file is required")
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(400, "File too large. Max 5MB")
    with psycopg.connect(DB_DSN) as conn:
        ensure_game_exists(conn, game_id)
        exec1(
            conn,
            "UPDATE app.game_titles SET logo_blob=%s, logo_mime=%s WHERE game_id=%s",
            (content, file.content_type, game_id),
        )
        conn.commit()
    return {"ok": True}

@app.delete("/games/{game_id}/logo")
def delete_game_logo(game_id: int, user: UserOut = Depends(require_role("admin"))):
    with psycopg.connect(DB_DSN) as conn:
        ensure_game_exists(conn, game_id)
        exec1(conn, "UPDATE app.game_titles SET logo_blob=NULL, logo_mime=NULL WHERE game_id=%s", (game_id,))
        conn.commit()
    return {"ok": True}

def run_games_import_job(job_id: str, content: bytes, owner: str):
    try:
        with psycopg.connect(DB_DSN) as conn:
            rows = read_games_from_excel(content)
            errors, warnings = validate_game_import_rows(conn, rows)
            if errors:
                set_import_progress(job_id, owner, {
                    "phase": "validate",
                    "current": len(rows),
                    "total": len(rows),
                    "done": True,
                    "result": {"ok": False, "total": len(rows), "errors": errors, "warnings": warnings},
                })
                return
            if is_import_cancelled(job_id):
                set_import_progress(job_id, owner, {
                    "phase": "cancelled",
                    "current": 0,
                    "total": len(rows),
                    "done": True,
                    "cancelled": True,
                    "result": {"ok": False, "cancelled": True},
                })
                return
            total = len(rows)
            set_import_progress(job_id, owner, {"phase": "upload", "current": 0, "total": total, "done": False})
            created = 0
            updated = 0
            skipped = 0
            failed = 0
            last_success_row = None
            row_errors = []
            upload_done = 0
            existing_game_by_row = {}
            for idx, item in enumerate(rows, start=2):
                title = (item.get("title") or "").strip()
                platform_codes = parse_import_platforms(item.get("platform_codes") or "")
                existing_game_id = find_game_id_by_title_platforms(conn, title, platform_codes)
                if existing_game_id:
                    existing_game_by_row[idx] = existing_game_id

            for idx, item in enumerate(rows, start=2):
                if is_import_cancelled(job_id):
                    set_import_progress(job_id, owner, {
                        "phase": "cancelled",
                        "current": upload_done,
                        "total": total,
                        "done": True,
                        "cancelled": True,
                        "result": {"ok": False, "cancelled": True},
                    })
                    return
                try:
                    title = (item.get("title") or "").strip()
                    platform_codes = parse_import_platforms(item.get("platform_codes") or "")
                    if not platform_codes:
                        skipped += 1
                        upload_done += 1
                        set_import_progress(job_id, owner, {"phase": "upload", "current": upload_done, "total": total, "done": False})
                        continue
                    existing_game_id = existing_game_by_row.get(idx) or find_game_id_by_title_platforms(conn, title, platform_codes)
                    if existing_game_id:
                        game_id = existing_game_id
                        updated += 1
                    else:
                        row = q1(
                            conn,
                            "INSERT INTO app.game_titles(title) VALUES (%s) RETURNING game_id",
                            (title,),
                        )
                        game_id = int(row[0])
                        created += 1
                    if platform_codes:
                        platform_ids = [get_platform_id(conn, code) for code in platform_codes]
                        with conn.cursor() as cur:
                            cur.executemany(
                                "INSERT INTO app.game_platforms(game_id, platform_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                                [(game_id, pid) for pid in platform_ids],
                            )
                    conn.commit()
                    upload_done += 1
                    last_success_row = idx
                    set_import_progress(job_id, owner, {"phase": "upload", "current": upload_done, "total": total, "done": False})
                except Exception as exc:
                    conn.rollback()
                    failed += 1
                    row_errors.append({"row": idx, "field": "Импорт", "message": str(exc)})
                    upload_done += 1
                    set_import_progress(job_id, owner, {"phase": "upload", "current": upload_done, "total": total, "done": False})
                    continue

            result = {
                "ok": len(row_errors) == 0,
                "total": len(rows),
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "failed": failed,
                "errors": row_errors,
                "warnings": warnings,
                "success_until_row": last_success_row,
            }
            set_import_progress(job_id, owner, {"phase": "upload", "current": total, "total": total, "done": True, "result": result})
    except Exception as exc:
        set_import_progress(job_id, owner, {"phase": "error", "current": 0, "total": 0, "done": True, "result": {"ok": False, "errors": [{"row": 0, "field": "Импорт", "message": str(exc)}]}})

@app.get("/games/import/template")
def games_import_template(user: UserOut = Depends(require_role("admin"))):
    wb = Workbook()
    ws = wb.active
    ws.title = "games"
    ws.append(GAME_IMPORT_HEADERS)
    for col in range(1, len(GAME_IMPORT_HEADERS) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=games_import_template.xlsx"},
    )

@app.get("/games/import/status")
def games_import_status(job_id: str, user: UserOut = Depends(require_role("admin"))):
    status = get_import_progress(job_id)
    if not status:
        return {"phase": "idle", "current": 0, "total": 0, "done": True}
    if status.get("owner") and status.get("owner") != user.username:
        raise HTTPException(403, "job not found")
    return {
        "phase": status.get("phase", "idle"),
        "current": int(status.get("current") or 0),
        "total": int(status.get("total") or 0),
        "done": bool(status.get("done")),
        "result": status.get("result"),
    }

@app.post("/games/import/validate")
def games_import_validate(file: UploadFile = File(...), user: UserOut = Depends(require_role("admin"))):
    if not file or not file.filename:
        raise HTTPException(400, "file is required")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls are supported")
    content = file.file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(400, "File too large. Max 5MB")
    with psycopg.connect(DB_DSN) as conn:
        rows = read_games_from_excel(content)
        errors, warnings = validate_game_import_rows(conn, rows, progress_cb=None, check_logo=False)
    return {"ok": len(errors) == 0, "total": len(rows), "errors": errors, "warnings": warnings}

@app.post("/games/import")
def games_import(file: UploadFile = File(...), user: UserOut = Depends(require_role("admin"))):
    if not file or not file.filename:
        raise HTTPException(400, "file is required")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls are supported")
    content = file.file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(400, "File too large. Max 5MB")
    job_id = uuid.uuid4().hex
    set_import_progress(job_id, user.username, {"phase": "queued", "current": 0, "total": 0, "done": False})
    thread = threading.Thread(target=run_games_import_job, args=(job_id, content, user.username), daemon=True)
    thread.start()
    return {"ok": True, "job_id": job_id}

@app.post("/games/import/cancel")
def games_import_cancel(job_id: str, user: UserOut = Depends(require_role("admin"))):
    status = get_import_progress(job_id)
    if not status:
        return {"ok": True}
    if status.get("owner") and status.get("owner") != user.username:
        raise HTTPException(403, "job not found")
    set_import_progress(job_id, user.username, {
        "phase": "cancelled",
        "current": int(status.get("current") or 0),
        "total": int(status.get("total") or 0),
        "done": True,
        "cancelled": True,
        "result": {"ok": False, "cancelled": True},
    })
    return {"ok": True}

def get_or_create_domain_id(conn, domain_name: str) -> int:
    row = q1(conn, "SELECT domain_id FROM app.domains WHERE name=%s", (domain_name,))
    if row:
        return int(row[0])
    row = q1(conn, "INSERT INTO app.domains(name) VALUES (%s) RETURNING domain_id", (domain_name,))
    return int(row[0])

def ensure_account_platforms(conn, account_id: int):
    ps4_id, ps4_slots = get_platform_info(conn, "ps4")
    ps5_id, ps5_slots = get_platform_info(conn, "ps5")
    exec1(
        conn,
        """
        INSERT INTO app.account_platforms(account_id, platform_id, slot_capacity)
        VALUES (%s, %s, %s), (%s, %s, %s)
        ON CONFLICT (account_id, platform_id) DO UPDATE SET slot_capacity=excluded.slot_capacity
        """,
        (account_id, ps4_id, ps4_slots, account_id, ps5_id, ps5_slots),
    )

def run_accounts_import_job(job_id: str, content: bytes, owner: str):
    try:
        with psycopg.connect(DB_DSN) as conn:
            rows = read_accounts_from_excel(content)
            errors, warnings = validate_account_import_rows(conn, rows)
            if errors:
                set_import_progress(job_id, owner, {"phase": "error", "current": 0, "total": 0, "done": True, "result": {"ok": False, "errors": errors, "warnings": warnings}})
                return
            total = len(rows)
            set_import_progress(job_id, owner, {"phase": "upload", "current": 0, "total": total, "done": False})
            created = 0
            updated = 0
            skipped = 0
            failed = 0
            upload_done = 0
            last_success_row = None
            row_errors = []
            for idx, item in enumerate(rows, start=2):
                if is_import_cancelled(job_id):
                    set_import_progress(job_id, owner, {"phase": "cancelled", "current": upload_done, "total": total, "done": True, "result": {"ok": False, "cancelled": True}})
                    return
                try:
                    account_val = (item.get("account") or "").strip()
                    password = (item.get("password") or "").strip()
                    game_title = (item.get("game") or "").strip()
                    login, domain = split_account(account_val)
                    if not login or not domain or not game_title:
                        skipped += 1
                        upload_done += 1
                        set_import_progress(job_id, owner, {"phase": "upload", "current": upload_done, "total": total, "done": False})
                        continue
                    game_row = q1(conn, "SELECT game_id FROM app.game_titles WHERE lower(title)=lower(%s)", (game_title,))
                    if not game_row:
                        skipped += 1
                        upload_done += 1
                        set_import_progress(job_id, owner, {"phase": "upload", "current": upload_done, "total": total, "done": False})
                        continue
                    game_id = int(game_row[0])
                    domain_id = get_or_create_domain_id(conn, domain)
                    account_row = q1(
                        conn,
                        """
                        SELECT account_id
                        FROM app.accounts
                        WHERE login_name=%s AND domain_id=%s
                        """,
                        (login, domain_id),
                    )
                    if account_row:
                        account_id = int(account_row[0])
                        updated += 1
                    else:
                        row = q1(
                            conn,
                            """
                            INSERT INTO app.accounts(login_name, domain_id, status_code)
                            VALUES (%s, %s, 'active')
                            RETURNING account_id
                            """,
                            (login, domain_id),
                        )
                        account_id = int(row[0])
                        ensure_account_platforms(conn, account_id)
                        created += 1
                    if password:
                        exec1(
                            conn,
                            """
                            INSERT INTO app.account_secrets(account_id, secret_key, secret_value)
                            VALUES (%s, %s, %s)
                            ON CONFLICT (account_id, secret_key)
                            DO UPDATE SET secret_value=excluded.secret_value
                            """,
                            (account_id, "account_password", b64_encode(password)),
                        )
                    exec1(
                        conn,
                        """
                        INSERT INTO app.account_assets(account_id, game_id, asset_type_code)
                        VALUES (%s, %s, 'game')
                        ON CONFLICT (account_id, game_id, asset_type_code) DO NOTHING
                        """,
                        (account_id, game_id),
                    )
                    conn.commit()
                    upload_done += 1
                    last_success_row = idx
                    set_import_progress(job_id, owner, {"phase": "upload", "current": upload_done, "total": total, "done": False})
                except Exception as exc:
                    conn.rollback()
                    failed += 1
                    row_errors.append({"row": idx, "field": "Импорт", "message": str(exc)})
                    upload_done += 1
                    set_import_progress(job_id, owner, {"phase": "upload", "current": upload_done, "total": total, "done": False})
            result = {
                "ok": len(row_errors) == 0,
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "failed": failed,
                "total": total,
                "errors": row_errors,
                "warnings": warnings,
                "success_until_row": last_success_row,
            }
            set_import_progress(job_id, owner, {"phase": "upload", "current": total, "total": total, "done": True, "result": result})
    except Exception as exc:
        set_import_progress(job_id, owner, {"phase": "error", "current": 0, "total": 0, "done": True, "result": {"ok": False, "errors": [{"row": 0, "field": "Импорт", "message": str(exc)}]}})

def run_slots_validate_job(job_id: str, content: bytes, owner: str, limit: Optional[int]):
    try:
        set_import_progress(job_id, owner, {"phase": "validate", "current": 0, "total": 0, "done": False})
        rows = read_slots_from_excel(content)
        if limit and limit > 0:
            rows = rows[:limit]
        total_rows = len(rows)
        set_import_progress(job_id, owner, {"phase": "validate", "current": 0, "total": total_rows, "done": False})
        with psycopg.connect(DB_DSN) as conn:
            errors, warnings, total = validate_slot_import_rows(
                conn,
                rows,
                progress_cb=lambda current: set_import_progress(
                    job_id,
                    owner,
                    {"phase": "validate", "current": current, "total": total_rows, "done": False},
                ),
            )
        result = {"ok": len(errors) == 0, "errors": errors, "warnings": warnings, "total": total}
        set_import_progress(job_id, owner, {"phase": "validate", "current": total_rows, "total": total_rows, "done": True, "result": result})
    except Exception as exc:
        set_import_progress(
            job_id,
            owner,
            {"phase": "error", "current": 0, "total": 0, "done": True, "result": {"ok": False, "errors": [{"row": 0, "field": "Проверка", "message": str(exc)}]}},
        )

def run_slots_import_job(job_id: str, content: bytes, owner: str):
    try:
        with psycopg.connect(DB_DSN) as conn:
            rows = read_slots_from_excel(content)
            errors, warnings, total_valid = validate_slot_import_rows(conn, rows, progress_cb=None)
            if errors:
                set_import_progress(job_id, owner, {"phase": "error", "current": 0, "total": 0, "done": True, "result": {"ok": False, "errors": errors, "warnings": warnings}})
                return
            total = len(rows)
            set_import_progress(job_id, owner, {"phase": "upload", "current": 0, "total": total, "done": False})

            created = 0
            released = 0
            skipped = 0
            failed = 0
            upload_done = 0
            row_errors = []

            DEFAULT_OLD_DT = datetime(MIN_DATE.year, MIN_DATE.month, MIN_DATE.day, tzinfo=timezone.utc)

            prepared = []
            for idx, row in enumerate(rows, start=2):
                status_raw = row.get("status")
                status = "" if status_raw is None else str(status_raw).strip().lower()
                if not status or status == "свободен":
                    skipped += 1
                    prepared.append({"skip": True})
                    continue

                account_val = normalize_cell_text(row.get("account"))
                slot_val = normalize_cell_text(row.get("slot"))
                game_title = normalize_cell_text(row.get("game"))
                customer = normalize_cell_text(row.get("customer"))
                source_val = normalize_cell_text(row.get("source"))
                date_val = row.get("date")

                if not account_val or not slot_val or not game_title:
                    skipped += 1
                    prepared.append({"skip": True})
                    continue

                login, domain = split_account(account_val)
                if not login or not domain:
                    skipped += 1
                    prepared.append({"skip": True})
                    continue

                account_row = q1(
                    conn,
                    """
                    SELECT a.account_id
                    FROM app.accounts a
                    JOIN app.domains d ON d.domain_id = a.domain_id
                    WHERE lower(a.login_name) = lower(%s) AND lower(d.name) = lower(%s)
                    """,
                    (login, domain),
                )
                if not account_row:
                    skipped += 1
                    prepared.append({"skip": True})
                    continue
                account_id = int(account_row[0])

                game_row = q1(conn, "SELECT game_id FROM app.game_titles WHERE lower(title)=lower(%s)", (game_title,))
                if not game_row:
                    skipped += 1
                    prepared.append({"skip": True})
                    continue
                game_id = int(game_row[0])

                slot_key = normalize_slot_file_value(slot_val)
                slot_mapping = SLOT_FILE_TO_TYPE.get(slot_key)
                if not slot_mapping:
                    skipped += 1
                    prepared.append({"skip": True})
                    continue
                slot_type_code, slot_instance = slot_mapping

                date_dt = normalize_slot_date_to_dt(date_val)
                assigned_at = date_dt or DEFAULT_OLD_DT
                source_id = resolve_source_id(conn, source_val) if source_val else None

                prepared.append({
                    "skip": False,
                    "row_idx": idx,
                    "account_id": account_id,
                    "game_id": game_id,
                    "slot_type_code": slot_type_code,
                    "slot_instance": slot_instance,
                    "customer": customer,
                    "source_id": source_id,
                    "assigned_at": assigned_at,
                    "completed_at": date_dt,
                })

            groups = {}
            for item in prepared:
                if item.get("skip"):
                    continue
                key = (item["account_id"], item["game_id"], item["slot_type_code"], item.get("slot_instance"))
                groups.setdefault(key, []).append(item)

            for key in groups:
                groups[key].sort(key=lambda x: (x["assigned_at"], x["row_idx"]))

            upload_done = skipped
            set_import_progress(job_id, owner, {"phase": "upload", "current": upload_done, "total": total, "done": False})

            for key, items in groups.items():
                prev_assignment_id = None
                prev_assigned_at = None
                for item in items:
                    if is_import_cancelled(job_id):
                        set_import_progress(job_id, owner, {"phase": "cancelled", "current": upload_done, "total": total, "done": True, "result": {"ok": False, "cancelled": True}})
                        return
                    try:
                        account_id = item["account_id"]
                        game_id = item["game_id"]
                        slot_type_code = item["slot_type_code"]
                        assigned_at = item["assigned_at"]
                        completed_at = item["completed_at"]
                        customer_nickname = item["customer"]
                        source_id = item["source_id"]

                        # ensure slot type and platform
                        slot_type = ensure_account_allows_slot_type(conn, account_id, slot_type_code)
                        platform_id = get_platform_id(conn, slot_type[1])

                        customer_id = None
                        if customer_nickname:
                            customer_id = ensure_customer(conn, customer_nickname, source_id)

                        region_row = q1(conn, "SELECT region_id FROM app.accounts WHERE account_id=%s", (account_id,))
                        region_id = int(region_row[0]) if region_row and region_row[0] is not None else None

                        deal_row = q1(conn, """
                            INSERT INTO app.deals(
                              deal_type_code, status_code, flow_status_code, region_id, customer_id, currency, total_amount, completed_at
                            )
                            VALUES ('rental', 'confirmed', 'completed', %s, %s, 'RUB', 0, %s)
                            RETURNING deal_id
                        """, (region_id, customer_id, completed_at))
                        deal_id = int(deal_row[0])

                        row_item = q1(conn, """
                            INSERT INTO app.deal_items(
                              deal_id, account_id, game_id, platform_id,
                              qty, price, purchase_cost, fee, purchase_at, start_at, end_at, slots_used, slot_type_code, notes, game_link
                            )
                            VALUES (%s, %s, %s, %s, 1, 0, 0, 0, %s, %s, NULL, 1, %s, NULL, NULL)
                            RETURNING deal_item_id
                        """, (
                            deal_id, account_id, game_id, platform_id,
                            completed_at, completed_at, slot_type_code
                        ))
                        deal_item_id = int(row_item[0])

                        assignment_row = q1(
                            conn,
                            """
                            INSERT INTO app.account_slot_assignments(
                              account_id, slot_type_code, customer_id, game_id, deal_id, deal_item_id, assigned_at, assigned_by
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            RETURNING assignment_id
                            """,
                            (account_id, slot_type_code, customer_id, game_id, deal_id, deal_item_id, assigned_at, owner),
                        )
                        assignment_id = int(assignment_row[0])

                        if prev_assignment_id is not None and prev_assigned_at is not None:
                            exec1(
                                conn,
                                """
                                UPDATE app.account_slot_assignments
                                SET released_at=%s, released_by=%s
                                WHERE assignment_id=%s
                                """,
                                (assigned_at, owner, prev_assignment_id),
                            )
                            released += 1

                        conn.commit()
                        created += 1
                        prev_assignment_id = assignment_id
                        prev_assigned_at = assigned_at
                    except Exception as exc:
                        conn.rollback()
                        failed += 1
                        row_errors.append({"row": item.get("row_idx"), "field": "Импорт", "message": str(exc)})
                    finally:
                        upload_done += 1
                        set_import_progress(job_id, owner, {"phase": "upload", "current": upload_done, "total": total, "done": False})

            result = {
                "ok": len(row_errors) == 0,
                "created": created,
                "released": released,
                "skipped": skipped,
                "failed": failed,
                "total": total,
                "errors": row_errors,
                "warnings": warnings,
            }
            set_import_progress(job_id, owner, {"phase": "upload", "current": total, "total": total, "done": True, "result": result})
    except Exception as exc:
        set_import_progress(job_id, owner, {"phase": "error", "current": 0, "total": 0, "done": True, "result": {"ok": False, "errors": [{"row": 0, "field": "Импорт", "message": str(exc)}]}})

@app.get("/accounts/import/template")
def accounts_import_template(user: UserOut = Depends(require_role("admin"))):
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"
    ws.append(["Аккаунт", "Пароль", "Игра"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=accounts_import_template.xlsx"},
    )

@app.post("/accounts/slots/clean")
def accounts_slots_clean(file: UploadFile = File(...), user: UserOut = Depends(require_role("admin"))):
    if not file or not file.filename:
        raise HTTPException(400, "file is required")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls are supported")
    content = file.file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, "File too large. Max 10MB")
    cleaned = clean_slots_excel(content)
    base = Path(file.filename or "slots_import").stem
    filename = f"{base}_cleaned.xlsx"
    return Response(
        cleaned,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

@app.post("/accounts/slots/validate")
def accounts_slots_validate(
    file: UploadFile = File(...),
    limit: Optional[int] = Form(None),
    user: UserOut = Depends(require_role("admin")),
):
    if not file or not file.filename:
        raise HTTPException(400, "file is required")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls are supported")
    content = file.file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, "File too large. Max 10MB")
    job_id = uuid.uuid4().hex
    set_import_progress(job_id, user.username, {"phase": "queued", "current": 0, "total": 0, "done": False})
    thread = threading.Thread(target=run_slots_validate_job, args=(job_id, content, user.username, limit), daemon=True)
    thread.start()
    return {"ok": True, "job_id": job_id}

@app.get("/accounts/slots/validate/status")
def accounts_slots_validate_status(job_id: str, user: UserOut = Depends(require_role("admin"))):
    status = get_import_progress(job_id)
    if not status:
        return {"phase": "idle", "current": 0, "total": 0, "done": True}
    if status.get("owner") and status.get("owner") != user.username:
        raise HTTPException(403, "job not found")
    return {
        "phase": status.get("phase", "idle"),
        "current": int(status.get("current") or 0),
        "total": int(status.get("total") or 0),
        "done": bool(status.get("done")),
        "result": status.get("result"),
    }

@app.post("/accounts/slots/validate/cancel")
def accounts_slots_validate_cancel(job_id: str, user: UserOut = Depends(require_role("admin"))):
    status = get_import_progress(job_id)
    if not status:
        return {"ok": True}
    if status.get("owner") and status.get("owner") != user.username:
        raise HTTPException(403, "job not found")
    set_import_progress(job_id, user.username, {
        "phase": "cancelled",
        "current": int(status.get("current") or 0),
        "total": int(status.get("total") or 0),
        "done": True,
        "cancelled": True,
        "result": {"ok": False, "cancelled": True},
    })
    return {"ok": True}

@app.post("/accounts/slots/report")
def accounts_slots_report(payload: ImportReportIn, user: UserOut = Depends(require_role("admin"))):
    content = build_import_report_xlsx(
        errors=payload.errors or [],
        warnings=payload.warnings or [],
    )
    return Response(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=slots_import_report.xlsx"},
    )

@app.post("/accounts/slots/import")
def accounts_slots_import(file: UploadFile = File(...), user: UserOut = Depends(require_role("admin"))):
    if not file or not file.filename:
        raise HTTPException(400, "file is required")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls are supported")
    content = file.file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, "File too large. Max 10MB")
    job_id = uuid.uuid4().hex
    set_import_progress(job_id, user.username, {"phase": "queued", "current": 0, "total": 0, "done": False})
    thread = threading.Thread(target=run_slots_import_job, args=(job_id, content, user.username), daemon=True)
    thread.start()
    return {"ok": True, "job_id": job_id}

@app.get("/accounts/slots/import/status")
def accounts_slots_import_status(job_id: str, user: UserOut = Depends(require_role("admin"))):
    status = get_import_progress(job_id)
    if not status:
        return {"phase": "idle", "current": 0, "total": 0, "done": True}
    if status.get("owner") and status.get("owner") != user.username:
        raise HTTPException(403, "job not found")
    return {
        "phase": status.get("phase", "idle"),
        "current": int(status.get("current") or 0),
        "total": int(status.get("total") or 0),
        "done": bool(status.get("done")),
        "result": status.get("result"),
    }

@app.post("/accounts/slots/import/cancel")
def accounts_slots_import_cancel(job_id: str, user: UserOut = Depends(require_role("admin"))):
    status = get_import_progress(job_id)
    if not status:
        return {"ok": True}
    if status.get("owner") and status.get("owner") != user.username:
        raise HTTPException(403, "job not found")
    set_import_progress(job_id, user.username, {
        "phase": "cancelled",
        "current": int(status.get("current") or 0),
        "total": int(status.get("total") or 0),
        "done": True,
        "cancelled": True,
        "result": {"ok": False, "cancelled": True},
    })
    return {"ok": True}

@app.get("/accounts/import/status")
def accounts_import_status(job_id: str, user: UserOut = Depends(require_role("admin"))):
    status = get_import_progress(job_id)
    if not status:
        return {"phase": "idle", "current": 0, "total": 0, "done": True}
    if status.get("owner") and status.get("owner") != user.username:
        raise HTTPException(403, "job not found")
    return {
        "phase": status.get("phase", "idle"),
        "current": int(status.get("current") or 0),
        "total": int(status.get("total") or 0),
        "done": bool(status.get("done")),
        "result": status.get("result"),
    }

@app.post("/accounts/import/validate")
def accounts_import_validate(file: UploadFile = File(...), user: UserOut = Depends(require_role("admin"))):
    if not file or not file.filename:
        raise HTTPException(400, "file is required")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls are supported")
    content = file.file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(400, "File too large. Max 5MB")
    with psycopg.connect(DB_DSN) as conn:
        rows = read_accounts_from_excel(content)
        errors, warnings = validate_account_import_rows(conn, rows, progress_cb=None)
    return {"ok": len(errors) == 0, "total": len(rows), "errors": errors, "warnings": warnings}

@app.post("/accounts/import")
def accounts_import(file: UploadFile = File(...), user: UserOut = Depends(require_role("admin"))):
    if not file or not file.filename:
        raise HTTPException(400, "file is required")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls are supported")
    content = file.file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(400, "File too large. Max 5MB")
    job_id = uuid.uuid4().hex
    set_import_progress(job_id, user.username, {"phase": "queued", "current": 0, "total": 0, "done": False})
    thread = threading.Thread(target=run_accounts_import_job, args=(job_id, content, user.username), daemon=True)
    thread.start()
    return {"ok": True, "job_id": job_id}

@app.post("/accounts/import/cancel")
def accounts_import_cancel(job_id: str, user: UserOut = Depends(require_role("admin"))):
    status = get_import_progress(job_id)
    if not status:
        return {"ok": True}
    if status.get("owner") and status.get("owner") != user.username:
        raise HTTPException(403, "job not found")
    set_import_progress(job_id, user.username, {
        "phase": "cancelled",
        "current": int(status.get("current") or 0),
        "total": int(status.get("total") or 0),
        "done": True,
        "cancelled": True,
        "result": {"ok": False, "cancelled": True},
    })
    return {"ok": True}

def build_import_report_xlsx(errors: List[dict], warnings: List[dict]) -> bytes:
    wb = Workbook()
    ws_err = wb.active
    ws_err.title = "Ошибки"
    ws_err.append(["Строка", "Поле", "Значение", "Сообщение"])
    def _get(issue, key):
        if isinstance(issue, dict):
            return issue.get(key)
        return getattr(issue, key, None)
    for e in errors or []:
        ws_err.append([_get(e, "row"), _get(e, "field"), _get(e, "value"), _get(e, "message")])
    ws_warn = wb.create_sheet("Предупреждения")
    ws_warn.append(["Строка", "Поле", "Значение", "Сообщение"])
    for w in warnings or []:
        ws_warn.append([_get(w, "row"), _get(w, "field"), _get(w, "value"), _get(w, "message")])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

@app.post("/accounts/import/report")
def accounts_import_report(payload: ImportReportIn, user: UserOut = Depends(require_role("admin"))):
    content = build_import_report_xlsx(
        [i.dict() for i in payload.errors],
        [i.dict() for i in payload.warnings],
    )
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=accounts_import_report.xlsx"},
    )

@app.post("/games/import/report")
def games_import_report(payload: ImportReportIn, user: UserOut = Depends(require_role("admin"))):
    content = build_import_report_xlsx(
        [i.dict() for i in payload.errors],
        [i.dict() for i in payload.warnings],
    )
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=games_import_report.xlsx"},
    )

@app.post("/rentals")
def create_rental(payload: RentalCreate, user: UserOut = Depends(get_current_user)):
    if not payload.slot_type_code:
        raise HTTPException(400, "slot_type_code is required for rental")
    validate_date_in_range(payload.purchase_at, "purchase_at")
    validate_date_in_range(payload.start_at, "start_at")
    validate_date_in_range(payload.end_at, "end_at")
    validate_date_range(payload.start_at, payload.end_at, "end_at")

    start_at = payload.start_at or now_utc()
    end_at = payload.end_at

    with psycopg.connect(DB_DSN) as conn:
        ensure_account_exists(conn, payload.account_id)
        ensure_game_active(conn, payload.game_id)
        ensure_source_exists(conn, payload.source_id)
        slot_type = ensure_account_allows_slot_type(conn, payload.account_id, payload.slot_type_code)
        platform_id = get_platform_id(conn, slot_type[1])
        # ensure customer exists
        row = q1(conn, "SELECT customer_id, source_id FROM app.customers WHERE nickname=%s", (payload.customer_nickname,))
        if row:
            customer_id = int(row[0])
            if payload.source_id and not row[1]:
                exec1(
                    conn,
                    "UPDATE app.customers SET source_id=%s WHERE customer_id=%s",
                    (payload.source_id, customer_id),
                )
        else:
            row = q1(
                conn,
                "INSERT INTO app.customers(nickname, source_id) VALUES (%s, %s) RETURNING customer_id",
                (payload.customer_nickname, payload.source_id),
            )
            customer_id = int(row[0])

        # check slots
        free = get_account_slot_free(conn, payload.account_id, payload.slot_type_code)
        if free < 1:
            raise HTTPException(409, "Not enough free slots for selected slot type")

        # create deal + item
        region_row = q1(conn, "SELECT region_id FROM app.accounts WHERE account_id=%s", (payload.account_id,))
        region_id = int(region_row[0]) if region_row and region_row[0] is not None else None
        deal_row = q1(conn, """
            INSERT INTO app.deals(deal_type_code, status_code, flow_status_code, region_id, customer_id, currency, total_amount)
            VALUES ('rental', 'confirmed', 'pending', %s, %s, 'RUB', %s)
            RETURNING deal_id
        """, (region_id, customer_id, payload.price))
        deal_id = int(deal_row[0])

        row_item = q1(conn, """
            INSERT INTO app.deal_items(
              deal_id, account_id, game_id, platform_id,
              qty, price, purchase_cost, start_at, end_at, slots_used, slot_type_code, purchase_at, notes, game_link
            )
            VALUES (%s, %s, %s, %s, 1, %s, 0, %s, %s, 1, %s, %s, %s, %s)
            RETURNING deal_item_id
        """, (
            deal_id, payload.account_id, payload.game_id, platform_id,
            payload.price, start_at, end_at, payload.slot_type_code, payload.purchase_at, None, None
        ))
        deal_item_id = int(row_item[0])
        exec1(
            conn,
            """
            INSERT INTO app.account_slot_assignments(
              account_id, slot_type_code, customer_id, game_id, deal_id, deal_item_id, assigned_by
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (payload.account_id, payload.slot_type_code, customer_id, payload.game_id, deal_id, deal_item_id, user.username),
        )
        conn.commit()

        conn.commit()

        # return updated slots
        occ2, free2 = slots_summary(conn, payload.account_id, platform_id)
        return {"deal_id": deal_id, "account_id": payload.account_id, "occupied_slots": occ2, "free_slots": free2}

@app.post("/deals")
def create_deal(payload: DealCreate, user: UserOut = Depends(get_current_user)):
    deal_type = payload.deal_type_code.strip().lower()
    if deal_type not in ("sale", "rental"):
        raise HTTPException(400, "deal_type_code must be sale or rental")
    if deal_type == "rental":
        if not payload.slot_type_code:
            raise HTTPException(400, "slot_type_code is required for rental")
        if not payload.account_id:
            raise HTTPException(400, "account_id is required for rental")
        if not payload.game_id:
            raise HTTPException(400, "game_id is required for rental")
    if deal_type == "sale":
        if not payload.region_code:
            raise HTTPException(400, "region_code is required for sale")
    if deal_type == "sale":
        payload.slots_used = 0
        if not payload.purchase_at:
            payload.purchase_at = now_utc()
    validate_date_in_range(payload.purchase_at, "purchase_at")
    validate_date_in_range(payload.start_at, "start_at")
    validate_date_in_range(payload.end_at, "end_at")
    validate_date_range(payload.start_at, payload.end_at, "end_at")

    with psycopg.connect(DB_DSN) as conn:
        if deal_type == "rental":
            ensure_account_exists(conn, payload.account_id)
            ensure_game_active(conn, payload.game_id)
        ensure_source_exists(conn, payload.source_id)
        platform_id = None
        if deal_type == "rental":
            slot_type = ensure_account_allows_slot_type(conn, payload.account_id, payload.slot_type_code)
            platform_id = get_platform_id(conn, slot_type[1])
        region_id = get_region_id(conn, payload.region_code)
        if region_id is None:
            if payload.account_id:
                region_row = q1(conn, "SELECT region_id FROM app.accounts WHERE account_id=%s", (payload.account_id,))
                region_id = int(region_row[0]) if region_row and region_row[0] is not None else None

        row = q1(conn, "SELECT customer_id, source_id FROM app.customers WHERE nickname=%s", (payload.customer_nickname,))
        if row:
            customer_id = int(row[0])
            if payload.source_id and not row[1]:
                exec1(
                    conn,
                    "UPDATE app.customers SET source_id=%s WHERE customer_id=%s",
                    (payload.source_id, customer_id),
                )
        else:
            row = q1(
                conn,
                "INSERT INTO app.customers(nickname, source_id) VALUES (%s, %s) RETURNING customer_id",
                (payload.customer_nickname, payload.source_id),
            )
            customer_id = int(row[0])

        if deal_type == "rental":
            free = get_account_slot_free(conn, payload.account_id, payload.slot_type_code)
            if free < 1:
                raise HTTPException(409, "Not enough free slots for selected slot type")

        deal_row = q1(conn, """
            INSERT INTO app.deals(deal_type_code, status_code, flow_status_code, region_id, customer_id, currency, total_amount)
            VALUES (%s, 'confirmed', 'pending', %s, %s, 'RUB', %s)
            RETURNING deal_id
        """, (deal_type, region_id, customer_id, payload.price))
        deal_id = int(deal_row[0])

        row_item = q1(conn, """
            INSERT INTO app.deal_items(
              deal_id, account_id, game_id, platform_id,
              qty, price, purchase_cost, fee, purchase_at, start_at, end_at, slots_used, slot_type_code, notes, game_link
            )
            VALUES (%s, %s, %s, %s, 1, %s, %s, 0, %s, %s, %s, %s, %s, %s, %s)
            RETURNING deal_item_id
        """, (
            deal_id, payload.account_id, payload.game_id, platform_id,
            payload.price, payload.purchase_cost, payload.purchase_at, payload.start_at, payload.end_at,
            (0 if deal_type == "sale" else 1), payload.slot_type_code, payload.notes, payload.game_link
        ))
        deal_item_id = int(row_item[0])
        if deal_type == "rental":
            exec1(
                conn,
                """
                INSERT INTO app.account_slot_assignments(
                  account_id, slot_type_code, customer_id, game_id, deal_id, deal_item_id, assigned_by
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (payload.account_id, payload.slot_type_code, customer_id, payload.game_id, deal_id, deal_item_id, user.username),
            )
        conn.commit()

        return {"deal_id": deal_id}

@app.put("/deals/{deal_id}")
def update_deal(deal_id: int, payload: DealUpdate, user: UserOut = Depends(get_current_user)):
    if payload.purchase_at is not None:
        validate_date_in_range(payload.purchase_at, "purchase_at")
    if payload.start_at is not None:
        validate_date_in_range(payload.start_at, "start_at")
    if payload.end_at is not None:
        validate_date_in_range(payload.end_at, "end_at")
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT set_config('app.user', %s, true)", (user.username,))
        row = q1(
            conn,
            """
            SELECT
              d.deal_type_code,
              d.status_code,
              d.flow_status_code,
              d.region_id,
              d.customer_id,
              d.total_amount,
              di.deal_item_id,
              di.account_id,
              di.game_id,
              di.platform_id,
              di.price,
              di.purchase_cost,
              di.purchase_at,
              di.start_at,
              di.end_at,
              di.slots_used,
              di.slot_type_code,
              di.notes,
              di.game_link
            FROM app.deals d
            JOIN app.deal_items di ON di.deal_id = d.deal_id
            WHERE d.deal_id=%s
            ORDER BY di.deal_item_id ASC
            LIMIT 1
            """,
            (deal_id,),
        )
        if not row:
            raise HTTPException(404, "Deal not found")

        current_type, status_code, flow_status_code, region_id, customer_id, total_amount, deal_item_id, \
            account_id, game_id, platform_id, price, purchase_cost, purchase_at, start_at, end_at, slots_used, slot_type_code, notes, game_link = row

        deal_type = (payload.deal_type_code or current_type or "").strip().lower()
        if deal_type not in ("sale", "rental"):
            raise HTTPException(400, "deal_type_code must be sale or rental")

        new_account_id = payload.account_id if payload.account_id is not None else account_id
        new_game_id = payload.game_id if payload.game_id is not None else game_id
        ensure_source_exists(conn, payload.source_id if payload.source_id is not None else None)
        if payload.region_code is not None:
            region_id = get_region_id(conn, payload.region_code)

        new_slot_type_code = payload.slot_type_code if payload.slot_type_code is not None else slot_type_code
        new_platform_id = platform_id
        slot_type_platform = None
        if deal_type == "rental":
            if not new_slot_type_code:
                raise HTTPException(400, "slot_type_code is required for rental")
            slot_type = get_slot_type(conn, new_slot_type_code)
            slot_type_platform = slot_type[1]
            new_platform_id = get_platform_id(conn, slot_type[1])

        # customer update
        cust_nickname = payload.customer_nickname
        cust_source = payload.source_id if payload.source_id is not None else None
        if cust_nickname:
            customer_id = ensure_customer(conn, cust_nickname, cust_source)

        new_price = payload.price if payload.price is not None else price
        new_purchase_cost = payload.purchase_cost if payload.purchase_cost is not None else purchase_cost
        new_purchase_at = payload.purchase_at if payload.purchase_at is not None else purchase_at
        new_start_at = payload.start_at if payload.start_at is not None else start_at
        new_end_at = payload.end_at if payload.end_at is not None else end_at
        new_notes = payload.notes if payload.notes is not None else notes
        new_game_link = payload.game_link if payload.game_link is not None else game_link
        new_slots_used = payload.slots_used if payload.slots_used is not None else slots_used
        new_flow_status = payload.flow_status_code if payload.flow_status_code is not None else flow_status_code
        if payload.flow_status_code is not None:
            row = q1(conn, "SELECT 1 FROM app.deal_flow_statuses WHERE code=%s", (payload.flow_status_code,))
            if not row:
                raise HTTPException(400, "Unknown flow_status_code")
        if deal_type == "sale":
            new_slots_used = 0
            new_account_id = None
            new_game_id = None
            new_platform_id = None
            new_slot_type_code = None
            if region_id is None:
                raise HTTPException(400, "region_code is required for sale")
        if deal_type == "rental":
            new_slots_used = 1
        validate_date_range(new_start_at, new_end_at, "end_at")

        # check slots for rental
        if deal_type == "rental":
            if not new_account_id:
                raise HTTPException(400, "account_id is required for rental")
            if not new_game_id:
                raise HTTPException(400, "game_id is required for rental")
            ensure_account_exists(conn, new_account_id)
            ensure_game_exists(conn, new_game_id)
            if slot_type_platform == "ps4" and not account_has_ps4(conn, new_account_id):
                raise HTTPException(400, "Account does not support PS4 slot type")
            free = get_account_slot_free(conn, new_account_id, new_slot_type_code)
            row_assign = q1(
                conn,
                """
                SELECT account_id, slot_type_code
                FROM app.account_slot_assignments
                WHERE deal_item_id=%s AND released_at IS NULL
                """,
                (deal_item_id,),
            )
            same_assignment = row_assign and int(row_assign[0]) == int(new_account_id) and row_assign[1] == new_slot_type_code
            free_adjusted = free + (1 if same_assignment else 0)
            if free_adjusted < 1:
                raise HTTPException(409, "Not enough free slots for selected slot type")

        completed_at_changed = flow_status_code != new_flow_status
        completed_at_value = None
        if completed_at_changed:
            completed_at_value = now_utc() if new_flow_status == "completed" else None

        exec1(
            conn,
            """
            UPDATE app.deals
            SET deal_type_code=%s,
                customer_id=%s,
                total_amount=%s,
                flow_status_code=%s,
                region_id=%s,
                completed_at=CASE WHEN %s THEN %s ELSE completed_at END
            WHERE deal_id=%s
            """,
            (deal_type, customer_id, new_price, new_flow_status, region_id, completed_at_changed, completed_at_value, deal_id),
        )

        exec1(
            conn,
            """
            UPDATE app.deal_items
            SET account_id=%s,
                game_id=%s,
                platform_id=%s,
                price=%s,
                purchase_cost=%s,
                purchase_at=%s,
                start_at=%s,
                end_at=%s,
                slots_used=%s,
                slot_type_code=%s,
                notes=%s,
                game_link=%s
            WHERE deal_item_id=%s
            """,
            (
                new_account_id,
                new_game_id,
                new_platform_id,
                new_price,
                new_purchase_cost,
                new_purchase_at,
                new_start_at,
                new_end_at,
                new_slots_used,
                new_slot_type_code,
                new_notes,
                new_game_link,
                deal_item_id,
            ),
        )
        if deal_type == "rental":
            row_assign = q1(
                conn,
                """
                SELECT assignment_id, account_id, slot_type_code, customer_id, game_id
                FROM app.account_slot_assignments
                WHERE deal_item_id=%s AND released_at IS NULL
                """,
                (deal_item_id,),
            )
            current_assign = row_assign[0] if row_assign else None
            need_new_assign = (
                (not row_assign)
                or int(row_assign[1]) != int(new_account_id)
                or row_assign[2] != new_slot_type_code
                or (customer_id is not None and row_assign[3] != customer_id)
                or (new_game_id is not None and row_assign[4] != new_game_id)
            )
            if need_new_assign and current_assign:
                release_slot_assignment(conn, deal_item_id, user.username)
            if need_new_assign:
                exec1(
                    conn,
                    """
                    INSERT INTO app.account_slot_assignments(
                      account_id, slot_type_code, customer_id, game_id, deal_id, deal_item_id, assigned_by
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (new_account_id, new_slot_type_code, customer_id, new_game_id, deal_id, deal_item_id, user.username),
                )
        conn.commit()

    return {"ok": True}

@app.get("/deals", response_model=DealListOut)
def list_deals(
    account_id: Optional[int] = None,
    game_id: Optional[int] = None,
    platform_code: Optional[str] = None,
    q: Optional[str] = None,
    deal_type_code: Optional[str] = None,
    status_code: Optional[str] = None,
    flow_status_code: Optional[str] = None,
    customer_q: Optional[str] = None,
    source_id: Optional[int] = None,
    purchase_from: Optional[date] = None,
    purchase_to: Optional[date] = None,
    price_min: Optional[float] = None,
    price_max: Optional[float] = None,
    notes_q: Optional[str] = None,
    account_q: Optional[str] = None,
    region_q: Optional[str] = None,
    game_q: Optional[str] = None,
    platform_q: Optional[str] = None,
    type_q: Optional[str] = None,
    status_q: Optional[str] = None,
    flow_status_q: Optional[str] = None,
    source_q: Optional[str] = None,
    date_q: Optional[str] = None,
    price_q: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    user: UserOut = Depends(get_current_user),
):
    if page < 1:
        raise HTTPException(400, "page must be >= 1")
    if page_size < 1 or page_size > 200:
        raise HTTPException(400, "page_size must be between 1 and 200")
    offset = (page - 1) * page_size

    with psycopg.connect(DB_DSN) as conn:
        where_sql, params = build_deals_filters(
            account_id,
            game_id,
            platform_code,
            q,
            deal_type_code,
            status_code,
            flow_status_code,
            customer_q,
            source_id,
            purchase_from,
            purchase_to,
            price_min,
            price_max,
            notes_q,
            account_q,
            region_q,
            game_q,
            platform_q,
            type_q,
            status_q,
            flow_status_q,
            source_q,
            date_q,
            price_q,
        )

        total_row = q1(conn, f"""
            SELECT COUNT(*)
            FROM app.deal_items di
            JOIN app.deals d ON d.deal_id = di.deal_id
            LEFT JOIN app.accounts a ON a.account_id = di.account_id
            LEFT JOIN app.domains dm ON dm.domain_id = a.domain_id
            LEFT JOIN app.regions ra ON ra.region_id = a.region_id
            LEFT JOIN app.regions rd ON rd.region_id = d.region_id
            LEFT JOIN app.game_titles g ON g.game_id = di.game_id
            LEFT JOIN app.platforms p ON p.platform_id = di.platform_id
            LEFT JOIN app.customers c ON c.customer_id = d.customer_id
            LEFT JOIN app.sources src ON src.source_id = c.source_id
            LEFT JOIN app.deal_flow_statuses fs ON fs.code = d.flow_status_code
            LEFT JOIN app.deal_statuses ds ON ds.code = d.status_code
            LEFT JOIN app.deal_types dt ON dt.code = d.deal_type_code
            {where_sql}
        """, params)
        total = int(total_row[0]) if total_row else 0

        rows = qall(conn, f"""
            SELECT
              d.deal_id,
              dt.name as deal_type_name,
              dt.code as deal_type_code,
              ds.name as status_name,
              d.flow_status_code,
              fs.name as flow_status_name,
              COALESCE(rd.code, ra.code) as region_code,
              di.account_id,
              a.login_name,
              dm.name as domain_name,
              di.game_id,
              g.title,
              g.short_title,
              p.code as platform_code,
              c.nickname,
              c.source_id,
              di.price,
              di.purchase_cost,
              di.purchase_at,
              d.created_at,
              di.slots_used,
              di.slot_type_code,
              di.notes,
              di.game_link
            FROM app.deal_items di
            JOIN app.deals d ON d.deal_id = di.deal_id
            LEFT JOIN app.deal_types dt ON dt.code = d.deal_type_code
            LEFT JOIN app.deal_statuses ds ON ds.code = d.status_code
            LEFT JOIN app.deal_flow_statuses fs ON fs.code = d.flow_status_code
            LEFT JOIN app.accounts a ON a.account_id = di.account_id
            LEFT JOIN app.domains dm ON dm.domain_id = a.domain_id
            LEFT JOIN app.regions ra ON ra.region_id = a.region_id
            LEFT JOIN app.regions rd ON rd.region_id = d.region_id
            LEFT JOIN app.game_titles g ON g.game_id = di.game_id
            LEFT JOIN app.platforms p ON p.platform_id = di.platform_id
            LEFT JOIN app.customers c ON c.customer_id = d.customer_id
            LEFT JOIN app.sources src ON src.source_id = c.source_id
            {where_sql}
            ORDER BY d.created_at DESC
            LIMIT %s OFFSET %s
        """, params + [page_size, offset])

    items = []
    for r in rows:
        login_full = f"{r[8]}@{r[9]}" if r[8] and r[9] else None
        items.append(
            DealListItem(
                deal_id=r[0],
                deal_type=r[1],
                deal_type_code=r[2],
                status=r[3],
                flow_status_code=r[4],
                flow_status=r[5],
                region_code=r[6],
                account_id=r[7],
                account_login=login_full,
                game_id=r[10],
                game_title=r[11],
                game_short_title=r[12],
                platform_code=r[13],
                customer_nickname=r[14],
                source_id=r[15],
                price=float(r[16] or 0),
                purchase_cost=float(r[17] or 0),
                purchase_at=r[18],
                created_at=r[19],
                slots_used=r[20],
                slot_type_code=r[21],
                notes=r[22],
                game_link=r[23],
            )
        )

    return DealListOut(total=total, items=items)

@app.get("/analytics/sales", response_model=SalesAnalyticsOut)
def analytics_sales(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    deal_type_code: Optional[str] = None,
    region_code: Optional[str] = None,
    source_id: Optional[int] = None,
    user: UserOut = Depends(get_current_user),
):
    params: List[Any] = []
    filters = ["activity_at IS NOT NULL", "status_code = 'confirmed'"]

    if deal_type_code:
        filters.append("deal_type_code = %s")
        params.append(deal_type_code)
    if region_code:
        filters.append("region_code = %s")
        params.append(region_code)
    if source_id:
        filters.append("source_id = %s")
        params.append(source_id)
    if date_from:
        filters.append("activity_at::date >= %s")
        params.append(date_from)
    if date_to:
        filters.append("activity_at::date <= %s")
        params.append(date_to)

    where_sql = " AND ".join(filters)

    with psycopg.connect(DB_DSN) as conn:
        totals_row = q1(conn, f"""
            WITH base AS (
                SELECT
                  d.deal_type_code,
                  d.status_code,
                  COALESCE(rd.code, ra.code) AS region_code,
                  COALESCE(rd.purchase_cost_rate, ra.purchase_cost_rate, 1.0) AS rate,
                  c.source_id,
                  di.price,
                  di.purchase_cost,
                  di.qty,
                  CASE
                    WHEN d.deal_type_code = 'sale' THEN d.completed_at
                    ELSE COALESCE(di.purchase_at, d.created_at)
                  END AS activity_at
                FROM app.deal_items di
                JOIN app.deals d ON d.deal_id = di.deal_id
                LEFT JOIN app.accounts a ON a.account_id = di.account_id
                LEFT JOIN app.regions ra ON ra.region_id = a.region_id
                LEFT JOIN app.regions rd ON rd.region_id = d.region_id
                LEFT JOIN app.customers c ON c.customer_id = d.customer_id
            )
            SELECT
              COALESCE(SUM(price * qty), 0) AS revenue,
              COALESCE(SUM(purchase_cost * qty * rate), 0) AS purchase_cost,
              COUNT(*) AS count
            FROM base
            WHERE {where_sql}
        """, params)

        revenue = float(totals_row[0] or 0) if totals_row else 0.0
        purchase_cost = float(totals_row[1] or 0) if totals_row else 0.0
        count = int(totals_row[2] or 0) if totals_row else 0
        margin = revenue - purchase_cost
        avg_check = (revenue / count) if count else 0.0

        by_day_rows = qall(conn, f"""
            WITH base AS (
                SELECT
                  d.deal_type_code,
                  d.status_code,
                  COALESCE(rd.code, ra.code) AS region_code,
                  COALESCE(rd.purchase_cost_rate, ra.purchase_cost_rate, 1.0) AS rate,
                  c.source_id,
                  di.price,
                  di.purchase_cost,
                  di.qty,
                  CASE
                    WHEN d.deal_type_code = 'sale' THEN d.completed_at
                    ELSE COALESCE(di.purchase_at, d.created_at)
                  END AS activity_at
                FROM app.deal_items di
                JOIN app.deals d ON d.deal_id = di.deal_id
                LEFT JOIN app.accounts a ON a.account_id = di.account_id
                LEFT JOIN app.regions ra ON ra.region_id = a.region_id
                LEFT JOIN app.regions rd ON rd.region_id = d.region_id
                LEFT JOIN app.customers c ON c.customer_id = d.customer_id
            )
            SELECT
              activity_at::date AS day,
              COALESCE(SUM(price * qty), 0) AS revenue,
              COALESCE(SUM(purchase_cost * qty * rate), 0) AS purchase_cost,
              COUNT(*) AS count
            FROM base
            WHERE {where_sql}
            GROUP BY day
            ORDER BY day
        """, params)

        by_type_rows = qall(conn, f"""
            WITH base AS (
                SELECT
                  d.deal_type_code,
                  d.status_code,
                  COALESCE(rd.code, ra.code) AS region_code,
                  COALESCE(rd.purchase_cost_rate, ra.purchase_cost_rate, 1.0) AS rate,
                  c.source_id,
                  di.price,
                  di.purchase_cost,
                  di.qty,
                  CASE
                    WHEN d.deal_type_code = 'sale' THEN d.completed_at
                    ELSE COALESCE(di.purchase_at, d.created_at)
                  END AS activity_at
                FROM app.deal_items di
                JOIN app.deals d ON d.deal_id = di.deal_id
                LEFT JOIN app.accounts a ON a.account_id = di.account_id
                LEFT JOIN app.regions ra ON ra.region_id = a.region_id
                LEFT JOIN app.regions rd ON rd.region_id = d.region_id
                LEFT JOIN app.customers c ON c.customer_id = d.customer_id
            )
            SELECT
              deal_type_code,
              COALESCE(SUM(price * qty), 0) AS revenue,
              COALESCE(SUM(purchase_cost * qty * rate), 0) AS purchase_cost,
              COUNT(*) AS count
            FROM base
            WHERE {where_sql}
            GROUP BY deal_type_code
            ORDER BY deal_type_code
        """, params)

    by_day = []
    for r in by_day_rows:
        rev = float(r[1] or 0)
        pc = float(r[2] or 0)
        by_day.append(
            SalesAnalyticsPoint(
                date=r[0],
                revenue=rev,
                purchase_cost=pc,
                margin=rev - pc,
                count=int(r[3] or 0),
            )
        )

    by_type = []
    for r in by_type_rows:
        rev = float(r[1] or 0)
        pc = float(r[2] or 0)
        by_type.append(
            SalesAnalyticsByType(
                deal_type_code=r[0],
                revenue=rev,
                purchase_cost=pc,
                margin=rev - pc,
                count=int(r[3] or 0),
            )
        )

    return SalesAnalyticsOut(
        totals=SalesAnalyticsTotals(
            revenue=revenue,
            purchase_cost=purchase_cost,
            margin=margin,
            count=count,
            avg_check=avg_check,
        ),
        by_day=by_day,
        by_type=by_type,
    )

@app.get("/analytics/sources", response_model=SourcesAnalyticsOut)
def analytics_sources(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    deal_type_code: Optional[str] = None,
    region_code: Optional[str] = None,
    source_id: Optional[int] = None,
    user: UserOut = Depends(get_current_user),
):
    params: List[Any] = []
    filters = ["activity_at IS NOT NULL", "status_code = 'confirmed'"]

    if deal_type_code:
        filters.append("deal_type_code = %s")
        params.append(deal_type_code)
    if region_code:
        filters.append("region_code = %s")
        params.append(region_code)
    if source_id:
        filters.append("source_id = %s")
        params.append(source_id)
    if date_from:
        filters.append("activity_at::date >= %s")
        params.append(date_from)
    if date_to:
        filters.append("activity_at::date <= %s")
        params.append(date_to)

    where_sql = " AND ".join(filters)

    with psycopg.connect(DB_DSN) as conn:
        top_count_rows = qall(conn, f"""
            WITH base AS (
                SELECT
                  d.deal_type_code,
                  d.status_code,
                  COALESCE(rd.code, ra.code) AS region_code,
                  c.source_id,
                  src.code as source_code,
                  src.name as source_name,
                  di.price,
                  di.qty,
                  CASE
                    WHEN d.deal_type_code = 'sale' THEN d.completed_at
                    ELSE COALESCE(di.purchase_at, d.created_at)
                  END AS activity_at
                FROM app.deal_items di
                JOIN app.deals d ON d.deal_id = di.deal_id
                LEFT JOIN app.accounts a ON a.account_id = di.account_id
                LEFT JOIN app.regions ra ON ra.region_id = a.region_id
                LEFT JOIN app.regions rd ON rd.region_id = d.region_id
                LEFT JOIN app.customers c ON c.customer_id = d.customer_id
                LEFT JOIN app.sources src ON src.source_id = c.source_id
            )
            SELECT
              source_id,
              source_code,
              source_name,
              COUNT(*) AS deals_count,
              COALESCE(SUM(price * qty), 0) AS revenue
            FROM base
            WHERE {where_sql}
            GROUP BY source_id, source_code, source_name
            ORDER BY deals_count DESC
            LIMIT 10
        """, params)

        top_revenue_rows = qall(conn, f"""
            WITH base AS (
                SELECT
                  d.deal_type_code,
                  d.status_code,
                  COALESCE(rd.code, ra.code) AS region_code,
                  c.source_id,
                  src.code as source_code,
                  src.name as source_name,
                  di.price,
                  di.qty,
                  CASE
                    WHEN d.deal_type_code = 'sale' THEN d.completed_at
                    ELSE COALESCE(di.purchase_at, d.created_at)
                  END AS activity_at
                FROM app.deal_items di
                JOIN app.deals d ON d.deal_id = di.deal_id
                LEFT JOIN app.accounts a ON a.account_id = di.account_id
                LEFT JOIN app.regions ra ON ra.region_id = a.region_id
                LEFT JOIN app.regions rd ON rd.region_id = d.region_id
                LEFT JOIN app.customers c ON c.customer_id = d.customer_id
                LEFT JOIN app.sources src ON src.source_id = c.source_id
            )
            SELECT
              source_id,
              source_code,
              source_name,
              COUNT(*) AS deals_count,
              COALESCE(SUM(price * qty), 0) AS revenue
            FROM base
            WHERE {where_sql}
            GROUP BY source_id, source_code, source_name
            ORDER BY revenue DESC
            LIMIT 10
        """, params)

        repeat_row = q1(conn, f"""
            WITH base AS (
                SELECT
                  d.deal_id,
                  d.status_code,
                  d.deal_type_code,
                  COALESCE(rd.code, ra.code) AS region_code,
                  c.source_id,
                  d.customer_id,
                  CASE
                    WHEN d.deal_type_code = 'sale' THEN d.completed_at
                    ELSE COALESCE(di.purchase_at, d.created_at)
                  END AS activity_at
                FROM app.deal_items di
                JOIN app.deals d ON d.deal_id = di.deal_id
                LEFT JOIN app.accounts a ON a.account_id = di.account_id
                LEFT JOIN app.regions ra ON ra.region_id = a.region_id
                LEFT JOIN app.regions rd ON rd.region_id = d.region_id
                LEFT JOIN app.customers c ON c.customer_id = d.customer_id
            ),
            filtered AS (
                SELECT DISTINCT deal_id, customer_id
                FROM base
                WHERE {where_sql}
            ),
            by_customer AS (
                SELECT customer_id, COUNT(*) AS deals_count
                FROM filtered
                WHERE customer_id IS NOT NULL
                GROUP BY customer_id
            )
            SELECT
              COALESCE(SUM(CASE WHEN deals_count > 1 THEN 1 ELSE 0 END), 0) AS repeat_count,
              COALESCE(COUNT(*), 0) AS total_customers
            FROM by_customer
        """, params)

    top_by_count = [
        SourceAnalyticsItem(
            source_id=r[0],
            source_code=r[1],
            source_name=r[2],
            deals_count=int(r[3] or 0),
            revenue=float(r[4] or 0),
        )
        for r in top_count_rows
    ]
    top_by_revenue = [
        SourceAnalyticsItem(
            source_id=r[0],
            source_code=r[1],
            source_name=r[2],
            deals_count=int(r[3] or 0),
            revenue=float(r[4] or 0),
        )
        for r in top_revenue_rows
    ]
    repeat_count = int(repeat_row[0] or 0) if repeat_row else 0
    total_customers = int(repeat_row[1] or 0) if repeat_row else 0
    repeat_share = (repeat_count / total_customers) if total_customers else 0.0

    return SourcesAnalyticsOut(
        top_by_count=top_by_count,
        top_by_revenue=top_by_revenue,
        repeat_customers=RepeatCustomersOut(
            repeat_count=repeat_count,
            total_customers=total_customers,
            repeat_share=repeat_share,
        ),
    )

@app.get("/analytics/audit", response_model=AuditAnalyticsOut)
def analytics_audit(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    deal_type_code: Optional[str] = None,
    region_code: Optional[str] = None,
    source_id: Optional[int] = None,
    limit: int = 200,
    user: UserOut = Depends(get_current_user),
):
    if limit < 1 or limit > 500:
        raise HTTPException(400, "limit must be between 1 and 500")

    params: List[Any] = []
    filters = ["da.changed_at IS NOT NULL"]

    if deal_type_code:
        filters.append("d.deal_type_code = %s")
        params.append(deal_type_code)
    if region_code:
        filters.append("COALESCE(rd.code, ra.code) = %s")
        params.append(region_code)
    if source_id:
        filters.append("c.source_id = %s")
        params.append(source_id)
    if date_from:
        filters.append("da.changed_at::date >= %s")
        params.append(date_from)
    if date_to:
        filters.append("da.changed_at::date <= %s")
        params.append(date_to)

    where_sql = " AND ".join(filters)

    with psycopg.connect(DB_DSN) as conn:
        total_row = q1(conn, f"""
            SELECT COUNT(*)
            FROM app.deal_audit da
            LEFT JOIN app.deals d ON d.deal_id = da.deal_id
            LEFT JOIN LATERAL (
              SELECT account_id
              FROM app.deal_items
              WHERE deal_id = d.deal_id
              ORDER BY deal_item_id ASC
              LIMIT 1
            ) di ON true
            LEFT JOIN app.accounts a ON a.account_id = di.account_id
            LEFT JOIN app.regions ra ON ra.region_id = a.region_id
            LEFT JOIN app.regions rd ON rd.region_id = d.region_id
            LEFT JOIN app.customers c ON c.customer_id = d.customer_id
            WHERE {where_sql}
        """, params)
        total = int(total_row[0]) if total_row else 0

        rows = qall(conn, f"""
            SELECT
              da.deal_id,
              da.table_name,
              da.action,
              da.changed_at,
              da.changed_by
            FROM app.deal_audit da
            LEFT JOIN app.deals d ON d.deal_id = da.deal_id
            LEFT JOIN LATERAL (
              SELECT account_id
              FROM app.deal_items
              WHERE deal_id = d.deal_id
              ORDER BY deal_item_id ASC
              LIMIT 1
            ) di ON true
            LEFT JOIN app.accounts a ON a.account_id = di.account_id
            LEFT JOIN app.regions ra ON ra.region_id = a.region_id
            LEFT JOIN app.regions rd ON rd.region_id = d.region_id
            LEFT JOIN app.customers c ON c.customer_id = d.customer_id
            WHERE {where_sql}
            ORDER BY da.changed_at DESC
            LIMIT %s
        """, params + [limit])

    items = [
        AuditItemOut(
            deal_id=r[0],
            table_name=r[1],
            action=r[2],
            changed_at=r[3],
            changed_by=r[4],
        )
        for r in rows
    ]

    return AuditAnalyticsOut(total=total, items=items)
