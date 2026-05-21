from io import BytesIO
import threading
import uuid
import base64
from collections import defaultdict
from datetime import datetime, timezone, timedelta

from fastapi import Depends, File, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook


def mount_accounts_import_routes(
    app,
    *,
    DB_DSN,
    psycopg,
    read_accounts_from_excel,
    validate_account_import_rows,
    set_import_progress,
    get_import_progress,
    run_accounts_import_job,
    build_import_report_xlsx,
    require_role,
    UserOut,
    ImportReportIn,
):
    # Нормализует строковые значения из ячеек для проверок импорта.
    def _norm(value):
        return str(value or "").strip()

    # Разбирает почту в формате login@domain.
    def _split_account(value):
        raw = _norm(value)
        if "@" not in raw:
            return None, None
        login, domain = raw.split("@", 1)
        login = login.strip()
        domain = domain.strip()
        if not login or not domain:
            return None, None
        return login, domain

    # Проверяет только лист "Пользователи" на полноту слотов по аккаунтам.
    # Правило:
    # - если у аккаунта встречается PS4-группа, требуем П4 + PS4(1) + PS4(2);
    # - если встречается PS5-группа, требуем П5 + PS5(1) + PS5(2).
    def _validate_users_slot_matrix(content: bytes):
        wb = load_workbook(BytesIO(content), data_only=True)
        ws = None
        for sheet_name in wb.sheetnames:
            if _norm(sheet_name).lower() == "пользователи":
                ws = wb[sheet_name]
                break
        if ws is None:
            return {
                "ok": False,
                "errors": [{"row": 0, "field": "Лист", "value": "Пользователи", "message": "Лист 'Пользователи' не найден"}],
                "warnings": [],
                "total": 0,
            }

        header_alias = {
            "почта": "account",
            "аккаунт": "account",
            "платформа": "slot",
            "слот": "slot",
            "статус": "status",
        }
        headers = []
        for cell in ws[1]:
            headers.append(header_alias.get(_norm(cell.value).lower(), _norm(cell.value).lower()))

        try:
            idx_account = headers.index("account")
            idx_slot = headers.index("slot")
        except ValueError:
            return {
                "ok": False,
                "errors": [{"row": 1, "field": "Заголовки", "value": ", ".join(headers), "message": "Нужны колонки: Почта/Аккаунт и Платформа/Слот"}],
                "warnings": [],
                "total": 0,
            }
        idx_status = headers.index("status") if "status" in headers else None

        known_slots = {"п4", "п5", "ps4(1)", "ps4(2)", "ps5(1)", "ps5(2)"}
        account_slots = {}
        # Собираем по аккаунту состояние каждого слота (свободен/занят) и строки, чтобы ловить конфликт дублирования.
        account_slot_state = {}
        account_first_row = {}
        warnings = []
        errors = []
        checked_rows = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            checked_rows += 1
            account_val = _norm(row[idx_account] if idx_account < len(row) else "")
            slot_val = _norm(row[idx_slot] if idx_slot < len(row) else "").lower()
            status_val = _norm(row[idx_status] if idx_status is not None and idx_status < len(row) else "").lower()
            is_free = status_val == "свободен"
            login, domain = _split_account(account_val)
            if not login or not domain:
                warnings.append({
                    "row": row_idx,
                    "field": "Аккаунт",
                    "value": account_val,
                    "message": "Нужно значение в формате login@domain",
                })
                continue
            if slot_val not in known_slots:
                warnings.append({
                    "row": row_idx,
                    "field": "Платформа",
                    "value": slot_val,
                    "message": "Неизвестное значение платформы/слота",
                })
                continue

            account_key = f"{login.lower()}@{domain.lower()}"
            account_slots.setdefault(account_key, set()).add(slot_val)
            slot_state = account_slot_state.setdefault(
                account_key,
                defaultdict(lambda: {"busy_count": 0, "busy_rows": [], "free_count": 0, "free_rows": []}),
            )
            if is_free:
                slot_state[slot_val]["free_count"] += 1
                slot_state[slot_val]["free_rows"].append(row_idx)
            else:
                slot_state[slot_val]["busy_count"] += 1
                slot_state[slot_val]["busy_rows"].append(row_idx)
            account_first_row.setdefault(account_key, row_idx)

        missing_order = ["п4", "п5", "ps4(1)", "ps4(2)", "ps5(1)", "ps5(2)"]
        for account_key, slots in account_slots.items():
            # Ловим только конфликт: занятый слот указан повторно, хотя в группе есть свободный альтернативный слот.
            duplicate_conflicts = []
            slot_state = account_slot_state.get(account_key, {})
            sibling_slots = {
                "ps4(1)": ["ps4(2)"],
                "ps4(2)": ["ps4(1)"],
                "ps5(1)": ["ps5(2)"],
                "ps5(2)": ["ps5(1)"],
            }
            for slot in ["ps4(1)", "ps4(2)", "ps5(1)", "ps5(2)"]:
                state = slot_state.get(slot) or {}
                busy_count = int(state.get("busy_count") or 0)
                if busy_count <= 1:
                    continue
                alternatives = sibling_slots.get(slot, [])
                free_alternatives = []
                for alt_slot in alternatives:
                    alt_state = slot_state.get(alt_slot) or {}
                    if int(alt_state.get("free_count") or 0) > 0:
                        free_alternatives.append(alt_slot.upper())
                if not free_alternatives:
                    continue
                busy_rows = state.get("busy_rows") or []
                rows_text = ", ".join(str(r) for r in busy_rows)
                duplicate_conflicts.append(
                    f"{slot.upper()} (строки: {rows_text}; свободный: {', '.join(free_alternatives)})"
                    if rows_text
                    else f"{slot.upper()} (свободный: {', '.join(free_alternatives)})"
                )
            if duplicate_conflicts:
                warnings.append({
                    "row": account_first_row.get(account_key, 0),
                    "field": "Слоты",
                    "value": account_key,
                    "message": f"Конфликт слотов: повторно занят слот при наличии свободного альтернативного — {', '.join(duplicate_conflicts)}",
                })

            has_ps4_group = any(s in slots for s in {"п4", "ps4(1)", "ps4(2)"})
            has_ps5_group = any(s in slots for s in {"п5", "ps5(1)", "ps5(2)"})
            required = set()
            if has_ps4_group:
                required.update({"п4", "ps4(1)", "ps4(2)"})
            if has_ps5_group:
                required.update({"п5", "ps5(1)", "ps5(2)"})
            if not required:
                continue
            missing = [slot for slot in missing_order if slot in required and slot not in slots]
            if not missing:
                continue
            warnings.append({
                "row": account_first_row.get(account_key, 0),
                "field": "Слоты",
                "value": account_key,
                "message": f"Не хватает слотов: {', '.join(slot.upper() for slot in missing)}",
            })

        return {
            "ok": len(errors) == 0 and len(warnings) == 0,
            "errors": errors,
            "warnings": warnings,
            "total": checked_rows,
        }

    @app.get("/accounts/import/template")
    def accounts_import_template(user: UserOut = Depends(require_role("admin", "owner"))):
        wb = Workbook()
        ws = wb.active
        ws.title = "Почты"
        ws.append(["Логин", "Пароль"] + [f"Резерв {idx}" for idx in range(1, 11)])
        # Добавляем второй лист для привязки игр к существующим аккаунтам.
        ws_bindings = wb.create_sheet("Аккаунты")
        ws_bindings.append(["Почта", "Игры"])
        # Листы сроков подписок: используем тот же файл, чтобы прогон был единым.
        ws_plus = wb.create_sheet("ПЛЮС")
        ws_plus.append(["Аккаунт", "Подписка", "Срок"])
        ws_ea = wb.create_sheet("EA PLAY")
        ws_ea.append(["Аккаунт", "Подписка", "Срок"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=accounts_import_template.xlsx"},
        )

    @app.get("/accounts/import/status")
    def accounts_import_status(job_id: str, user: UserOut = Depends(require_role("admin", "owner"))):
        status = get_import_progress(job_id)
        if not status:
            return {"phase": "idle", "current": 0, "total": 0, "done": True}
        if status.get("owner") and status.get("owner") != user.username:
            raise HTTPException(403, "job not found")
        return {
            "phase": status.get("phase", "idle"),
            "current": int(status.get("current") or 0),
            "total": int(status.get("total") or 0),
            "done": bool(status.get("done")),
            "result": status.get("result"),
        }

    @app.post("/accounts/import/validate")
    def accounts_import_validate(file: UploadFile = File(...), user: UserOut = Depends(require_role("admin", "owner"))):
        if not file or not file.filename:
            raise HTTPException(400, "file is required")
        if not file.filename.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(400, "Only .xlsx/.xls are supported")
        content = file.file.read()
        if len(content) > 5 * 1024 * 1024:
            raise HTTPException(400, "File too large. Max 5MB")
        with psycopg.connect(DB_DSN) as conn:
            rows = read_accounts_from_excel(content)
            errors, warnings = validate_account_import_rows(conn, rows, progress_cb=None)
        return {"ok": len(errors) == 0, "total": len(rows), "errors": errors, "warnings": warnings}

    @app.post("/accounts/import/slots-check")
    def accounts_import_slots_check(file: UploadFile = File(...), user: UserOut = Depends(require_role("admin", "owner"))):
        if not file or not file.filename:
            raise HTTPException(400, "file is required")
        if not file.filename.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(400, "Only .xlsx/.xls are supported")
        content = file.file.read()
        if len(content) > 5 * 1024 * 1024:
            raise HTTPException(400, "File too large. Max 5MB")
        return _validate_users_slot_matrix(content)

    @app.post("/accounts/import")
    def accounts_import(file: UploadFile = File(...), user: UserOut = Depends(require_role("admin", "owner"))):
        if not file or not file.filename:
            raise HTTPException(400, "file is required")
        if not file.filename.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(400, "Only .xlsx/.xls are supported")
        content = file.file.read()
        if len(content) > 5 * 1024 * 1024:
            raise HTTPException(400, "File too large. Max 5MB")
        job_id = uuid.uuid4().hex
        set_import_progress(job_id, user.username, {"phase": "queued", "current": 0, "total": 0, "done": False})
        thread = threading.Thread(target=run_accounts_import_job, args=(job_id, content, user.username), daemon=True)
        thread.start()
        return {"ok": True, "job_id": job_id}

    @app.post("/accounts/import/cancel")
    def accounts_import_cancel(job_id: str, user: UserOut = Depends(require_role("admin", "owner"))):
        status = get_import_progress(job_id)
        if not status:
            return {"ok": True}
        if status.get("owner") and status.get("owner") != user.username:
            raise HTTPException(403, "job not found")
        set_import_progress(
            job_id,
            user.username,
            {
                "phase": "cancelled",
                "current": int(status.get("current") or 0),
                "total": int(status.get("total") or 0),
                "done": True,
                "cancelled": True,
                "result": {"ok": False, "cancelled": True},
            },
        )
        return {"ok": True}

    @app.post("/accounts/import/report")
    def accounts_import_report(payload: ImportReportIn, user: UserOut = Depends(require_role("admin", "owner"))):
        content = build_import_report_xlsx(
            [i.dict() for i in payload.errors],
            [i.dict() for i in payload.warnings],
        )
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=accounts_import_report.xlsx"},
        )

    # Декодирует секреты аккаунтов из base64, чтобы экспорт можно было сразу использовать как импорт.
    def _decode_secret(value):
        raw = str(value or "").strip()
        if not raw:
            return ""
        try:
            return base64.b64decode(raw.encode("utf-8")).decode("utf-8")
        except Exception:
            return raw

    # Формирует строку источника в формате, который понимает импорт слотов.
    def _format_source(name, code):
        name_val = str(name or "").strip()
        code_val = str(code or "").strip()
        if name_val and code_val:
            return f"{name_val} {code_val.upper()}"
        return name_val or code_val

    @app.get("/accounts/import/export")
    def accounts_import_export(user: UserOut = Depends(require_role("admin", "owner"))):
        wb = Workbook()
        ws_games = wb.active
        ws_games.title = "Игры"
        ws_games.append(["Товар", "Платформа"])

        ws_mails = wb.create_sheet("Почты")
        ws_mails.append(["Логин", "Пароль"] + [f"Резерв {idx}" for idx in range(1, 11)])

        ws_bindings = wb.create_sheet("Аккаунты")
        ws_bindings.append(["Почта", "Игры"])

        ws_plus = wb.create_sheet("ПЛЮС")
        ws_plus.append(["Аккаунт", "Подписка", "Срок"])

        ws_ea = wb.create_sheet("EA PLAY")
        ws_ea.append(["Аккаунт", "Подписка", "Срок"])

        ws_users = wb.create_sheet("Пользователи")
        ws_users.append(["Почта", "Комментарий покупателя", "Дубль", "Статус", "Пользователь", "Откуда", "Покупка", "Платформа"])

        with psycopg.connect(DB_DSN) as conn:
            # В старых БД может не быть колонки accounts.is_archived, поэтому условие собираем динамически.
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT 1
                    FROM information_schema.columns
                    WHERE table_schema='app' AND table_name='accounts' AND column_name='is_archived'
                    LIMIT 1
                    """
                )
                has_account_is_archived = cur.fetchone() is not None
            account_active_clause = "a.is_archived IS NOT TRUE" if has_account_is_archived else "TRUE"

            # Выгружаем игры в формате импорта товаров.
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT
                      p.title,
                      COALESCE(string_agg(DISTINCT lower(pl.code), ', ' ORDER BY lower(pl.code)), '')
                    FROM app.products p
                    LEFT JOIN app.product_platforms pp ON pp.product_id = p.product_id
                    LEFT JOIN app.platforms pl ON pl.platform_id = pp.platform_id
                    WHERE lower(p.type_code) = 'game'
                      AND p.is_archived IS NOT TRUE
                    GROUP BY p.product_id, p.title
                    ORDER BY lower(p.title), p.product_id
                    """
                )
                for title, platforms in cur.fetchall():
                    ws_games.append([title or "", platforms or ""])

            # Выгружаем аккаунты и резервы в формате листа Почты.
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT
                      a.login_name,
                      max(CASE WHEN s.secret_key = 'account_password' THEN s.secret_value END) AS account_password,
                      max(CASE WHEN s.secret_key = 'reserve1' THEN s.secret_value END) AS reserve1,
                      max(CASE WHEN s.secret_key = 'reserve2' THEN s.secret_value END) AS reserve2,
                      max(CASE WHEN s.secret_key = 'reserve3' THEN s.secret_value END) AS reserve3,
                      max(CASE WHEN s.secret_key = 'reserve4' THEN s.secret_value END) AS reserve4,
                      max(CASE WHEN s.secret_key = 'reserve5' THEN s.secret_value END) AS reserve5,
                      max(CASE WHEN s.secret_key = 'reserve6' THEN s.secret_value END) AS reserve6,
                      max(CASE WHEN s.secret_key = 'reserve7' THEN s.secret_value END) AS reserve7,
                      max(CASE WHEN s.secret_key = 'reserve8' THEN s.secret_value END) AS reserve8,
                      max(CASE WHEN s.secret_key = 'reserve9' THEN s.secret_value END) AS reserve9,
                      max(CASE WHEN s.secret_key = 'reserve10' THEN s.secret_value END) AS reserve10
                    FROM app.accounts a
                    LEFT JOIN app.account_secrets s ON s.account_id = a.account_id
                    WHERE {account_active_clause}
                    GROUP BY a.account_id, a.login_name
                    ORDER BY lower(a.login_name), a.account_id
                    """
                )
                for row in cur.fetchall():
                    login_name = row[0] or ""
                    decoded = [_decode_secret(value) for value in row[1:]]
                    ws_mails.append([login_name, *decoded])

            # Выгружаем связки аккаунт-игра в формате листа Аккаунты.
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT
                      a.login_name || '@' || d.name AS account_full,
                      p.title
                    FROM app.account_assets aa
                    JOIN app.accounts a ON a.account_id = aa.account_id
                    JOIN app.domains d ON d.domain_id = a.domain_id
                    JOIN app.products p ON p.product_id = aa.product_id
                    WHERE aa.asset_type_code = 'game'
                      AND {account_active_clause}
                      AND p.is_archived IS NOT TRUE
                    ORDER BY lower(a.login_name || '@' || d.name), lower(p.title), aa.account_asset_id
                    """
                )
                for account_full, title in cur.fetchall():
                    ws_bindings.append([account_full or "", title or ""])

            # Выгружаем сроки подписок по листам ПЛЮС и EA PLAY.
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT
                      a.login_name || '@' || d.name AS account_full,
                      p.title,
                      st.valid_until
                    FROM app.subscription_terms st
                    JOIN app.accounts a ON a.account_id = st.account_id
                    JOIN app.domains d ON d.domain_id = a.domain_id
                    JOIN app.products p ON p.product_id = st.product_id
                    WHERE st.is_archived IS NOT TRUE
                      AND {account_active_clause}
                      AND p.is_archived IS NOT TRUE
                    ORDER BY lower(a.login_name || '@' || d.name), st.valid_until, p.product_id, st.term_id
                    """
                )
                for account_full, title, valid_until in cur.fetchall():
                    valid_str = valid_until.isoformat() if valid_until else ""
                    title_val = str(title or "").strip()
                    if title_val.lower() == "ea play":
                        ws_ea.append([account_full or "", "EA PLAY", valid_str])
                    else:
                        ws_plus.append([account_full or "", title_val, valid_str])

            # Выгружаем историю слотов в формате листа Пользователи для повторного импорта.
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT
                      a.login_name || '@' || d.name AS account_full,
                      c.nickname,
                      src.name,
                      src.code,
                      coalesce(di.start_at, asa.assigned_at) AS purchase_at,
                      asa.released_at,
                      asa.slot_type_code,
                      st.capacity,
                      asa.assignment_id
                    FROM app.account_slot_assignments asa
                    JOIN app.accounts a ON a.account_id = asa.account_id
                    JOIN app.domains d ON d.domain_id = a.domain_id
                    LEFT JOIN app.slot_types st ON st.code = asa.slot_type_code
                    LEFT JOIN app.customers c ON c.customer_id = asa.customer_id
                    LEFT JOIN app.sources src ON src.source_id = c.source_id
                    LEFT JOIN app.deal_items di ON di.deal_item_id = asa.deal_item_id
                    WHERE {account_active_clause}
                    ORDER BY lower(a.login_name || '@' || d.name), asa.slot_type_code, coalesce(di.start_at, asa.assigned_at), asa.assignment_id
                    """
                )
                raw_rows = cur.fetchall()
                slot_map = {
                    "activate_ps4": "П4",
                    "activate_ps5": "П5",
                    "play_ps4": "PS4",
                    "play_ps5": "PS5",
                }
                grouped = defaultdict(list)
                for row in raw_rows:
                    slot_type_code = str(row[6] or "").strip().lower()
                    grouped[(str(row[0] or "").strip().lower(), slot_type_code)].append(row)

                # Восстанавливаем номер инстанса (1/2) по временной шкале назначений.
                # Это нужно, чтобы round-trip экспорт->импорт не сваливал все play-слоты в (1).
                far_future = datetime.max.replace(tzinfo=timezone.utc) - timedelta(days=1)
                for _, rows_group in grouped.items():
                    rows_group.sort(key=lambda x: (x[4] or x[5] or datetime.min.replace(tzinfo=timezone.utc), x[8]))
                    slot_type_code = str(rows_group[0][6] or "").strip().lower() if rows_group else ""
                    base_label = slot_map.get(slot_type_code, "")
                    capacity_raw = rows_group[0][7] if rows_group else None
                    capacity = int(capacity_raw) if isinstance(capacity_raw, int) and capacity_raw > 0 else 1
                    lane_until = [None] * capacity
                    exported = []
                    for account_full, nickname, source_name, source_code, purchase_at, released_at, slot_type_code_row, _capacity_row, _assignment_id in rows_group:
                        assigned_at = purchase_at or released_at
                        if assigned_at is None:
                            continue

                        # Освобождаем дорожки, где прошлое назначение уже завершилось к моменту текущего.
                        for idx, busy_until in enumerate(lane_until):
                            if busy_until is not None and busy_until <= assigned_at:
                                lane_until[idx] = None

                        lane_idx = None
                        for idx, busy_until in enumerate(lane_until):
                            if busy_until is None:
                                lane_idx = idx
                                break
                        if lane_idx is None:
                            # Если все дорожки заняты (грязная история), берем дорожку с самым ранним окончанием.
                            min_end_idx = 0
                            min_end_val = lane_until[0]
                            for idx in range(1, len(lane_until)):
                                curr = lane_until[idx]
                                if min_end_val is None:
                                    min_end_idx = idx
                                    min_end_val = curr
                                    continue
                                if curr is not None and curr < min_end_val:
                                    min_end_idx = idx
                                    min_end_val = curr
                            lane_idx = min_end_idx

                        # Открытое назначение (released_at is NULL) блокирует дорожку до "бесконечности".
                        lane_until[lane_idx] = released_at if released_at is not None else far_future
                        if slot_type_code_row in {"play_ps4", "play_ps5"}:
                            platform_value = f"{base_label}({lane_idx + 1})"
                        else:
                            platform_value = base_label

                        date_value = assigned_at.date().isoformat()
                        exported.append([
                            account_full or "",
                            "",
                            "",
                            "ЗАНЯТ",
                            nickname or "",
                            _format_source(source_name, source_code),
                            date_value,
                            platform_value,
                        ])
                    for row_out in exported:
                        ws_users.append(row_out)

        for ws in wb.worksheets:
            for col_idx in range(1, 9):
                ws.column_dimensions[chr(64 + col_idx)].width = 22

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=gamesales_import_export.xlsx"},
        )
