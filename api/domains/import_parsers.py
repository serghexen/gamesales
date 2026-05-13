from dataclasses import dataclass
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any, Callable
import os
import re
import ssl
import urllib.error
import urllib.request

from fastapi import HTTPException
from openpyxl import load_workbook


@dataclass
class ImportParsers:
    GAME_IMPORT_HEADERS: list[str]
    parse_import_platforms: Callable[[str], list[str]]
    validate_game_import_rows: Callable[..., tuple[list[dict[str, Any]], list[dict[str, Any]]]]
    read_games_from_excel: Callable[[bytes], list[dict[str, Any]]]
    read_accounts_from_excel: Callable[[bytes], list[dict[str, Any]]]
    read_slots_from_excel: Callable[[bytes], list[dict[str, Any]]]
    clean_slots_excel: Callable[[bytes], bytes]
    split_account: Callable[[str], tuple[str | None, str | None]]
    validate_account_import_rows: Callable[..., tuple[list[dict[str, Any]], list[dict[str, Any]]]]
    SLOT_FILE_TO_TYPE: dict[str, tuple[str, int]]
    normalize_slot_file_value: Callable[[str], str]
    normalize_cell_text: Callable[[Any], str]
    normalize_slot_date: Callable[[Any], str | None]
    normalize_slot_date_to_dt: Callable[[Any], datetime | None]
    resolve_source_id: Callable[[Any, str], int | None]
    validate_slot_import_rows: Callable[..., tuple[list[dict[str, Any]], list[dict[str, Any]], int]]
    detect_logo_mime_from_bytes: Callable[[bytes], str]
    fetch_logo_from_url: Callable[[str], tuple[bytes, str]]


def build_import_parsers(*, q1, qall, normalize_platform_codes):
    max_logo_bytes = 5 * 1024 * 1024
    allowed_logo_mime = {"image/jpeg", "image/png", "image/webp"}
    # Локальный лимит строк для листов Почты*: помогает быстро гонять тестовые прогоны.
    accounts_import_limit_raw = str(os.getenv("ACCOUNTS_IMPORT_LIMIT_PER_SHEET", "") or "").strip()
    accounts_import_limit = int(accounts_import_limit_raw) if accounts_import_limit_raw.isdigit() and int(accounts_import_limit_raw) > 0 else None

    game_import_headers = ["Товар", "Платформа"]
    game_import_header_map = {
        "Товар": "title",
        "товар": "title",
        "Игра": "title",
        "игра": "title",
        "Игры": "title",
        "игры": "title",
        "Платформа": "platform_codes",
        "title": "title",
        "platform_codes": "platform_codes",
    }
    account_import_header_map = {
        "Аккаунт": "account",
        "аккаунт": "account",
        "Login": "account",
        "Логин": "account",
        "логин": "account",
        "Почта": "account",
        "почта": "account",
        "Пароль": "password",
        "пароль": "password",
        "Password": "password",
        "Товар": "game",
        "товар": "game",
        "Игра": "game",
        "игра": "game",
        "Игры": "game",
        "игры": "game",
        "Game": "game",
        "Подписка": "subscription",
        "подписка": "subscription",
        "Название подписки": "subscription",
        "название подписки": "subscription",
        "Срок": "valid_until",
        "срок": "valid_until",
        "Дата": "valid_until",
        "дата": "valid_until",
        "Активна до": "valid_until",
        "активна до": "valid_until",
        "Valid Until": "valid_until",
    }
    slot_import_header_map = {
        "Аккаунт": "account",
        "аккаунт": "account",
        "Почта": "account",
        "почта": "account",
        "Товар": "game",
        "товар": "game",
        "Игра": "game",
        "игра": "game",
        "Статус": "status",
        "статус": "status",
        "Пользователь": "customer",
        "пользователь": "customer",
        "Источник": "source",
        "источник": "source",
        "Откуда": "source",
        "откуда": "source",
        "Дата": "date",
        "дата": "date",
        "Покупка": "date",
        "покупка": "date",
        "Слот": "slot",
        "слот": "slot",
        "Платформа": "slot",
        "платформа": "slot",
    }
    slot_file_to_type = {
        "п4": ("activate_ps4", 1),
        "п5": ("activate_ps5", 1),
        "ps4(1)": ("play_ps4", 1),
        "ps4(2)": ("play_ps4", 2),
        "ps5(1)": ("play_ps5", 1),
        "ps5(2)": ("play_ps5", 2),
    }
    ru_months = {
        "январь": 1,
        "февраль": 2,
        "март": 3,
        "апрель": 4,
        "май": 5,
        "июнь": 6,
        "июль": 7,
        "август": 8,
        "сентябрь": 9,
        "октябрь": 10,
        "ноябрь": 11,
        "декабрь": 12,
    }

    def parse_import_platforms(value: str) -> list[str]:
        if not value:
            return []
        raw = str(value).strip()
        if raw in {"-", "—"}:
            return []
        parts = []
        for chunk in raw.replace(";", ",").split(","):
            val = chunk.strip().lower()
            if val:
                parts.append(val)
        return normalize_platform_codes(parts)

    # Нормализует заголовки импорта аккаунтов и распознает колонки резервов.
    def normalize_account_import_header(raw_header: Any) -> str:
        if raw_header is None:
            return ""
        raw = str(raw_header).strip()
        if not raw:
            return ""
        mapped = account_import_header_map.get(raw)
        if mapped:
            return mapped
        reserve_match = re.match(r"^(?:резерв|reserve)\s*([1-9]|10)$", raw, flags=re.IGNORECASE)
        if reserve_match:
            return f"reserve{int(reserve_match.group(1))}"
        return raw

    # Проверяет, относится ли строка к подписке/плюсу, чтобы пропустить ее в импорте игр.
    def is_subscription_like_game_title(value: str) -> bool:
        title = str(value or "").strip().lower()
        if not title:
            return False
        return "подписк" in title or "плюс" in title

    # Проверяет, что товар (игра или подписка) существует в каталоге.
    def resolve_slot_product_meta_by_title(conn, product_title: str):
        if not product_title:
            return None
        row = q1(
            conn,
            """
            SELECT p.product_id, lower(p.type_code)
            FROM app.products p
            WHERE lower(p.title) = lower(%s)
              AND lower(p.type_code) IN ('game', 'subscription')
              AND p.is_archived IS NOT TRUE
            ORDER BY p.product_id
            LIMIT 1
            """,
            (product_title,),
        )
        if not row or row[0] is None:
            return None
        # Поддерживаем старые тестовые моки, где возвращается только product_id без type_code.
        type_code = str(row[1] or "").strip().lower() if len(row) > 1 else "game"
        return {"product_id": int(row[0]), "type_code": type_code}

    def detect_logo_mime_from_bytes(data: bytes) -> str:
        if not data:
            return ""
        if data.startswith(b"\xff\xd8\xff"):
            return "image/jpeg"
        if data.startswith(b"\x89PNG\r\n\x1a\n"):
            return "image/png"
        if len(data) >= 12 and data[:4] == b"RIFF" and data[8:12] == b"WEBP":
            return "image/webp"
        return ""

    def fetch_logo_from_url(url: str) -> tuple[bytes, str]:
        if not url:
            raise ValueError("logo url is empty")
        context = ssl.create_default_context()
        context.check_hostname = False
        context.verify_mode = ssl.CERT_NONE
        req = urllib.request.Request(url, headers={"User-Agent": "gamesales-import"})
        try:
            with urllib.request.urlopen(req, timeout=10, context=context) as resp:
                content_type = (resp.headers.get("Content-Type") or "").split(";")[0].strip().lower()
                data = resp.read(max_logo_bytes + 1)
        except urllib.error.URLError as exc:
            raise ValueError(f"logo download failed: {exc}") from exc
        if len(data) > max_logo_bytes:
            raise ValueError("logo too large")
        mime = content_type if content_type in allowed_logo_mime else ""
        if not mime:
            mime = detect_logo_mime_from_bytes(data)
        if mime not in allowed_logo_mime:
            raise ValueError("logo type not allowed")
        return data, mime

    def validate_game_import_rows(conn, rows: list[dict], progress_cb=None, check_logo=False):
        errors = []
        warnings = []
        platform_rows = qall(conn, "SELECT code FROM app.platforms")
        platforms = {str(r[0]).strip().lower() for r in platform_rows}
        for idx, row in enumerate(rows, start=2):
            report_row = idx - 1
            title = (row.get("title") or "").strip()
            # Строки подписок в листе "Игры" пропускаем, чтобы не плодить нецелевые товары.
            if bool(row.get("skip_for_import")):
                warnings.append({
                    "row": report_row,
                    "field": "Товар",
                    "value": title,
                    "message": "Строка с подпиской/плюсом пропущена",
                })
                if progress_cb:
                    progress_cb(idx - 1)
                continue
            platform_raw = row.get("platform_codes") or ""
            platform_codes = parse_import_platforms(platform_raw)
            if not title:
                errors.append({"row": report_row, "field": "Товар", "value": title, "message": "Название обязательно"})
            if not platform_codes:
                warnings.append({"row": report_row, "field": "Платформа", "value": str(platform_raw).strip(), "message": "Платформы не указаны — строка будет пропущена"})
            for code in platform_codes:
                if code not in platforms:
                    errors.append({"row": report_row, "field": "Платформа", "value": code, "message": f"Неизвестная платформа: {code}"})
            if progress_cb:
                progress_cb(idx - 1)
        return errors, warnings

    def read_games_from_excel(content: bytes) -> list[dict]:
        wb = load_workbook(BytesIO(content), data_only=True)
        # Если в файле есть лист "Игры", читаем его приоритетно; иначе оставляем active для совместимости.
        ws = wb.active
        for sheet_name in wb.sheetnames:
            if str(sheet_name or "").strip().lower() == "игры":
                ws = wb[sheet_name]
                break
        headers = []
        for cell in ws[1]:
            if cell.value is None:
                headers.append("")
            else:
                raw = str(cell.value).strip()
                headers.append(game_import_header_map.get(raw, raw))
        data = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            item = {}
            for idx, key in enumerate(headers):
                if not key:
                    continue
                item[key] = row[idx] if idx < len(row) else None
            # Помечаем строки подписок, чтобы валидация и импорт синхронно их пропускали.
            item["skip_for_import"] = is_subscription_like_game_title(item.get("title"))
            data.append(item)
        return data

    def read_accounts_from_excel(content: bytes, limit_per_sheet: int | None = None) -> list[dict]:
        wb = load_workbook(BytesIO(content), data_only=True)
        data = []
        # Даем возможность задать лимит через аргумент или env-переменную для локальных прогонов.
        effective_limit = limit_per_sheet if isinstance(limit_per_sheet, int) and limit_per_sheet > 0 else accounts_import_limit
        # Читаем все нужные листы и фиксируем порядок прогона: Почты* -> ПЛЮС/EA PLAY -> Аккаунты.
        mail_sheets = []
        term_sheets = []
        binding_sheets = []
        for sheet_name in wb.sheetnames:
            name_norm = str(sheet_name or "").strip().lower()
            if name_norm.startswith("почты"):
                mail_sheets.append(wb[sheet_name])
            elif name_norm in {"плюс", "ea play"}:
                term_sheets.append(wb[sheet_name])
            elif name_norm == "аккаунты":
                binding_sheets.append(wb[sheet_name])
        selected_sheets = [*mail_sheets, *term_sheets, *binding_sheets]
        if not selected_sheets:
            selected_sheets = [wb.active]
        for ws in selected_sheets:
            sheet_name_norm = str(ws.title or "").strip().lower()
            is_bindings_sheet = sheet_name_norm == "аккаунты"
            is_plus_sheet = sheet_name_norm == "плюс"
            is_ea_play_sheet = sheet_name_norm == "ea play"
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            first_row = rows[0]
            first_account = str(first_row[0] or "").strip() if len(first_row) > 0 else ""
            first_looks_like_data = bool(split_account(first_account)[0] and split_account(first_account)[1])
            # Без заголовка поддерживаем только листы ПЛЮС и EA PLAY.
            allow_no_header = is_plus_sheet or is_ea_play_sheet
            has_header = (not first_looks_like_data) or (not allow_no_header)

            if has_header:
                headers = [normalize_account_import_header(cell) for cell in rows[0]]
                iter_rows = list(enumerate(rows[1:], start=2))
            else:
                # Для листов ПЛЮС/EA PLAY без шапки берем фиксированный порядок колонок.
                if is_plus_sheet or is_ea_play_sheet:
                    headers = ["account", "subscription", "valid_until"]
                else:
                    headers = ["account", "password", "reserve1", "reserve2", "reserve3", "reserve4", "reserve5", "reserve6", "reserve7", "reserve8", "reserve9", "reserve10"]
                iter_rows = list(enumerate(rows, start=1))

            for row_idx, row in iter_rows:
                # Ограничиваем чтение верхними строками листа, чтобы локальный прогон был предсказуемым.
                if effective_limit is not None:
                    max_row_idx = (1 + effective_limit) if has_header else effective_limit
                    if row_idx > max_row_idx:
                        break
                if not any(row):
                    continue
                item = {}
                for idx, key in enumerate(headers):
                    if not key:
                        continue
                    item[key] = row[idx] if idx < len(row) else None
                # Для сроков подписок поддерживаем фиксированный формат колонок, даже если заголовки нестандартные.
                if is_plus_sheet or is_ea_play_sheet:
                    if "account" not in item and len(row) >= 1:
                        item["account"] = row[0]
                    if "subscription" not in item and len(row) >= 2:
                        item["subscription"] = row[1]
                    if "valid_until" not in item and len(row) >= 3:
                        item["valid_until"] = row[2]
                # Складываем непустые резервы в единый словарь, чтобы джоба одинаково их обрабатывала.
                reserve_values = {}
                for reserve_idx in range(1, 11):
                    reserve_key = f"reserve{reserve_idx}"
                    reserve_raw = item.get(reserve_key)
                    reserve_val = str(reserve_raw or "").strip()
                    if reserve_val:
                        reserve_values[reserve_key] = reserve_val
                item["reserve_values"] = reserve_values
                item["_sheet_name"] = ws.title
                item["_sheet_row"] = row_idx
                # Храним тип строки, чтобы джоба и валидация применяли нужную логику.
                if is_bindings_sheet:
                    item["_import_kind"] = "account_game_binding"
                elif is_plus_sheet:
                    item["_import_kind"] = "subscription_term_plus"
                elif is_ea_play_sheet:
                    item["_import_kind"] = "subscription_term_ea_play"
                else:
                    item["_import_kind"] = "account_credentials"
                data.append(item)
        return data

    # Приводит название подписки из листов к тому виду, как товар хранится в каталоге.
    def normalize_subscription_slot_title(raw_title: Any, sheet_name: str) -> str:
        title = str(raw_title or "").strip()
        if not title:
            return ""
        if str(sheet_name or "").strip().lower() == "ea play":
            return "EA PLAY"
        return title

    # Собирает кандидатов товара по аккаунту из вспомогательных листов файла.
    def build_account_slot_product_candidates(wb) -> dict[tuple[str, str], list[dict[str, Any]]]:
        candidates: dict[tuple[str, str], list[dict[str, Any]]] = {}

        def add_candidate(account_value: Any, title_value: Any, valid_until_value: Any = None):
            account_raw = str(account_value or "").strip()
            title_raw = str(title_value or "").strip()
            login, domain = split_account(account_raw)
            if not login or not domain or not title_raw:
                return
            key = (login.lower(), domain.lower())
            bucket = candidates.setdefault(key, [])
            normalized_valid_until = normalize_slot_date(valid_until_value)
            exists = any(
                str(item.get("title") or "").strip().lower() == title_raw.lower()
                and str(item.get("valid_until") or "").strip() == str(normalized_valid_until or "").strip()
                for item in bucket
            )
            if not exists:
                bucket.append({"title": title_raw, "valid_until": normalized_valid_until})

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            name_norm = str(sheet_name or "").strip().lower()
            if name_norm not in {"аккаунты", "плюс", "ea play"}:
                continue
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            # Листы могут быть без шапки (данные сразу с первой строки), поэтому определяем это эвристикой.
            first_row = rows[0]
            first_account = str(first_row[0] or "").strip() if len(first_row) > 0 else ""
            first_looks_like_data = bool(split_account(first_account)[0] and split_account(first_account)[1])
            # В вспомогательном разборе без заголовка поддерживаем только ПЛЮС/EA PLAY.
            allow_no_header = name_norm in {"плюс", "ea play"}
            has_header = (not first_looks_like_data) or (not allow_no_header)

            # Для листа "Аккаунты" используем account-заголовки, где колонка игры обычно называется "Игры".
            if has_header:
                if name_norm == "аккаунты":
                    headers = [normalize_account_import_header(cell) for cell in rows[0]]
                else:
                    headers = [slot_import_header_map.get(str(cell).strip(), str(cell).strip()) if cell is not None else "" for cell in rows[0]]
                data_rows = rows[1:]
            else:
                # Без заголовка используем фиксированную схему колонок, как в шаблоне импорта.
                headers = ["account", "subscription", "valid_until"]
                data_rows = rows

            for row in data_rows:
                if not any(row):
                    continue
                item = {}
                for idx, key in enumerate(headers):
                    if key:
                        item[key] = row[idx] if idx < len(row) else None
                if name_norm == "аккаунты":
                    add_candidate(item.get("account"), item.get("game"))
                else:
                    add_candidate(
                        item.get("account"),
                        normalize_subscription_slot_title(item.get("subscription"), sheet_name),
                        item.get("valid_until"),
                    )
        return candidates

    # Собирает аккаунты, которые будут созданы в этом же файле (листы Почты*), чтобы не ругаться на "не найден в БД".
    def build_staged_slot_accounts(wb) -> set[tuple[str, str]]:
        staged: set[tuple[str, str]] = set()
        for sheet_name in wb.sheetnames:
            name_norm = str(sheet_name or "").strip().lower()
            if not name_norm.startswith("почты"):
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [normalize_account_import_header(cell) for cell in rows[0]]
            for row in rows[1:]:
                if not any(row):
                    continue
                item = {}
                for idx, key in enumerate(headers):
                    if key:
                        item[key] = row[idx] if idx < len(row) else None
                login, domain = split_account(str(item.get("account") or "").strip())
                if login and domain:
                    staged.add((login.lower(), domain.lower()))
        return staged

    def read_slots_from_excel(content: bytes) -> list[dict]:
        wb = load_workbook(BytesIO(content), data_only=True)
        # Для импорта слотов приоритетно читаем лист "Пользователи", иначе используем active.
        ws = wb.active
        for sheet_name in wb.sheetnames:
            if str(sheet_name or "").strip().lower() == "пользователи":
                ws = wb[sheet_name]
                break
        account_product_candidates = build_account_slot_product_candidates(wb)
        staged_accounts = build_staged_slot_accounts(wb)
        headers = []
        for cell in ws[1]:
            if cell.value is None:
                headers.append("")
            else:
                raw = str(cell.value).strip()
                headers.append(slot_import_header_map.get(raw, raw))
        data = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            item = {}
            for idx, key in enumerate(headers):
                if not key:
                    continue
                item[key] = row[idx] if idx < len(row) else None
            # Если в листе нет явной колонки товара, пробуем взять его по аккаунту из листов Аккаунты/ПЛЮС/EA PLAY.
            if not str(item.get("game") or "").strip():
                login, domain = split_account(str(item.get("account") or "").strip())
                if login and domain:
                    key = (login.lower(), domain.lower())
                    candidates = account_product_candidates.get(key, [])
                    unique_titles = []
                    seen_titles = set()
                    for candidate in candidates:
                        title = str(candidate.get("title") or "").strip()
                        if not title:
                            continue
                        title_l = title.lower()
                        if title_l in seen_titles:
                            continue
                        seen_titles.add(title_l)
                        unique_titles.append(title)
                    if len(unique_titles) == 1:
                        selected_title = unique_titles[0]
                        item["game"] = selected_title
                        selected_title_l = selected_title.lower()
                        selected_dates = []
                        seen_dates = set()
                        for candidate in candidates:
                            if str(candidate.get("title") or "").strip().lower() != selected_title_l:
                                continue
                            valid_until = str(candidate.get("valid_until") or "").strip()
                            if not valid_until or valid_until in seen_dates:
                                continue
                            seen_dates.add(valid_until)
                            selected_dates.append(valid_until)
                        if len(selected_dates) == 1:
                            item["_subscription_valid_until"] = selected_dates[0]
                        elif len(selected_dates) > 1:
                            item["_subscription_term_dates"] = selected_dates
                    elif len(unique_titles) > 1:
                        # Сохраняем список кандидатов, чтобы валидация показала понятное предупреждение.
                        item["_game_candidates"] = unique_titles
                    if key in staged_accounts:
                        # Помечаем аккаунт как "будущий" из этого же файла, чтобы валидация не требовала наличие в БД.
                        item["_account_declared_in_file"] = True
            # Сохраняем реальный номер строки в листе для точного отчета предупреждений/ошибок.
            item["_sheet_row"] = row_idx
            data.append(item)
        return data

    def clean_slots_excel(content: bytes) -> bytes:
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        status_col = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value is None:
                continue
            header = str(cell.value).strip()
            if header.lower() == "статус":
                status_col = idx
                break
        if not status_col:
            raise HTTPException(400, 'Колонка "Статус" не найдена')
        for row_idx in range(ws.max_row, 1, -1):
            raw = ws.cell(row=row_idx, column=status_col).value
            value = "" if raw is None else str(raw).strip()
            if not value or value.lower() == "свободен":
                ws.delete_rows(row_idx, 1)
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    def split_account(value: str):
        if not value:
            return None, None
        raw = str(value).strip()
        if not raw or "@" not in raw:
            return None, None
        login, domain = raw.split("@", 1)
        login = login.strip()
        domain = domain.strip()
        if not login or not domain:
            return None, None
        return login, domain

    def validate_account_import_rows(conn, rows: list[dict], progress_cb=None):
        errors = []
        warnings = []
        # Собираем аккаунты из листов Почты* этого же файла, чтобы валидация сроков учитывала "будущие" вставки.
        staged_accounts: set[tuple[str, str]] = set()
        for staged_row in rows:
            staged_kind = str(staged_row.get("_import_kind") or "account_credentials")
            if staged_kind != "account_credentials":
                continue
            staged_account_val = (staged_row.get("account") or "").strip()
            staged_login, staged_domain = split_account(staged_account_val)
            if not staged_login or not staged_domain:
                continue
            staged_accounts.add((staged_login.lower(), staged_domain.lower()))
        for idx, row in enumerate(rows, start=2):
            report_row = int(row.get("_sheet_row") or (idx - 1))
            report_sheet = str(row.get("_sheet_name") or "").strip() or None
            row_kind = str(row.get("_import_kind") or "account_credentials")
            account_val = (row.get("account") or "").strip()
            login, domain = split_account(account_val)
            if row_kind == "account_game_binding":
                game_title = (row.get("game") or "").strip()
                if not account_val or not login or not domain:
                    warnings.append({
                        "sheet": report_sheet,
                        "row": report_row,
                        "field": "Почта",
                        "value": account_val,
                        "message": "Нужно значение в формате login@domain — строка будет пропущена",
                    })
                if not game_title:
                    warnings.append({
                        "sheet": report_sheet,
                        "row": report_row,
                        "field": "Игры",
                        "value": game_title,
                        "message": "Игра не указана — строка будет пропущена",
                    })
                elif resolve_slot_product_meta_by_title(conn, game_title) is None:
                    warnings.append({
                        "sheet": report_sheet,
                        "row": report_row,
                        "field": "Игры",
                        "value": game_title,
                        "message": "Игра не найдена — строка будет пропущена",
                    })
            elif row_kind in {"subscription_term_plus", "subscription_term_ea_play"}:
                subscription_title = (row.get("subscription") or "").strip()
                valid_until_raw = row.get("valid_until")
                valid_until_text = "" if valid_until_raw is None else str(valid_until_raw).strip()
                if not account_val or not login or not domain:
                    warnings.append({
                        "sheet": report_sheet,
                        "row": report_row,
                        "field": "Аккаунт",
                        "value": account_val,
                        "message": "Нужно значение в формате login@domain — строка будет пропущена",
                    })
                else:
                    row_acc = q1(
                        conn,
                        """
                        SELECT 1
                        FROM app.accounts a
                        JOIN app.domains d ON d.domain_id = a.domain_id
                        WHERE lower(a.login_name) = lower(%s) AND lower(d.name) = lower(%s)
                        LIMIT 1
                        """,
                        (login, domain),
                    )
                    in_file_accounts = (login.lower(), domain.lower()) in staged_accounts
                    if not row_acc and not in_file_accounts:
                        warnings.append({
                            "sheet": report_sheet,
                            "row": report_row,
                            "field": "Аккаунт",
                            "value": account_val,
                            "message": "Аккаунт не найден — строка будет пропущена",
                        })
                if not subscription_title:
                    warnings.append({
                        "sheet": report_sheet,
                        "row": report_row,
                        "field": "Подписка",
                        "value": subscription_title,
                        "message": "Название подписки не указано — строка будет пропущена",
                    })
                if not valid_until_text:
                    warnings.append({
                        "sheet": report_sheet,
                        "row": report_row,
                        "field": "Срок",
                        "value": valid_until_text,
                        "message": "Дата окончания обязательна — строка будет пропущена",
                    })
                elif not normalize_slot_date(valid_until_raw):
                    warnings.append({
                        "sheet": report_sheet,
                        "row": report_row,
                        "field": "Срок",
                        "value": valid_until_text,
                        "message": "Не удалось распознать дату окончания — строка будет пропущена",
                    })
            else:
                if not account_val or not login or not domain:
                    warnings.append({
                        "sheet": report_sheet,
                        "row": report_row,
                        "field": "Аккаунт",
                        "value": account_val,
                        "message": "Нужно значение в формате login@domain — строка будет пропущена",
                    })
            if progress_cb:
                progress_cb(idx - 1)
        return errors, warnings

    def normalize_slot_file_value(value: str) -> str:
        return str(value or "").strip().lower()

    def normalize_cell_text(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def normalize_slot_date(value):
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        raw = str(value).strip()
        if not raw:
            return None
        # Сначала обрабатываем ISO-формат строго как YYYY-MM-DD.
        iso = re.match(r"^\s*(\d{4})-(\d{1,2})-(\d{1,2})\s*$", raw)
        if iso:
            y = int(iso.group(1))
            mo = int(iso.group(2))
            d = int(iso.group(3))
            try:
                return date(y, mo, d).isoformat()
            except ValueError:
                return None
        # Затем обрабатываем только полный формат D/M/Y (без совпадений по подстроке).
        m = re.match(r"^\s*(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})\s*$", raw)
        if m:
            d = int(m.group(1))
            mo = int(m.group(2))
            y = int(m.group(3))
            if y < 100:
                y += 2000
            try:
                return date(y, mo, d).isoformat()
            except ValueError:
                return None
        m = re.match(r"^\s*([A-Za-zА-Яа-яёЁ]+)\s+(\d{2,4})\s*$", raw)
        if m:
            month_name = m.group(1).strip().lower()
            year = int(m.group(2))
            if year < 100:
                year += 2000
            month = ru_months.get(month_name)
            if month:
                try:
                    return date(year, month, 1).isoformat()
                except ValueError:
                    return None
        return None

    def normalize_slot_date_to_dt(value):
        normalized = normalize_slot_date(value)
        if not normalized:
            return None
        try:
            return datetime.fromisoformat(normalized).replace(tzinfo=timezone.utc)
        except ValueError:
            return None

    def resolve_source_id(conn, source_val: str):
        if not source_val:
            return None
        parts = [p for p in str(source_val).strip().split() if p]
        row = None
        if len(parts) >= 2:
            name_part = parts[0]
            code_part = parts[-1]
            row = q1(
                conn,
                """
                SELECT source_id
                FROM app.sources
                WHERE lower(name) = lower(%s) AND lower(code) = lower(%s)
                """,
                (name_part, code_part),
            )
            if not row:
                row = q1(
                    conn,
                    """
                    SELECT source_id
                    FROM app.sources
                    WHERE lower(name) = lower(%s) AND lower(code) = lower(%s)
                    """,
                    (code_part, name_part),
                )
        if not row:
            row = q1(
                conn,
                """
                SELECT source_id
                FROM app.sources
                WHERE lower(code) = lower(%s) OR lower(name) = lower(%s)
                """,
                (source_val, source_val),
            )
        return int(row[0]) if row else None

    def validate_slot_import_rows(conn, rows: list[dict], progress_cb=None):
        errors = []
        warnings = []
        total = 0
        # Запоминаем уже показанные предупреждения по аккаунту, чтобы не дублировать одно и то же на каждой строке.
        warned_accounts: set[tuple[str, str]] = set()
        warned_account_game_missing: set[tuple[str, str]] = set()
        warned_account_game_ambiguous: set[tuple[str, str]] = set()

        slot_rows = qall(conn, "SELECT code FROM app.slot_types")
        slot_types = {str(r[0]).strip().lower() for r in slot_rows}

        for idx, row in enumerate(rows, start=2):
            report_row = int(row.get("_sheet_row") or idx)
            if progress_cb:
                progress_cb(idx - 1)
            status_raw = row.get("status")
            status = "" if status_raw is None else str(status_raw).strip().lower()
            if not status or status == "свободен":
                continue

            total += 1
            account_val = normalize_cell_text(row.get("account"))
            slot_val = normalize_cell_text(row.get("slot"))
            game_title = normalize_cell_text(row.get("game"))
            customer = normalize_cell_text(row.get("customer"))
            source_val = normalize_cell_text(row.get("source"))
            date_val = row.get("date")
            login = None
            domain = None
            account_key = None

            if not account_val:
                errors.append({"row": report_row, "field": "Аккаунт", "value": account_val, "account": account_val, "message": "Аккаунт обязателен"})
            else:
                login, domain = split_account(account_val)
                if not login or not domain:
                    warnings.append({"row": report_row, "field": "Аккаунт", "value": account_val, "account": account_val, "message": "Нужно значение в формате login@domain"})
                else:
                    account_key = (login.lower(), domain.lower())
                    row_acc = q1(
                        conn,
                        """
                        SELECT 1
                        FROM app.accounts a
                        JOIN app.domains d ON d.domain_id = a.domain_id
                        WHERE lower(a.login_name) = lower(%s) AND lower(d.name) = lower(%s)
                        """,
                        (login, domain),
                    )
                    if not row_acc:
                        if not bool(row.get("_account_declared_in_file")) and account_key not in warned_accounts:
                            warnings.append({"row": report_row, "field": "Аккаунт", "value": account_val, "account": account_val, "message": "Аккаунт не найден"})
                            warned_accounts.add(account_key)

            if not slot_val:
                errors.append({"row": report_row, "field": "Слот", "value": slot_val, "account": account_val, "message": "Слот обязателен"})
            else:
                normalized_slot = normalize_slot_file_value(slot_val)
                slot_mapping = slot_file_to_type.get(normalized_slot)
                if not slot_mapping:
                    warnings.append({"row": report_row, "field": "Слот", "value": slot_val, "account": account_val, "message": "Неизвестный слот"})
                elif slot_mapping[0].lower() not in slot_types:
                    warnings.append({"row": report_row, "field": "Слот", "value": slot_val, "account": account_val, "message": "Слот не найден в БД"})

            if game_title:
                product_meta = resolve_slot_product_meta_by_title(conn, game_title)
                if product_meta is None:
                    if account_key:
                        if account_key not in warned_account_game_missing:
                            warnings.append({"row": report_row, "field": "Товар", "value": game_title, "account": account_val, "message": "Товар не найден"})
                            warned_account_game_missing.add(account_key)
                    else:
                        warnings.append({"row": report_row, "field": "Товар", "value": game_title, "account": account_val, "message": "Товар не найден"})
                elif str(product_meta.get("type_code") or "") == "subscription":
                    # Для подписки обязательно определяем и проверяем срок из листов ПЛЮС/EA PLAY.
                    if isinstance(row.get("_subscription_term_dates"), list) and len(row.get("_subscription_term_dates")) > 1:
                        warnings.append({
                            "row": report_row,
                            "field": "Срок подписки",
                            "value": ", ".join(str(x) for x in row.get("_subscription_term_dates")),
                            "account": account_val,
                            "message": "Найдено несколько дат срока подписки — выберите срок однозначно",
                        })
                    else:
                        term_date = normalize_slot_date(row.get("_subscription_valid_until"))
                        if not term_date:
                            warnings.append({
                                "row": report_row,
                                "field": "Срок подписки",
                                "value": str(row.get("_subscription_valid_until") or ""),
                                "account": account_val,
                                "message": "Дата срока подписки не найдена в листах ПЛЮС/EA PLAY",
                            })
                        elif account_key:
                            term_row = q1(
                                conn,
                                """
                                SELECT st.term_id
                                FROM app.subscription_terms st
                                JOIN app.accounts a ON a.account_id = st.account_id
                                JOIN app.domains d ON d.domain_id = a.domain_id
                                WHERE lower(a.login_name) = lower(%s)
                                  AND lower(d.name) = lower(%s)
                                  AND st.product_id = %s
                                  AND st.valid_until = %s::date
                                  AND st.is_archived IS NOT TRUE
                                LIMIT 1
                                """,
                                (login, domain, int(product_meta["product_id"]), term_date),
                            )
                            if not term_row:
                                warnings.append({
                                    "row": report_row,
                                    "field": "Срок подписки",
                                    "value": term_date,
                                    "account": account_val,
                                    "message": "Срок подписки не найден в БД для аккаунта и товара",
                                })
            else:
                candidates = row.get("_game_candidates") or []
                if isinstance(candidates, list) and len(candidates) > 1:
                    if account_key:
                        if account_key not in warned_account_game_ambiguous:
                            warnings.append({
                                "row": report_row,
                                "field": "Товар",
                                "value": ", ".join(str(x) for x in candidates),
                                "account": account_val,
                                "message": "Найдено несколько товаров/подписок для аккаунта — выберите товар однозначно",
                            })
                            warned_account_game_ambiguous.add(account_key)
                    else:
                        warnings.append({
                            "row": report_row,
                            "field": "Товар",
                            "value": ", ".join(str(x) for x in candidates),
                            "account": account_val,
                            "message": "Найдено несколько товаров/подписок для аккаунта — выберите товар однозначно",
                        })
                else:
                    if account_key:
                        if account_key not in warned_account_game_missing:
                            warnings.append({
                                "row": report_row,
                                "field": "Товар",
                                "value": game_title,
                                "account": account_val,
                                "message": "Товар не найден (нет совпадения в листах Аккаунты/ПЛЮС/EA PLAY)",
                            })
                            warned_account_game_missing.add(account_key)
                    else:
                        warnings.append({
                            "row": report_row,
                            "field": "Товар",
                            "value": game_title,
                            "account": account_val,
                            "message": "Товар не найден (нет совпадения в листах Аккаунты/ПЛЮС/EA PLAY)",
                        })

            if not customer:
                warnings.append({"row": report_row, "field": "Пользователь", "value": customer, "account": account_val, "message": "Пользователь не указан"})

            if source_val:
                parts = [p for p in str(source_val).strip().split() if p]
                row_source = None
                if len(parts) >= 2:
                    name_part = parts[0]
                    code_part = parts[-1]
                    row_source = q1(
                        conn,
                        """
                        SELECT 1
                        FROM app.sources
                        WHERE lower(name) = lower(%s) AND lower(code) = lower(%s)
                        """,
                        (name_part, code_part),
                    )
                    if not row_source:
                        row_source = q1(
                            conn,
                            """
                            SELECT 1
                            FROM app.sources
                            WHERE lower(name) = lower(%s) AND lower(code) = lower(%s)
                            """,
                            (code_part, name_part),
                        )
                if not row_source:
                    row_source = q1(
                        conn,
                        """
                        SELECT 1
                        FROM app.sources
                        WHERE lower(code) = lower(%s) OR lower(name) = lower(%s)
                        """,
                        (source_val, source_val),
                    )
                if not row_source:
                    warnings.append({"row": report_row, "field": "Источник", "value": source_val, "account": account_val, "message": "Источник не найден"})

            if date_val not in (None, ""):
                normalized_date = normalize_slot_date(date_val)
                if not normalized_date:
                    original_date = normalize_cell_text(date_val)
                    warnings.append({"row": report_row, "field": "Дата", "value": original_date, "account": account_val, "message": "Не удалось распознать дату"})

            if progress_cb:
                progress_cb(idx - 1)

        return errors, warnings, total

    return ImportParsers(
        GAME_IMPORT_HEADERS=game_import_headers,
        parse_import_platforms=parse_import_platforms,
        validate_game_import_rows=validate_game_import_rows,
        read_games_from_excel=read_games_from_excel,
        read_accounts_from_excel=read_accounts_from_excel,
        read_slots_from_excel=read_slots_from_excel,
        clean_slots_excel=clean_slots_excel,
        split_account=split_account,
        validate_account_import_rows=validate_account_import_rows,
        SLOT_FILE_TO_TYPE=slot_file_to_type,
        normalize_slot_file_value=normalize_slot_file_value,
        normalize_cell_text=normalize_cell_text,
        normalize_slot_date=normalize_slot_date,
        normalize_slot_date_to_dt=normalize_slot_date_to_dt,
        resolve_source_id=resolve_source_id,
        validate_slot_import_rows=validate_slot_import_rows,
        detect_logo_mime_from_bytes=detect_logo_mime_from_bytes,
        fetch_logo_from_url=fetch_logo_from_url,
    )
