from datetime import datetime
from io import BytesIO
from typing import Any
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PRICE_TYPES = {"VOUCHER", "TOP_UP_FIXED"}


def collect_price_targets(services: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Собирает активные номиналы, для которых InterHub поддерживает calculate."""
    targets: list[dict[str, Any]] = []
    for service in services:
        service_type = str(service.get("type") or "").upper()
        if service_type not in PRICE_TYPES:
            continue
        nominal_field = next(
            (field for field in service.get("fields") or [] if str(field.get("name") or "") == "nominal"),
            None,
        )
        if not nominal_field:
            continue
        for nominal in nominal_field.get("value_list") or []:
            if not isinstance(nominal, dict) or nominal.get("active") is False:
                continue
            try:
                nominal_id = int(nominal.get("id"))
                service_id = int(service.get("service_id"))
            except (TypeError, ValueError):
                continue
            if service_id <= 0 or nominal_id <= 0:
                continue
            targets.append(
                {
                    "service_id": service_id,
                    "service_title": str(service.get("title") or ""),
                    "category": str(service.get("category") or ""),
                    "service_type": service_type,
                    "nominal_id": nominal_id,
                    "nominal_title": str(nominal.get("title") or nominal_id),
                }
            )
    return targets


def build_interhub_prices_xlsx(prices: list[dict[str, Any]], errors: list[dict[str, Any]]) -> bytes:
    """Создаёт один файл с закупочными ценами и ошибками последнего запуска."""
    workbook = Workbook()
    prices_sheet = workbook.active
    prices_sheet.title = "Закупочные цены"
    prices_sheet.append(["ID услуги", "Услуга", "Категория", "Тип", "ID номинала", "Номинал", "Закупочная цена, ₽", "Рассчитано", "Полный ответ calculate (JSON)"])
    for row in prices:
        prices_sheet.append([
            row.get("service_id"), row.get("service_title"), row.get("category"), row.get("service_type"),
            row.get("nominal_id"), row.get("nominal_title"), row.get("fixed_amount"), format_datetime(row.get("calculated_at")),
            format_provider_response(row.get("provider_response")),
        ])
    errors_sheet = workbook.create_sheet("Ошибки calculate")
    errors_sheet.append(["ID услуги", "Услуга", "Тип", "ID номинала", "Номинал", "Статус InterHub", "Сообщение", "Время", "Полный ответ calculate (JSON)"])
    for row in errors:
        errors_sheet.append([
            row.get("service_id"), row.get("service_title"), row.get("service_type"), row.get("nominal_id"),
            row.get("nominal_title"), row.get("provider_status"), row.get("provider_message"), format_datetime(row.get("calculated_at")),
            format_provider_response(row.get("provider_response")),
        ])
    for sheet in (prices_sheet, errors_sheet):
        style_sheet(sheet)
    prices_sheet.column_dimensions["G"].width = 22
    prices_sheet.column_dimensions["I"].width = 52
    errors_sheet.column_dimensions["I"].width = 52
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def format_datetime(value: Any) -> str:
    """Приводит дату из базы к читаемому виду в Excel."""
    if isinstance(value, datetime):
        return value.astimezone().strftime("%d.%m.%Y %H:%M:%S")
    return str(value or "")


def format_provider_response(value: Any) -> str:
    """Сохраняет ответ поставщика в читаемом JSON без удаления неизвестных полей."""
    if isinstance(value, str):
        return value
    try:
        return json.dumps(value or {}, ensure_ascii=False, sort_keys=True)
    except (TypeError, ValueError):
        return str(value or "")


def style_sheet(sheet) -> None:
    """Оформляет лист, чтобы отчёт было удобно фильтровать и читать."""
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        longest = max((len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1)), default=10)
        sheet.column_dimensions[get_column_letter(column)].width = min(max(longest + 2, 12), 42)
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 7).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row, sheet.max_column).alignment = Alignment(vertical="top", wrap_text=True)
