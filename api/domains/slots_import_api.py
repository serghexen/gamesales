from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
import threading
import uuid

from fastapi import Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


SLOT_EXPORT_LABELS = {
    "activate_ps4": "П4",
    "activate_ps5": "П5",
    "play_ps4": "PS4",
    "play_ps5": "PS5",
}


# Приводит дату к timezone-aware значению, чтобы корректно сравнивать интервалы занятости.
def _as_aware_datetime(value):
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


# Возвращает подпись конкретного физического слота с номером дорожки для емкости больше одного.
def _slot_export_label(slot_type_code, lane_idx, capacity):
    base = SLOT_EXPORT_LABELS.get(str(slot_type_code or "").strip().lower(), str(slot_type_code or ""))
    if int(capacity or 0) > 1:
        return f"{base}({lane_idx + 1})"
    return base


# Оставляет только календарную дату, потому что Excel не поддерживает timezone в datetime.
def _excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    return value


# Восстанавливает историю физических слотов и добавляет строки для свободных мест на текущий момент.
def _build_slots_export_rows(db_rows):
    grouped = defaultdict(list)
    for row in db_rows:
        account_full, slot_type_code, capacity = row[:3]
        grouped[(str(account_full or ""), str(slot_type_code or ""), max(int(capacity or 0), 1))].append(row)

    export_rows = []
    far_future = datetime.max.replace(tzinfo=timezone.utc)
    now_utc = datetime.now(timezone.utc)
    for (account_full, slot_type_code, capacity), rows in sorted(grouped.items(), key=lambda item: (item[0][0].lower(), item[0][1])):
        assignments = [row for row in rows if row[3] is not None]
        assignments.sort(key=lambda row: (_as_aware_datetime(row[7]) or far_future, int(row[3] or 0)))
        lane_intervals = [[] for _ in range(capacity)]

        for row in assignments:
            assignment_id, product_title, customer_nickname, purchase_at, assigned_at, released_at = row[3:9]
            assigned_dt = _as_aware_datetime(assigned_at) or far_future
            released_dt = _as_aware_datetime(released_at) or far_future

            # Убираем завершенные интервалы, чтобы повторное использование освобожденного слота не считалось дублем.
            for lane in lane_intervals:
                lane[:] = [end_at for end_at in lane if end_at > assigned_dt]

            lane_idx = next((idx for idx, lane in enumerate(lane_intervals) if not lane), None)
            is_duplicate = lane_idx is None
            if lane_idx is None:
                lane_idx = min(range(capacity), key=lambda idx: min(lane_intervals[idx]))
            lane_intervals[lane_idx].append(released_dt)

            export_rows.append({
                "account": account_full,
                "duplicate": "Да" if is_duplicate else "",
                "product": product_title or "",
                "status": "Занят",
                "customer": customer_nickname or "",
                "purchase_at": purchase_at or assigned_at,
                "slot_type": _slot_export_label(slot_type_code, lane_idx, capacity),
                "is_free": False,
                "assignment_id": int(assignment_id),
            })

        # После проигрывания истории оставляем только интервалы, которые заняты сейчас.
        for lane_idx, lane in enumerate(lane_intervals):
            active_intervals = [end_at for end_at in lane if end_at > now_utc]
            if active_intervals:
                continue
            export_rows.append({
                "account": account_full,
                "duplicate": "",
                "product": "",
                "status": "Свободен",
                "customer": "",
                "purchase_at": None,
                "slot_type": _slot_export_label(slot_type_code, lane_idx, capacity),
                "is_free": True,
                "assignment_id": None,
            })

    return export_rows


# Собирает оформленный XLSX-файл с историей занятости и текущими свободными слотами.
def _build_slots_export_xlsx(db_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Слоты"
    headers = ["Аккаунт", "Дубль", "Товар", "Статус", "Пользователь", "Дата покупки (сделки)", "Тип слота"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    free_fill = PatternFill("solid", fgColor="C6EFCE")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in _build_slots_export_rows(db_rows):
        ws.append([
            item["account"],
            item["duplicate"],
            item["product"],
            item["status"],
            item["customer"],
            _excel_date(item["purchase_at"]),
            item["slot_type"],
        ])
        row_idx = ws.max_row
        if item["is_free"]:
            for cell in ws[row_idx]:
                cell.fill = free_fill
        ws.cell(row=row_idx, column=6).number_format = "DD.MM.YYYY"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(ws.max_row, 1)}"
    widths = {"A": 32, "B": 12, "C": 36, "D": 14, "E": 24, "F": 24, "G": 16}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def mount_slots_import_routes(
    app,
    *,
    DB_DSN,
    psycopg,
    clean_slots_excel,
    set_import_progress,
    get_import_progress,
    run_slots_validate_job,
    run_slots_import_job,
    build_import_report_xlsx,
    require_role,
    UserOut,
    ImportReportIn,
):
    # Подключает маршруты импорта и отдельную выгрузку полной истории слотов.
    @app.get("/accounts/slots/export")
    def accounts_slots_export(user: UserOut = Depends(require_role("admin", "owner"))):
        # Выбираем всю историю назначений и пустые типы слотов, чтобы добавить актуальные свободные строки.
        with psycopg.connect(DB_DSN) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    WITH account_products_linked AS (
                      SELECT aa.account_id, aa.product_id
                      FROM app.account_assets aa
                      WHERE aa.asset_type_code IN ('game', 'subscription')
                      UNION
                      SELECT terms.account_id, terms.product_id
                      FROM app.subscription_terms terms
                      WHERE terms.is_archived IS NOT TRUE
                    ),
                    account_platforms AS (
                      SELECT
                        links.account_id,
                        bool_or(lower(pl.code) = 'ps4') AS has_ps4
                      FROM account_products_linked links
                      LEFT JOIN app.product_platforms pp ON pp.product_id = links.product_id
                      LEFT JOIN app.platforms pl ON pl.platform_id = pp.platform_id
                      GROUP BY links.account_id
                    ),
                    export_slots AS (
                      SELECT
                        a.account_id,
                        a.login_name || '@' || d.name AS account_full,
                        st.code AS slot_type_code,
                        st.capacity
                      FROM app.accounts a
                      JOIN app.domains d ON d.domain_id = a.domain_id
                      CROSS JOIN app.slot_types st
                      LEFT JOIN account_platforms ap ON ap.account_id = a.account_id
                      WHERE COALESCE(a.is_archived, false) IS NOT TRUE
                        AND (
                          st.platform_code = 'ps5'
                          OR COALESCE(ap.has_ps4, false)
                          OR EXISTS (
                            SELECT 1
                            FROM app.account_slot_assignments asa_history
                            WHERE asa_history.account_id = a.account_id
                              AND asa_history.slot_type_code = st.code
                          )
                        )
                    )
                    SELECT
                      es.account_full,
                      es.slot_type_code,
                      es.capacity,
                      asa.assignment_id,
                      p.title AS product_title,
                      c.nickname AS customer_nickname,
                      COALESCE(di.purchase_at, deals.completed_at, deals.created_at, asa.assigned_at) AS purchase_at,
                      asa.assigned_at,
                      asa.released_at
                    FROM export_slots es
                    LEFT JOIN app.account_slot_assignments asa
                      ON asa.account_id = es.account_id
                     AND asa.slot_type_code = es.slot_type_code
                    LEFT JOIN app.deal_items di ON di.deal_item_id = asa.deal_item_id
                    LEFT JOIN app.deals deals ON deals.deal_id = COALESCE(asa.deal_id, di.deal_id)
                    LEFT JOIN app.products p ON p.product_id = COALESCE(asa.product_id, di.product_id)
                    LEFT JOIN app.customers c ON c.customer_id = COALESCE(asa.customer_id, deals.customer_id)
                    ORDER BY lower(es.account_full), es.slot_type_code, asa.assigned_at, asa.assignment_id
                    """
                )
                rows = cur.fetchall()

        content = _build_slots_export_xlsx(rows)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=slots_history.xlsx"},
        )

    @app.post("/accounts/slots/clean")
    def accounts_slots_clean(file: UploadFile = File(...), user: UserOut = Depends(require_role("admin", "owner"))):
        if not file or not file.filename:
            raise HTTPException(400, "file is required")
        if not file.filename.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(400, "Only .xlsx/.xls are supported")
        content = file.file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(400, "File too large. Max 10MB")
        cleaned = clean_slots_excel(content)
        base = Path(file.filename or "slots_import").stem
        filename = f"{base}_cleaned.xlsx"
        return Response(
            cleaned,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    @app.post("/accounts/slots/validate")
    def accounts_slots_validate(
        file: UploadFile = File(...),
        limit: int | None = Form(None),
        user: UserOut = Depends(require_role("admin", "owner")),
    ):
        if not file or not file.filename:
            raise HTTPException(400, "file is required")
        if not file.filename.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(400, "Only .xlsx/.xls are supported")
        content = file.file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(400, "File too large. Max 10MB")
        job_id = uuid.uuid4().hex
        set_import_progress(job_id, user.username, {"phase": "queued", "current": 0, "total": 0, "done": False})
        thread = threading.Thread(target=run_slots_validate_job, args=(job_id, content, user.username, limit), daemon=True)
        thread.start()
        return {"ok": True, "job_id": job_id}

    @app.get("/accounts/slots/validate/status")
    def accounts_slots_validate_status(job_id: str, user: UserOut = Depends(require_role("admin", "owner"))):
        status = get_import_progress(job_id)
        if not status:
            return {"phase": "idle", "current": 0, "total": 0, "done": True}
        if status.get("owner") and status.get("owner") != user.username:
            raise HTTPException(403, "job not found")
        return {
            "phase": status.get("phase", "idle"),
            "current": int(status.get("current") or 0),
            "total": int(status.get("total") or 0),
            "done": bool(status.get("done")),
            "result": status.get("result"),
        }

    @app.post("/accounts/slots/validate/cancel")
    def accounts_slots_validate_cancel(job_id: str, user: UserOut = Depends(require_role("admin", "owner"))):
        status = get_import_progress(job_id)
        if not status:
            return {"ok": True}
        if status.get("owner") and status.get("owner") != user.username:
            raise HTTPException(403, "job not found")
        set_import_progress(
            job_id,
            user.username,
            {
                "phase": "cancelled",
                "current": int(status.get("current") or 0),
                "total": int(status.get("total") or 0),
                "done": True,
                "cancelled": True,
                "result": {"ok": False, "cancelled": True},
            },
        )
        return {"ok": True}

    @app.post("/accounts/slots/report")
    def accounts_slots_report(payload: ImportReportIn, user: UserOut = Depends(require_role("admin", "owner"))):
        content = build_import_report_xlsx(
            errors=payload.errors or [],
            warnings=payload.warnings or [],
        )
        return Response(
            content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=slots_import_report.xlsx"},
        )

    @app.post("/accounts/slots/import")
    def accounts_slots_import(file: UploadFile = File(...), user: UserOut = Depends(require_role("admin", "owner"))):
        if not file or not file.filename:
            raise HTTPException(400, "file is required")
        if not file.filename.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(400, "Only .xlsx/.xls are supported")
        content = file.file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(400, "File too large. Max 10MB")
        job_id = uuid.uuid4().hex
        set_import_progress(job_id, user.username, {"phase": "queued", "current": 0, "total": 0, "done": False})
        thread = threading.Thread(target=run_slots_import_job, args=(job_id, content, user.username), daemon=True)
        thread.start()
        return {"ok": True, "job_id": job_id}

    @app.get("/accounts/slots/import/status")
    def accounts_slots_import_status(job_id: str, user: UserOut = Depends(require_role("admin", "owner"))):
        status = get_import_progress(job_id)
        if not status:
            return {"phase": "idle", "current": 0, "total": 0, "done": True}
        if status.get("owner") and status.get("owner") != user.username:
            raise HTTPException(403, "job not found")
        return {
            "phase": status.get("phase", "idle"),
            "current": int(status.get("current") or 0),
            "total": int(status.get("total") or 0),
            "done": bool(status.get("done")),
            "result": status.get("result"),
        }

    @app.post("/accounts/slots/import/cancel")
    def accounts_slots_import_cancel(job_id: str, user: UserOut = Depends(require_role("admin", "owner"))):
        status = get_import_progress(job_id)
        if not status:
            return {"ok": True}
        if status.get("owner") and status.get("owner") != user.username:
            raise HTTPException(403, "job not found")
        set_import_progress(
            job_id,
            user.username,
            {
                "phase": "cancelled",
                "current": int(status.get("current") or 0),
                "total": int(status.get("total") or 0),
                "done": True,
                "cancelled": True,
                "result": {"ok": False, "cancelled": True},
            },
        )
        return {"ok": True}
