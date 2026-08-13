"""Проверяет доступность номиналов InterHub без оплаты и формирует Excel-отчёт."""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def resolve_runtime_paths(script_path: Path, current_dir: Path) -> tuple[Path, Path]:
    # Находит модули API как в репозитории, так и при временном запуске файла из /tmp контейнера.
    resolved_script = script_path.resolve()
    for parent in resolved_script.parents:
        if (parent / "api" / "domains").is_dir():
            return parent, parent / "api"
    runtime_dir = current_dir.resolve()
    return runtime_dir, runtime_dir


ROOT_DIR, API_DIR = resolve_runtime_paths(Path(__file__), Path.cwd())
if str(API_DIR) not in sys.path:
    sys.path.insert(0, str(API_DIR))

from domains.interhub_price_cache import collect_price_targets  # noqa: E402
from domains.interhub_service import build_interhub_service  # noqa: E402


REPORT_HEADERS = [
    "ID услуги",
    "Услуга",
    "Категория",
    "Тип",
    "ID номинала",
    "Номинал",
    "Закупочная цена, ₽",
    "Сумма номинала",
    "Calculate успешен",
    "Статус calculate",
    "Сообщение calculate",
    "Check выполнен",
    "Доступность",
    "Статус check",
    "Сообщение check",
    "Transaction ID check",
]


def parse_bool(value: Any, default: bool = False) -> bool:
    # Приводит строковую настройку окружения к булеву значению без неявных вариантов.
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def error_result(exc: Exception) -> dict[str, Any]:
    # Превращает сетевую ошибку в обычную строку отчёта, чтобы обход продолжился.
    status = int(getattr(exc, "status_code", -1) or -1)
    message = str(getattr(exc, "detail", exc) or exc)
    return {"success": False, "status": status, "message": message, "raw": {}}


def json_text(value: Any) -> str:
    # Сохраняет полный ответ поставщика для последующей ручной сверки в Excel.
    try:
        return json.dumps(value or {}, ensure_ascii=False, sort_keys=True)
    except (TypeError, ValueError):
        return str(value or "")


def audit_targets(
    service,
    targets: list[dict[str, Any]],
    *,
    delay_ms: int,
    run_id: str,
    progress=None,
) -> list[dict[str, Any]]:
    # Последовательно вызывает calculate и check, но никогда не вызывает pay.
    rows: list[dict[str, Any]] = []
    for index, target in enumerate(targets, start=1):
        if index > 1 and delay_ms > 0:
            time.sleep(delay_ms / 1000)

        base_payload = {
            "service_id": target["service_id"],
            "account": "",
            "params": {"nominal": target["nominal_id"]},
        }
        try:
            calculated = service.calculate({
                **base_payload,
                "agent_transaction_id": f"gamesales-stock-calc-{run_id}-{index}",
            })
        except Exception as exc:  # noqa: BLE001 - каждая внешняя ошибка должна попасть в отчёт
            calculated = error_result(exc)

        fixed_amount = float(calculated.get("fixed_amount") or 0)
        check_attempted = bool(calculated.get("success")) and fixed_amount > 0
        checked: dict[str, Any] = {}
        if check_attempted:
            if delay_ms > 0:
                time.sleep(delay_ms / 1000)
            try:
                checked = service.check({
                    **base_payload,
                    "agent_transaction_id": f"gamesales-stock-check-{run_id}-{index}",
                    "amount": fixed_amount,
                })
            except Exception as exc:  # noqa: BLE001 - сбой одного номинала не останавливает аудит
                checked = error_result(exc)

        if checked.get("success"):
            availability = "Доступно"
        elif check_attempted:
            availability = "Недоступно / ошибка"
        else:
            availability = "Не проверено"

        row = {
            **target,
            "fixed_amount": fixed_amount,
            "amount_in_currency": float(calculated.get("amount_in_currency") or 0),
            "calculate_success": bool(calculated.get("success")),
            "calculate_status": int(calculated.get("status") or 0),
            "calculate_message": str(calculated.get("message") or ""),
            "check_attempted": check_attempted,
            "availability": availability,
            "check_status": int(checked.get("status") or 0) if check_attempted else None,
            "check_message": str(checked.get("message") or "") if check_attempted else "Calculate не вернул цену",
            "check_transaction_id": str(checked.get("transaction_id") or ""),
            "calculate_response": calculated.get("raw") or {},
            "check_response": checked.get("raw") or {},
        }
        rows.append(row)
        if progress:
            progress(index, len(targets), row)
    return rows


def build_report(rows: list[dict[str, Any]], *, generated_at: datetime) -> Workbook:
    # Создаёт короткую сводку и подробный лист, пригодный для фильтрации по ошибкам.
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Сводка"
    # Excel хранит локальное время без timezone, поэтому убираем tzinfo после перевода в локальную зону.
    excel_generated_at = generated_at.astimezone().replace(tzinfo=None) if generated_at.tzinfo else generated_at
    available = sum(row["availability"] == "Доступно" for row in rows)
    unavailable = sum(row["availability"] == "Недоступно / ошибка" for row in rows)
    skipped = sum(row["availability"] == "Не проверено" for row in rows)
    summary_rows = [
        ("Проверка остатков InterHub", "calculate → check, без pay"),
        ("Сформировано", excel_generated_at),
        ("Всего номиналов", len(rows)),
        ("Доступно", available),
        ("Недоступно / ошибка", unavailable),
        ("Не проверено", skipped),
    ]
    for item in summary_rows:
        summary.append(item)
    summary["A1"].font = Font(bold=True, color="FFFFFF")
    summary["B1"].font = Font(bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary["B1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary["B2"].number_format = "yyyy-mm-dd hh:mm:ss"
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 34

    report = workbook.create_sheet("Проверка остатков")
    report.append(REPORT_HEADERS)
    for row in rows:
        report.append([
            row.get("service_id"), row.get("service_title"), row.get("category"), row.get("service_type"),
            row.get("nominal_id"), row.get("nominal_title"), row.get("fixed_amount"), row.get("amount_in_currency"),
            row.get("calculate_success"), row.get("calculate_status"), row.get("calculate_message"),
            row.get("check_attempted"), row.get("availability"), row.get("check_status"), row.get("check_message"),
            row.get("check_transaction_id"),
        ])
    style_report_sheet(report)

    responses = workbook.create_sheet("Ответы API")
    responses.append(["ID услуги", "Услуга", "ID номинала", "Номинал", "Доступность", "Ответ calculate (JSON)", "Ответ check (JSON)"])
    for row in rows:
        responses.append([
            row.get("service_id"), row.get("service_title"), row.get("nominal_id"), row.get("nominal_title"),
            row.get("availability"), json_text(row.get("calculate_response")), json_text(row.get("check_response")),
        ])
    style_response_sheet(responses)
    return workbook


def style_report_sheet(sheet) -> None:
    # Оформляет таблицу и подсвечивает итог проверки, чтобы ошибки были заметны сразу.
    header_fill = PatternFill("solid", fgColor="1F4E78")
    status_fills = {
        "Доступно": PatternFill("solid", fgColor="C6EFCE"),
        "Недоступно / ошибка": PatternFill("solid", fgColor="FFC7CE"),
        "Не проверено": PatternFill("solid", fgColor="FFEB9C"),
    }
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 7).number_format = "#,##0.00"
        sheet.cell(row_index, 8).number_format = "#,##0.00"
        status_cell = sheet.cell(row_index, 13)
        status_cell.fill = status_fills.get(str(status_cell.value or ""), PatternFill())
        for column in (11, 15):
            sheet.cell(row_index, column).alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, sheet.max_column + 1):
        longest = max((len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1)), default=10)
        sheet.column_dimensions[get_column_letter(column)].width = min(max(longest + 2, 12), 48)


def style_response_sheet(sheet) -> None:
    # Выносит длинные JSON-ответы на отдельный лист, чтобы основной отчёт оставался компактным.
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row_index in range(2, sheet.max_row + 1):
        for column in (6, 7):
            sheet.cell(row_index, column).alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in {"A": 14, "B": 28, "C": 16, "D": 24, "E": 24, "F": 54, "G": 54}.items():
        sheet.column_dimensions[column].width = width


def build_service_from_env():
    # Собирает тот же клиент InterHub, который использует API, не печатая токен и адрес proxy.
    return build_interhub_service(
        HTTPException=HTTPException,
        interhub_api_url=os.getenv("INTERHUB_API_URL", ""),
        interhub_token=os.getenv("INTERHUB_TOKEN", ""),
        timeout_sec=int(os.getenv("INTERHUB_TIMEOUT_SEC", "20") or "20"),
        ssl_verify=parse_bool(os.getenv("INTERHUB_SSL_VERIFY"), True),
        ca_cert_path=os.getenv("INTERHUB_CA_CERT_PATH", ""),
        proxy_url=os.getenv("INTERHUB_PROXY_URL", ""),
        calculate_path=os.getenv("INTERHUB_CALCULATE_PATH", "/api/agent/payment/check/calculate"),
        check_path=os.getenv("INTERHUB_CHECK_PATH", "/api/agent/payment/check"),
        pay_path=os.getenv("INTERHUB_PAY_PATH", "/api/agent/payment/pay"),
        check_status_path=os.getenv("INTERHUB_CHECK_STATUS_PATH", "/api/agent/payment/check_status"),
        deposit_path=os.getenv("INTERHUB_DEPOSIT_PATH", "/api/agent/deposit"),
    )


def parse_args() -> argparse.Namespace:
    # Читает только безопасные параметры запуска и оставляет секреты в окружении.
    parser = argparse.ArgumentParser(description="Проверить номиналы InterHub через calculate + check без pay")
    parser.add_argument("--output", type=Path, help="Путь к итоговому .xlsx")
    parser.add_argument("--delay-ms", type=int, default=None, help="Пауза между запросами к InterHub")
    parser.add_argument("--limit", type=int, default=0, help="Проверить только первые N номиналов")
    parser.add_argument("--service-id", action="append", type=int, default=[], help="Проверять только указанный ID услуги")
    return parser.parse_args()


def main() -> int:
    # Выполняет аудит из CLI и сохраняет один отчёт без каких-либо вызовов оплаты.
    load_dotenv(ROOT_DIR / ".env.dev", override=False)
    args = parse_args()
    delay_ms = args.delay_ms
    if delay_ms is None:
        delay_ms = int(os.getenv("INTERHUB_PRICE_CALCULATE_DELAY_MS", "700") or "700")
    if delay_ms < 0 or args.limit < 0:
        print("--delay-ms и --limit не могут быть отрицательными", file=sys.stderr)
        return 2

    service = build_service_from_env()
    targets = collect_price_targets(service.get_services())
    if args.service_id:
        allowed_services = set(args.service_id)
        targets = [target for target in targets if target["service_id"] in allowed_services]
    if args.limit:
        targets = targets[:args.limit]
    if not targets:
        print("В каталоге нет подходящих активных номиналов", file=sys.stderr)
        return 1

    started_at = datetime.now().astimezone()
    run_id = f"{started_at:%Y%m%d%H%M%S}-{uuid.uuid4().hex[:6]}"

    def print_progress(index: int, total: int, row: dict[str, Any]) -> None:
        # Показывает ход выполнения без полных ответов и секретов поставщика.
        print(f"[{index}/{total}] {row['service_title']} · {row['nominal_title']}: {row['availability']}")

    rows = audit_targets(service, targets, delay_ms=delay_ms, run_id=run_id, progress=print_progress)
    workbook = build_report(rows, generated_at=started_at)
    output = args.output or Path.cwd() / f"interhub-stock-audit-{started_at:%Y%m%d-%H%M%S}.xlsx"
    output = output.expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    available = sum(row["availability"] == "Доступно" for row in rows)
    print(f"Готово: {output}")
    print(f"Доступно: {available}; проблемы: {len(rows) - available}; pay не вызывался")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
