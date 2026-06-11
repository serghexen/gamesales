import unittest
from io import BytesIO
from datetime import date, datetime, timezone
from unittest.mock import patch

try:
    from fastapi.testclient import TestClient
except Exception:  # pragma: no cover
    TestClient = None

from openpyxl import Workbook, load_workbook

import app as app_module


class _ScriptedCursor:
    def __init__(self, script):
        self._script = script
        self._current = None
        self.rowcount = 0

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        return False

    def execute(self, sql, params=None):
        if not self._script:
            raise AssertionError(f"Unexpected SQL without scripted response: {sql}")
        self._current = self._script.pop(0)
        self.rowcount = int(self._current.get("rowcount", 0))

    def fetchone(self):
        if not self._current:
            return None
        return self._current.get("one")

    def fetchall(self):
        if not self._current:
            return []
        return self._current.get("all", [])


class _ScriptedConnCtx:
    def __init__(self, script):
        self._script = list(script)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        return False

    def commit(self):
        return None

    def cursor(self):
        return _ScriptedCursor(self._script)


@unittest.skipIf(TestClient is None, "fastapi.testclient requires httpx")
class AccountsImportFormatTests(unittest.TestCase):
    def _client(self):
        return TestClient(app_module.app)

    def _auth_headers(self, role="admin", username="admin", user_id=1):
        token = app_module.create_access_token(user_id, username, role)
        return {"Authorization": f"Bearer {token}"}

    # Собирает общий xlsx с произвольными листами.
    def _build_workbook(self):
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Логин", "Пароль", "Резерв 1"])
        ws_summary.append(["summary-login", "p", "r"])

        ws_mail_0 = wb.create_sheet("Почты")
        ws_mail_0.append(["Логин", "Пароль", "Резерв 1", "Резерв 2"])
        ws_mail_0.append(["user1@gmail.com", "pass-1", "res-1", "res-2"])

        ws_mail_1 = wb.create_sheet("Почты1")
        ws_mail_1.append(["Логин", "Пароль", "Резерв 1"])
        ws_mail_1.append(["user2@yahoo.com", "pass-2", "res-3"])

        ws_other = wb.create_sheet("Другое")
        ws_other.append(["Логин", "Пароль"])
        ws_other.append(["other@domain.com", "pass-x"])

        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    # Проверяет, что validate читает только листы Почты* и не берет посторонние листы.
    def test_accounts_import_validate_reads_only_pochty_sheets(self):
        content = self._build_workbook()
        with (
            patch.object(app_module, "ensure_analytics_schema", return_value=None),
            patch.object(app_module.psycopg, "connect", return_value=_ScriptedConnCtx([])),
            patch.object(app_module, "JWT_SECRET", "test-secret"),
            patch.object(app_module, "JWT_ALG", "HS256"),
        ):
            with self._client() as client:
                res = client.post(
                    "/accounts/import/validate",
                    headers=self._auth_headers(role="admin"),
                    files={"file": ("accounts.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            self.assertEqual(res.status_code, 200)
            body = res.json()
            self.assertEqual(body["ok"], True)
            self.assertEqual(body["total"], 2)
            self.assertEqual(len(body.get("warnings", [])), 0)

    # Проверяет предупреждение для логина без домена.
    def test_accounts_import_validate_warns_for_login_without_domain(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Почты"
        ws.append(["Логин", "Пароль", "Резерв 1"])
        ws.append(["broken-login", "pass-1", "res-1"])
        out = BytesIO()
        wb.save(out)
        content = out.getvalue()

        with (
            patch.object(app_module, "ensure_analytics_schema", return_value=None),
            patch.object(app_module.psycopg, "connect", return_value=_ScriptedConnCtx([])),
            patch.object(app_module, "JWT_SECRET", "test-secret"),
            patch.object(app_module, "JWT_ALG", "HS256"),
        ):
            with self._client() as client:
                res = client.post(
                    "/accounts/import/validate",
                    headers=self._auth_headers(role="admin"),
                    files={"file": ("accounts.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            self.assertEqual(res.status_code, 200)
            body = res.json()
            self.assertEqual(body["total"], 1)
            self.assertTrue(any("login@domain" in str(w.get("message") or "") for w in body.get("warnings", [])))

    # Проверяет, что парсер складывает резервы в reserve_values.
    def test_read_accounts_from_excel_collects_reserves(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Почты"
        ws.append(["Логин", "Пароль", "Резерв 1", "Резерв 2", "Резерв 10"])
        ws.append(["user@gmail.com", "pass-1", "one", "two", "ten"])
        out = BytesIO()
        wb.save(out)

        rows = app_module.read_accounts_from_excel(out.getvalue())
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["account"], "user@gmail.com")
        self.assertEqual(rows[0]["password"], "pass-1")
        self.assertEqual(rows[0]["reserve_values"], {"reserve1": "one", "reserve2": "two", "reserve10": "ten"})

    # Проверяет лимит строк на каждом листе для быстрого локального прогона.
    def test_read_accounts_from_excel_respects_limit_per_sheet(self):
        wb = Workbook()
        ws_0 = wb.active
        ws_0.title = "Почты"
        ws_0.append(["Логин", "Пароль"])
        ws_0.append(["u1@gmail.com", "p1"])
        ws_0.append(["u2@gmail.com", "p2"])
        ws_1 = wb.create_sheet("Почты1")
        ws_1.append(["Логин", "Пароль"])
        ws_1.append(["u3@gmail.com", "p3"])
        ws_1.append(["u4@gmail.com", "p4"])
        out = BytesIO()
        wb.save(out)

        rows = app_module.read_accounts_from_excel(out.getvalue(), limit_per_sheet=1)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["account"], "u1@gmail.com")
        self.assertEqual(rows[1]["account"], "u3@gmail.com")

    # Проверяет, что лист "Аккаунты" тоже читается и помечается как привязка игры.
    def test_read_accounts_from_excel_includes_akkaunty_sheet(self):
        wb = Workbook()
        ws_mail = wb.active
        ws_mail.title = "Почты"
        ws_mail.append(["Логин", "Пароль"])
        ws_mail.append(["u1@gmail.com", "p1"])
        ws_bind = wb.create_sheet("Аккаунты")
        ws_bind.append(["Почта", "Игры"])
        ws_bind.append(["u1@gmail.com", "EA FC 26"])
        out = BytesIO()
        wb.save(out)

        rows = app_module.read_accounts_from_excel(out.getvalue())
        self.assertEqual(len(rows), 2)
        binding_row = [r for r in rows if r.get("_sheet_name") == "Аккаунты"][0]
        self.assertEqual(binding_row["account"], "u1@gmail.com")
        self.assertEqual(binding_row["game"], "EA FC 26")
        self.assertEqual(binding_row["_import_kind"], "account_game_binding")

    # Проверяет, что листы подписок читаются в том же прогоне и получают свои типы строк.
    def test_read_accounts_from_excel_includes_subscription_sheets(self):
        wb = Workbook()
        ws_mail = wb.active
        ws_mail.title = "Почты"
        ws_mail.append(["Логин", "Пароль"])
        ws_mail.append(["u1@gmail.com", "p1"])
        ws_plus = wb.create_sheet("ПЛЮС")
        ws_plus.append(["Аккаунт", "Подписка", "Срок"])
        ws_plus.append(["u1@gmail.com", "ПЛЮС DELUXE", "05.02.2027"])
        ws_ea = wb.create_sheet("EA PLAY")
        ws_ea.append(["Аккаунт", "Подписка", "Срок"])
        ws_ea.append(["u1@gmail.com", "ПОДПИСКА EA PLAY до 26.02.2026", "26.02.2026"])
        ws_bind = wb.create_sheet("Аккаунты")
        ws_bind.append(["Почта", "Игры"])
        ws_bind.append(["u1@gmail.com", "EA FC 26"])
        out = BytesIO()
        wb.save(out)

        rows = app_module.read_accounts_from_excel(out.getvalue())
        self.assertEqual(len(rows), 4)
        self.assertEqual(rows[0]["_import_kind"], "account_credentials")
        self.assertEqual(rows[1]["_import_kind"], "subscription_term_plus")
        self.assertEqual(rows[1]["subscription"], "ПЛЮС DELUXE")
        self.assertEqual(rows[2]["_import_kind"], "subscription_term_ea_play")
        self.assertEqual(rows[2]["subscription"], "ПОДПИСКА EA PLAY до 26.02.2026")
        self.assertEqual(rows[3]["_import_kind"], "account_game_binding")

    # Проверяет валидацию листа "Аккаунты": проверяем игру, а отсутствие аккаунта не считаем ошибкой.
    def test_accounts_import_validate_checks_game_for_bindings_and_allows_missing_account(self):
        wb = Workbook()
        ws_bind = wb.active
        ws_bind.title = "Аккаунты"
        ws_bind.append(["Почта", "Игры"])
        ws_bind.append(["ok@gmail.com", "EA FC 26"])  # ok
        ws_bind.append(["missing@gmail.com", "EA FC 26"])  # account will be auto-created
        ws_bind.append(["ok@gmail.com", "Unknown Game"])  # game not found
        out = BytesIO()
        wb.save(out)
        content = out.getvalue()

        script = [
            {"one": (55,)},  # row1 game exists
            {"one": (55,)},  # row2 game exists
            {"one": None},   # row3 game missing
        ]
        with (
            patch.object(app_module, "ensure_analytics_schema", return_value=None),
            patch.object(app_module.psycopg, "connect", return_value=_ScriptedConnCtx(script)),
            patch.object(app_module, "JWT_SECRET", "test-secret"),
            patch.object(app_module, "JWT_ALG", "HS256"),
        ):
            with self._client() as client:
                res = client.post(
                    "/accounts/import/validate",
                    headers=self._auth_headers(role="admin"),
                    files={"file": ("accounts.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            self.assertEqual(res.status_code, 200)
            body = res.json()
            self.assertEqual(body["total"], 3)
            warnings = body.get("warnings", [])
            self.assertFalse(any(w.get("field") == "Почта" and "Аккаунт не найден" in str(w.get("message") or "") for w in warnings))
            self.assertTrue(any(w.get("field") == "Игры" and "Игра не найдена" in str(w.get("message") or "") for w in warnings))

    # Проверяет, что slots-парсер подтягивает товар из листа "Аккаунты" по почте.
    def test_read_slots_from_excel_resolves_game_from_akkaunty_sheet(self):
        wb = Workbook()
        ws_users = wb.active
        ws_users.title = "Пользователи"
        ws_users.append(["Почта", "Статус", "Платформа", "Пользователь", "Откуда", "Покупка"])
        ws_users.append(["user1@gmail.com", "ЗАНЯТ", "PS5(1)", "Buyer", "ASAT TG", "01.04.2026"])
        ws_accounts = wb.create_sheet("Аккаунты")
        ws_accounts.append(["Почта", "Игры"])
        ws_accounts.append(["user1@gmail.com", "EA FC 26"])
        out = BytesIO()
        wb.save(out)

        rows = app_module.read_slots_from_excel(out.getvalue())
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["account"], "user1@gmail.com")
        self.assertEqual(rows[0]["slot"], "PS5(1)")
        self.assertEqual(rows[0]["game"], "EA FC 26")

    # Проверяет предупреждение, когда товар не найден ни в строке, ни в листах Аккаунты/ПЛЮС/EA PLAY.
    def test_validate_slot_import_rows_warns_when_game_not_found_from_related_sheets(self):
        rows = [{
            "account": "user1@gmail.com",
            "status": "ЗАНЯТ",
            "slot": "PS5(1)",
            "game": "",
            "customer": "Buyer",
            "source": "ASAT TG",
            "date": "01.04.2026",
        }]
        script = [
            {"all": [("play_ps5",)]},   # slot_types
            {"one": (1,)},              # account exists
            {"one": (1,)},              # source exists
        ]
        with _ScriptedConnCtx(script) as conn:
            errors, warnings, total = app_module.validate_slot_import_rows(conn, rows)

        self.assertEqual(errors, [])
        self.assertEqual(total, 1)
        self.assertTrue(any("Товар не найден (нет совпадения в листах Аккаунты/ПЛЮС/EA PLAY)" in str(w.get("message") or "") for w in warnings))

    # Проверяет предупреждение, если по аккаунту найдено несколько кандидатов товара/подписки.
    def test_validate_slot_import_rows_warns_for_multiple_game_candidates(self):
        rows = [{
            "account": "user1@gmail.com",
            "status": "ЗАНЯТ",
            "slot": "PS5(1)",
            "game": "",
            "_game_candidates": ["EA FC 26", "EA PLAY"],
            "customer": "Buyer",
            "source": "ASAT TG",
            "date": "01.04.2026",
        }]
        script = [
            {"all": [("play_ps5",)]},   # slot_types
            {"one": (1,)},              # account exists
            {"one": (1,)},              # source exists
        ]
        with _ScriptedConnCtx(script) as conn:
            errors, warnings, total = app_module.validate_slot_import_rows(conn, rows)

        self.assertEqual(errors, [])
        self.assertEqual(total, 1)
        self.assertTrue(any("Найдено несколько товаров/подписок для аккаунта" in str(w.get("message") or "") for w in warnings))

    # Проверяет, что для "будущего" аккаунта из этого же файла не показываем предупреждение "Аккаунт не найден".
    def test_validate_slot_import_rows_skips_missing_account_warning_for_staged_account(self):
        rows = [{
            "account": "future@gmail.com",
            "_account_declared_in_file": True,
            "status": "ЗАНЯТ",
            "slot": "PS5(1)",
            "game": "EA FC 26",
            "customer": "Buyer",
            "source": "ASAT TG",
            "date": "01.04.2026",
        }]
        script = [
            {"all": [("play_ps5",)]},   # slot_types
            {"one": None},              # account missing in DB
            {"one": (77,)},             # product exists
            {"one": (1,)},              # source exists
        ]
        with _ScriptedConnCtx(script) as conn:
            errors, warnings, total = app_module.validate_slot_import_rows(conn, rows)

        self.assertEqual(errors, [])
        self.assertEqual(total, 1)
        self.assertFalse(any(w.get("field") == "Аккаунт" and "Аккаунт не найден" in str(w.get("message") or "") for w in warnings))

    # Проверяет валидацию сроков подписок: аккаунт должен существовать, дата обязательна и парсится.
    def test_accounts_import_validate_subscription_terms_requires_account_and_date(self):
        wb = Workbook()
        ws_plus = wb.active
        ws_plus.title = "ПЛЮС"
        ws_plus.append(["Аккаунт", "Подписка", "Срок"])
        ws_plus.append(["ok@gmail.com", "ПЛЮС DELUXE", "05.02.2027"])  # ok
        ws_plus.append(["missing@gmail.com", "ПЛЮС DELUXE", "05.02.2027"])  # account missing
        ws_plus.append(["ok@gmail.com", "ПЛЮС DELUXE", ""])  # empty date
        ws_plus.append(["ok@gmail.com", "ПЛЮС DELUXE", "abc"])  # invalid date
        out = BytesIO()
        wb.save(out)
        content = out.getvalue()

        script = [
            {"one": (1,)},    # row1 account exists
            {"one": None},    # row2 account missing
            {"one": (1,)},    # row3 account exists
            {"one": (1,)},    # row4 account exists
        ]
        with (
            patch.object(app_module, "ensure_analytics_schema", return_value=None),
            patch.object(app_module.psycopg, "connect", return_value=_ScriptedConnCtx(script)),
            patch.object(app_module, "JWT_SECRET", "test-secret"),
            patch.object(app_module, "JWT_ALG", "HS256"),
        ):
            with self._client() as client:
                res = client.post(
                    "/accounts/import/validate",
                    headers=self._auth_headers(role="admin"),
                    files={"file": ("accounts.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            self.assertEqual(res.status_code, 200)
            body = res.json()
            self.assertEqual(body["total"], 4)
            warnings = body.get("warnings", [])
            self.assertTrue(any(w.get("field") == "Аккаунт" and "Аккаунт не найден" in str(w.get("message") or "") for w in warnings))
            self.assertTrue(any(w.get("field") == "Срок" and "обязательна" in str(w.get("message") or "") for w in warnings))
            self.assertTrue(any(w.get("field") == "Срок" and "Не удалось распознать" in str(w.get("message") or "") for w in warnings))

    # Проверяет, что валидация сроков учитывает аккаунты, которые будут созданы из листов Почты* этого же файла.
    def test_accounts_import_validate_subscription_terms_accepts_accounts_staged_in_same_file(self):
        wb = Workbook()
        ws_mail = wb.active
        ws_mail.title = "Почты"
        ws_mail.append(["Логин", "Пароль"])
        ws_mail.append(["newacc@gmail.com", "pwd"])
        ws_plus = wb.create_sheet("ПЛЮС")
        ws_plus.append(["Аккаунт", "Подписка", "Срок"])
        ws_plus.append(["newacc@gmail.com", "ПЛЮС DELUXE", "05.02.2027"])
        out = BytesIO()
        wb.save(out)
        content = out.getvalue()

        # Для строки из ПЛЮС в БД аккаунт пока отсутствует, но он есть в этом же файле на листе Почты.
        script = [
            {"one": None},
        ]
        with (
            patch.object(app_module, "ensure_analytics_schema", return_value=None),
            patch.object(app_module.psycopg, "connect", return_value=_ScriptedConnCtx(script)),
            patch.object(app_module, "JWT_SECRET", "test-secret"),
            patch.object(app_module, "JWT_ALG", "HS256"),
        ):
            with self._client() as client:
                res = client.post(
                    "/accounts/import/validate",
                    headers=self._auth_headers(role="admin"),
                    files={"file": ("accounts.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            self.assertEqual(res.status_code, 200)
            body = res.json()
            warnings = body.get("warnings", [])
            self.assertFalse(any(w.get("field") == "Аккаунт" and "Аккаунт не найден" in str(w.get("message") or "") for w in warnings))

    # Проверяет конфликт: один и тот же занятый слот указан повторно, когда соседний слот свободен.
    def test_accounts_import_slots_check_warns_for_duplicate_busy_slot_with_free_alternative(self):
        wb = Workbook()
        ws_users = wb.active
        ws_users.title = "Пользователи"
        ws_users.append(["Почта", "Статус", "Платформа", "Пользователь", "Откуда", "Покупка"])
        ws_users.append(["winner@example.com", "ЗАНЯТ", "PS5(1)", "Buyer1", "ASAT TG", "17.05.2026"])
        ws_users.append(["winner@example.com", "СВОБОДЕН", "PS5(2)", "", "", ""])
        ws_users.append(["winner@example.com", "ЗАНЯТ", "PS5(1)", "Buyer2", "VK", "21.05.2026"])
        out = BytesIO()
        wb.save(out)
        content = out.getvalue()

        with (
            patch.object(app_module, "ensure_analytics_schema", return_value=None),
            patch.object(app_module, "JWT_SECRET", "test-secret"),
            patch.object(app_module, "JWT_ALG", "HS256"),
        ):
            with self._client() as client:
                res = client.post(
                    "/accounts/import/slots-check",
                    headers=self._auth_headers(role="admin"),
                    files={"file": ("accounts.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            self.assertEqual(res.status_code, 200)
            body = res.json()
            self.assertFalse(body.get("ok"))
            warnings = body.get("warnings", [])
            self.assertTrue(any("Конфликт слотов" in str(w.get("message") or "") for w in warnings))
            self.assertTrue(any("PS5(1)" in str(w.get("message") or "") for w in warnings))
            self.assertTrue(any("PS5(2)" in str(w.get("message") or "") for w in warnings))

    # Проверяет, что исторический повтор без свободной альтернативы не помечается как конфликт.
    def test_accounts_import_slots_check_does_not_warn_when_no_free_alternative(self):
        wb = Workbook()
        ws_users = wb.active
        ws_users.title = "Пользователи"
        ws_users.append(["Почта", "Статус", "Платформа", "Пользователь", "Откуда", "Покупка"])
        ws_users.append(["winner@example.com", "ЗАНЯТ", "PS5(1)", "Buyer1", "ASAT TG", "17.05.2026"])
        ws_users.append(["winner@example.com", "ЗАНЯТ", "PS5(1)", "Buyer2", "VK", "21.05.2026"])
        ws_users.append(["winner@example.com", "ЗАНЯТ", "PS5(2)", "Buyer3", "VK", "22.05.2026"])
        out = BytesIO()
        wb.save(out)
        content = out.getvalue()

        with (
            patch.object(app_module, "ensure_analytics_schema", return_value=None),
            patch.object(app_module, "JWT_SECRET", "test-secret"),
            patch.object(app_module, "JWT_ALG", "HS256"),
        ):
            with self._client() as client:
                res = client.post(
                    "/accounts/import/slots-check",
                    headers=self._auth_headers(role="admin"),
                    files={"file": ("accounts.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            self.assertEqual(res.status_code, 200)
            body = res.json()
            warnings = body.get("warnings", [])
            self.assertFalse(any("Конфликт слотов" in str(w.get("message") or "") for w in warnings))

    # Проверяет XLSX-выгрузку истории, отметку дубля и зеленую строку свободного слота.
    def test_slots_export_contains_history_duplicates_and_free_rows(self):
        dt_1 = datetime(2026, 5, 17, tzinfo=timezone.utc)
        dt_2 = datetime(2026, 5, 18, tzinfo=timezone.utc)
        dt_3 = datetime(2026, 5, 19, tzinfo=timezone.utc)
        script = [{
            "all": [
                ("winner@example.com", "play_ps5", 2, 1, "Game A", "Buyer 1", dt_1, dt_1, None),
                ("winner@example.com", "play_ps5", 2, 2, "Game B", "Buyer 2", dt_2, dt_2, None),
                ("winner@example.com", "play_ps5", 2, 3, "Game C", "Buyer 3", dt_3, dt_3, None),
                ("winner@example.com", "activate_ps5", 1, None, None, None, None, None, None),
                ("winner@example.com", "activate_ps4", 1, 4, "Game D", "Buyer 4", dt_1, dt_1, dt_2),
                ("winner@example.com", "activate_ps4", 1, 5, "Game E", "Buyer 5", dt_2, dt_2, dt_3),
            ],
        }]
        with (
            patch.object(app_module, "ensure_analytics_schema", return_value=None),
            patch.object(app_module.psycopg, "connect", return_value=_ScriptedConnCtx(script)),
            patch.object(app_module, "JWT_SECRET", "test-secret"),
            patch.object(app_module, "JWT_ALG", "HS256"),
        ):
            with self._client() as client:
                res = client.get("/accounts/slots/export", headers=self._auth_headers(role="admin"))

        self.assertEqual(res.status_code, 200)
        wb = load_workbook(BytesIO(res.content))
        ws = wb["Слоты"]
        self.assertEqual([cell.value for cell in ws[1]], [
            "Аккаунт", "Дубль", "Товар", "Статус", "Пользователь", "Дата покупки (сделки)", "Тип слота",
        ])
        values = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertTrue(any(row[1] == "Да" and row[2] == "Game C" for row in values))
        self.assertTrue(any(row[1] is None and row[2] == "Game E" for row in values))
        free_row_idx = next(idx for idx, row in enumerate(values, start=2) if row[3] == "Свободен")
        self.assertEqual(ws.cell(free_row_idx, 1).fill.fgColor.rgb, "00C6EFCE")

    # Проверяет поиск сделок по связке "дата + ник" в отдельном файле проверки.
    def test_accounts_import_deals_check_warns_when_deal_not_found(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Проверка"
        ws.append(["Номер заказа", "Источник", "Мессенджер", "Дата", "Ник"])
        ws.append(["51281052034", "SPS", "TG ASAT", "24.11.2025", "AndiLino"])
        ws.append(["51252120576", "SPS", "TG ASAT", "24.11.2025", "Дмитрий"])
        out = BytesIO()
        wb.save(out)
        content = out.getvalue()

        # Пакетная сверка: в БД найдена только первая связка.
        script = [
            {"all": [("andilino", date(2025, 11, 24))]},
        ]
        with (
            patch.object(app_module, "ensure_analytics_schema", return_value=None),
            patch.object(app_module.psycopg, "connect", return_value=_ScriptedConnCtx(script)),
            patch.object(app_module, "JWT_SECRET", "test-secret"),
            patch.object(app_module, "JWT_ALG", "HS256"),
        ):
            with self._client() as client:
                res = client.post(
                    "/accounts/import/deals-check",
                    headers=self._auth_headers(role="admin"),
                    files={"file": ("deals-check.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            self.assertEqual(res.status_code, 200)
            body = res.json()
            self.assertFalse(body.get("ok"))
            warnings = body.get("warnings", [])
            self.assertTrue(any(w.get("field") == "Сделка" for w in warnings))
            self.assertTrue(any("дата + ник" in str(w.get("message") or "") for w in warnings))

    # Проверяет, что при полном совпадении связок ошибок и предупреждений нет.
    def test_accounts_import_deals_check_passes_when_all_deals_found(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Проверка"
        ws.append(["Номер заказа", "Источник", "Мессенджер", "Дата", "Ник"])
        ws.append(["51281052034", "SPS", "TG ASAT", "24.11.2025", "AndiLino"])
        out = BytesIO()
        wb.save(out)
        content = out.getvalue()

        script = [{"all": [("andilino", date(2025, 11, 24))]}]
        with (
            patch.object(app_module, "ensure_analytics_schema", return_value=None),
            patch.object(app_module.psycopg, "connect", return_value=_ScriptedConnCtx(script)),
            patch.object(app_module, "JWT_SECRET", "test-secret"),
            patch.object(app_module, "JWT_ALG", "HS256"),
        ):
            with self._client() as client:
                res = client.post(
                    "/accounts/import/deals-check",
                    headers=self._auth_headers(role="admin"),
                    files={"file": ("deals-check.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            self.assertEqual(res.status_code, 200)
            body = res.json()
            self.assertTrue(body.get("ok"))
            self.assertEqual(body.get("warnings", []), [])

    # Проверяет заливку номера заявки в найденную сделку.
    def test_accounts_import_deals_fill_updates_found_deals(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Проверка"
        ws.append(["Номер заявки", "Дата", "Ник"])
        ws.append(["REQ-100", "24.11.2025", "AndiLino"])
        out = BytesIO()
        wb.save(out)
        content = out.getvalue()

        script = [
            {"all": [("andilino", date(2025, 11, 24))]},
            {"rowcount": 1},
        ]
        with (
            patch.object(app_module, "ensure_analytics_schema", return_value=None),
            patch.object(app_module.psycopg, "connect", return_value=_ScriptedConnCtx(script)),
            patch.object(app_module, "JWT_SECRET", "test-secret"),
            patch.object(app_module, "JWT_ALG", "HS256"),
        ):
            with self._client() as client:
                res = client.post(
                    "/accounts/import/deals-fill",
                    headers=self._auth_headers(role="admin"),
                    files={"file": ("deals-fill.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            self.assertEqual(res.status_code, 200)
            body = res.json()
            self.assertTrue(body.get("ok"))
            self.assertEqual(body.get("updated"), 1)
            self.assertEqual(body.get("skipped"), 0)

    # Проверяет, что номер заявки нормализуется и пустой номер пропускается.
    def test_accounts_import_deals_fill_skips_rows_without_order_number(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Проверка"
        ws.append(["Номер заявки", "Дата", "Ник"])
        ws.append(["   ", "24.11.2025", "AndiLino"])
        ws.append(["  REQ   200  ", "24.11.2025", "AndiLino"])
        out = BytesIO()
        wb.save(out)
        content = out.getvalue()

        script = [
            {"all": [("andilino", date(2025, 11, 24))]},
            {"rowcount": 1},
        ]
        with (
            patch.object(app_module, "ensure_analytics_schema", return_value=None),
            patch.object(app_module.psycopg, "connect", return_value=_ScriptedConnCtx(script)),
            patch.object(app_module, "JWT_SECRET", "test-secret"),
            patch.object(app_module, "JWT_ALG", "HS256"),
        ):
            with self._client() as client:
                res = client.post(
                    "/accounts/import/deals-fill",
                    headers=self._auth_headers(role="admin"),
                    files={"file": ("deals-fill.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            self.assertEqual(res.status_code, 200)
            body = res.json()
            self.assertEqual(body.get("updated"), 1)
            self.assertEqual(body.get("skipped"), 1)


if __name__ == "__main__":
    unittest.main()
