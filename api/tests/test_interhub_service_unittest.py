import json
import unittest
from io import BytesIO
from unittest.mock import patch

from fastapi import HTTPException
from openpyxl import load_workbook

from domains.interhub_api import interhub_status_check_interval
from domains.interhub_price_cache import build_interhub_prices_xlsx, collect_price_targets
from domains.interhub_service import build_interhub_service


class _Response:
    def __init__(self, payload):
        self._raw = json.dumps(payload).encode("utf-8")

    def read(self):
        # Имитируем байтовый ответ urllib без обращения к реальному InterHub.
        return self._raw

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        return False


class InterHubServiceTests(unittest.TestCase):
    def test_status_polling_uses_documented_backoff(self):
        # Первый статус проверяем через минуту, затем три раза через пять и дальше через тридцать минут.
        self.assertEqual(interhub_status_check_interval(0), "1 minute")
        self.assertEqual(interhub_status_check_interval(1), "5 minutes")
        self.assertEqual(interhub_status_check_interval(3), "5 minutes")
        self.assertEqual(interhub_status_check_interval(4), "30 minutes")

    def test_catalog_normalizes_nested_service_and_dynamic_fields(self):
        # Преобразуем типовой ответ провайдера в контракт, удобный для динамической формы.
        payload = {
            "success": True,
            "data": [
                {
                    "service_id": 7,
                    "name": "Mobile top up",
                    "category_name": "Mobile",
                    "type": "TOP_UP_FIXED",
                    "min_amount": "10.50",
                    "max_amount": 100,
                    "fields": [
                        {
                            "name": "nominal",
                            "type": "LIST",
                            "required": True,
                            "value_list": [{"id": 15, "title": "15"}],
                        }
                    ],
                }
            ],
        }
        service = build_interhub_service(
            HTTPException=HTTPException,
            interhub_api_url="https://api.interhub.ae",
            interhub_token="test-token",
            timeout_sec=20,
            ssl_verify=False,
            ca_cert_path="",
            calculate_path="/api/agent/payment/check/calculate",
            check_path="/api/agent/payment/check",
            deposit_path="/api/agent/deposit",
        )
        with patch("domains.interhub_service.urllib.request.urlopen", return_value=_Response(payload)) as urlopen_mock:
            items = service.get_services()

        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["service_id"], 7)
        self.assertEqual(items[0]["min_amount"], 10.5)
        self.assertTrue(items[0]["fields"][0]["required"])
        request = urlopen_mock.call_args.args[0]
        self.assertEqual(request.get_header("Token"), "test-token")
        self.assertTrue(request.full_url.endswith("/api/agent/service/list"))

    def test_catalog_rejects_empty_token_before_network_call(self):
        # Блокируем запрос, если секрет не задан в серверном окружении.
        service = build_interhub_service(
            HTTPException=HTTPException,
            interhub_api_url="https://api.interhub.ae",
            interhub_token="",
            timeout_sec=20,
            ssl_verify=True,
            ca_cert_path="",
            calculate_path="/api/agent/payment/check/calculate",
            check_path="/api/agent/payment/check",
            deposit_path="/api/agent/deposit",
        )
        with self.assertRaises(HTTPException) as error:
            service.get_services()
        self.assertEqual(error.exception.status_code, 500)

    def test_pay_preserves_gift_code_and_check_status(self):
        # Сохраняем код ваучера из pay и статус processing без потери полей провайдера.
        service = build_interhub_service(
            HTTPException=HTTPException,
            interhub_api_url="https://api.interhub.ae",
            interhub_token="test-token",
            timeout_sec=20,
            ssl_verify=False,
            ca_cert_path="",
            calculate_path="/api/agent/payment/check/calculate",
            check_path="/api/agent/payment/check",
            deposit_path="/api/agent/deposit",
        )
        with patch(
            "domains.interhub_service.urllib.request.urlopen",
            return_value=_Response({"success": True, "status": 0, "params": {"gift_code": "TESTGIFTCODE"}}),
        ) as urlopen_mock:
            result = service.pay({"agent_transaction_id": "test-1"})

        self.assertEqual(result["params"]["gift_code"], "TESTGIFTCODE")
        request = urlopen_mock.call_args.args[0]
        self.assertTrue(request.full_url.endswith("/api/agent/payment/pay"))

    def test_price_targets_keep_only_active_voucher_and_top_up_fixed_nominals(self):
        # В массовый calculate берём только поддерживаемые типы и активные номиналы из каталога.
        targets = collect_price_targets([
            {
                "service_id": 11, "title": "Voucher", "category": "Games", "type": "VOUCHER",
                "fields": [{"name": "nominal", "value_list": [{"id": 101, "title": "TRY 100", "active": True}, {"id": 102, "title": "TRY 200", "active": False}]}],
            },
            {
                "service_id": 12, "title": "Steam", "category": "Games", "type": "TOP_UP",
                "fields": [{"name": "nominal", "value_list": [{"id": 201, "title": "USD 10", "active": True}]}],
            },
        ])
        self.assertEqual(targets, [{
            "service_id": 11, "service_title": "Voucher", "category": "Games", "service_type": "VOUCHER",
            "nominal_id": 101, "nominal_title": "TRY 100",
        }])

    def test_price_export_has_prices_and_calculate_errors(self):
        # Excel разделяет полезные цены и ошибки, которые нужно отправить поставщику.
        content = build_interhub_prices_xlsx(
            [{"service_id": 11, "service_title": "Voucher", "category": "Games", "service_type": "VOUCHER", "nominal_id": 101, "nominal_title": "TRY 100", "fixed_amount": 117.47, "provider_response": {"success": True, "fixed_amount": 117.47}}],
            [{"service_id": 12, "service_title": "India", "service_type": "VOUCHER", "nominal_id": 202, "nominal_title": "INR 2500", "provider_status": -142, "provider_message": "The service not active", "provider_response": {"success": False, "status": -142}}],
        )
        workbook = load_workbook(BytesIO(content))
        self.assertEqual(workbook.sheetnames, ["Закупочные цены", "Ошибки calculate"])
        self.assertEqual(workbook["Закупочные цены"]["G2"].value, 117.47)
        self.assertEqual(workbook["Закупочные цены"]["I2"].value, '{"fixed_amount": 117.47, "success": true}')
        self.assertEqual(workbook["Ошибки calculate"]["G2"].value, "The service not active")


if __name__ == "__main__":
    unittest.main()
