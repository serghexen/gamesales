import sys
import unittest
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


SCRIPTS_DIR = Path(__file__).resolve().parents[1] / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from audit_interhub_stock import audit_targets, build_report, resolve_runtime_paths


class _FakeInterHubService:
    def __init__(self):
        # Хранит внешние вызовы, чтобы тест доказал отсутствие pay.
        self.calculate_calls = []
        self.check_calls = []
        self.pay_calls = []

    def calculate(self, payload):
        # Возвращает цену только для первого номинала и ошибку для второго.
        self.calculate_calls.append(payload)
        if payload["params"]["nominal"] == 101:
            return {
                "success": True,
                "status": 0,
                "message": "",
                "fixed_amount": 117.47,
                "amount_in_currency": 100,
                "raw": {"success": True, "fixed_amount": 117.47},
            }
        return {"success": False, "status": -142, "message": "The service not active", "raw": {"status": -142}}

    def check(self, payload):
        # Подтверждает доступность и возвращает ID только для проверенного номинала.
        self.check_calls.append(payload)
        return {"success": True, "status": 0, "message": "Available", "transaction_id": "provider-check-1", "raw": {"success": True}}

    def pay(self, payload):
        # Делает случайный вызов оплаты видимым как немедленную ошибку теста.
        self.pay_calls.append(payload)
        raise AssertionError("pay не должен вызываться аудитом")


class InterHubStockAuditScriptTests(unittest.TestCase):
    def setUp(self):
        # Готовит два номинала с разными результатами calculate.
        self.targets = [
            {"service_id": 11, "service_title": "Voucher", "category": "Games", "service_type": "VOUCHER", "nominal_id": 101, "nominal_title": "TRY 100"},
            {"service_id": 12, "service_title": "Inactive", "category": "Games", "service_type": "VOUCHER", "nominal_id": 202, "nominal_title": "TRY 200"},
        ]

    def test_audit_calls_check_only_after_successful_calculate_and_never_pays(self):
        # Проверяет безопасную последовательность calculate → check без pay.
        service = _FakeInterHubService()
        rows = audit_targets(service, self.targets, delay_ms=0, run_id="test")

        self.assertEqual(len(service.calculate_calls), 2)
        self.assertEqual(len(service.check_calls), 1)
        self.assertEqual(service.check_calls[0]["amount"], 117.47)
        self.assertEqual(service.check_calls[0]["params"], {"nominal": 101})
        self.assertEqual(service.pay_calls, [])
        self.assertEqual(rows[0]["availability"], "Доступно")
        self.assertEqual(rows[1]["availability"], "Не проверено")

    def test_runtime_paths_support_script_copied_to_container_tmp(self):
        # При запуске из /tmp использует рабочий каталог контейнера вместо отсутствующего корня репозитория.
        root_dir, api_dir = resolve_runtime_paths(Path("/tmp/audit_interhub_stock.py"), Path("/app"))

        self.assertEqual(root_dir, Path("/app"))
        self.assertEqual(api_dir, Path("/app"))

    def test_report_contains_summary_and_filterable_audit_rows(self):
        # Проверяет листы и сохранение локальной даты, переданной вместе с timezone.
        service = _FakeInterHubService()
        rows = audit_targets(service, self.targets, delay_ms=0, run_id="test")
        generated_at = datetime(2026, 8, 12, 20, 30, tzinfo=timezone(timedelta(hours=3)))
        workbook = build_report(rows, generated_at=generated_at)
        buffer = BytesIO()
        workbook.save(buffer)
        loaded = load_workbook(BytesIO(buffer.getvalue()))

        self.assertEqual(loaded.sheetnames, ["Сводка", "Проверка остатков", "Ответы API"])
        self.assertEqual(loaded["Сводка"]["B3"].value, 2)
        self.assertEqual(loaded["Сводка"]["B4"].value, 1)
        self.assertIsNone(loaded["Сводка"]["B2"].value.tzinfo)
        report = loaded["Проверка остатков"]
        self.assertEqual(report["M2"].value, "Доступно")
        self.assertEqual(report["M3"].value, "Не проверено")
        self.assertEqual(report["G2"].value, 117.47)
        self.assertEqual(report.freeze_panes, "A2")
        self.assertEqual(report.auto_filter.ref, report.dimensions)
        responses = loaded["Ответы API"]
        self.assertEqual(responses["F2"].value, '{"fixed_amount": 117.47, "success": true}')
        self.assertEqual(responses["G2"].value, '{"success": true}')


if __name__ == "__main__":
    unittest.main()
